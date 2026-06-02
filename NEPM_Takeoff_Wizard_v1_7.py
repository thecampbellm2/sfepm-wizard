"""
NEPM Takeoff Wizard  v1.7
==================
Double-click to run. Requires Python 3.8+ with openpyxl and Pillow.
If missing, run:  pip install openpyxl pillow

Changes in v1.7:
  - Spreadsheet logo enlarged to 5.21 cm × 5.21 cm to fill the header row
  - Row 1 height increased to match the larger logo

Changes in v1.6:
  - Fixed: selected radio button now shows a visible white dot
    (was invisible against the dark indicator ring in v1.5)

Changes in v1.5:
  - Monochrome (black/white/grey) colour palette throughout wizard and spreadsheet
  - Spreadsheet title block redesigned: tall mega-header with large logo + job name
  - Wizard logo enlarged to fill header band

Changes in v1.4:
  - Rebranded: Sydney Fitout Estimation & Project Management →
    National Estimation & Project Management (NEPM)
  - NEPM circular badge logo embedded in wizard header and spreadsheet title block
  - All SFEPM references updated to NEPM throughout

Changes in v1.3:
  - Save dialog pre-fills to H:\\My Drive\\Jobs\\[YYYY]\\[MM MMM]\\[Job Name]\\
    Folders are created automatically if they don't exist.
    Falls back to standard folder picker if H: drive is unavailable.
    Invalid Windows filename characters stripped from job name in path.

Changes in v1.2:
  - Brickwork tab: Opening Count and Opening Width (LM) columns added
    next to Deductions M2 for lintel reference
  - Blockwork tab: Same opening columns added
  - George Summary: Total Openings and Total Opening Width (LM) per
    brick/block type, summed from takeoff sheets

Changes in v1.1:
  - Auto-updater: wizard checks GitHub for updates on launch
  - Forced update with changelog shown after restart

Changes in v1.0 (internal v5):
  - Font changed to Aptos throughout spreadsheet
  - Total ($) now calculates as Wastage qty × Rate (rate per brick/block)
  - Dynamic brick/block calculator on Rates & Assumptions sheet
  - 10 brick/block type slots in wizard (up from 5)
"""

import sys
import os
import tkinter as tk
from tkinter import messagebox

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    import subprocess
    root = tk.Tk(); root.withdraw()
    ans = messagebox.askyesno(
        "Missing dependency",
        "openpyxl is required but not installed.\n\nInstall it now? (requires internet)"
    )
    if ans:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    else:
        sys.exit(1)



# ══════════════════════════════════════════════════════════════════════════════
#  VERSION & USER LOOKUP
# ══════════════════════════════════════════════════════════════════════════════
VERSION = "v1.7"

# ── Auto-updater config ───────────────────────────────────────────────────────
# Set these to your GitHub repo's raw file URLs after first publish.
# Replace YOUR_USERNAME and YOUR_REPO with your actual GitHub details.
GITHUB_VERSION_URL   = "https://raw.githubusercontent.com/thecampbellm2/sfepm-wizard/main/version.txt"
GITHUB_EXE_URL       = "https://raw.githubusercontent.com/thecampbellm2/sfepm-wizard/main/NEPM_Takeoff_Wizard.exe"
GITHUB_CHANGELOG_URL = "https://raw.githubusercontent.com/thecampbellm2/sfepm-wizard/main/changelog.txt"
# ─────────────────────────────────────────────────────────────────────────────

# Map Windows usernames to display names for "Prepared by" field
# Add new users here: "windows_username": "Display Name"
USER_DISPLAY_NAMES = {
    "theca": "Mike Campbell",
    # "george": "Georgina Murison",   # ← uncomment and set correct username
}

def get_display_name():
    """Return display name for current Windows user, fallback to username."""
    import os
    try:
        username = os.environ.get("USERNAME", os.environ.get("USER", "")).lower()
        return USER_DISPLAY_NAMES.get(username, username.title())
    except Exception:
        return ""


# ══════════════════════════════════════════════════════════════════════════════
#  PERSISTENT CONFIG  (remembers save folder between sessions)
# ══════════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════════
#  EMBEDDED LOGO  (NEPM badge — base64-encoded JPEG, Pillow required)
# ══════════════════════════════════════════════════════════════════════════════
import base64
LOGO_B64 = "/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCAH0AfQDASIAAhEBAxEB/8QAHQABAAEEAwEAAAAAAAAAAAAAAAUDBAcIAgYJAf/EAF8QAAEDAwEDBgcIDgcGAwUJAAEAAgMEBREGBxIhCBMxQVFhFBUiMnGBkQkWN0JVdaGzIzM0NjhSYnJ0grGywdEYJENTc5LTFyU1VpSiVFelRGNktOFFZ4STo8LD0uT/xAAYAQEBAQEBAAAAAAAAAAAAAAAAAgMBBP/EADMRAQACAQEEBwgDAAMBAQAAAAABAhEDEiFBUSIxYXGxwfAEIzIzgZGh0bLh8RNCQ4I0/9oADAMBAAIRAxEAPwDTJERAREQEREBERAREQEREBERAREQEREBERAREQEREBEX0AnoGUHxFzEUh6GFchTyHsHrQUkVcUzutwX3wb8v6EFuindL2EXi7MozM6OPdL5HAcQ0dnf0Ls+ptC0FJaZaygnnbJA3ec2QhwcB09XAoMdornwYfj/Qvhpj1P+hBboqxp39RaVxMMg+Ln1oKaLkWOHS0j1LigIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIubI3v80etBwRXLKcDzznuCrNY1vmtAQWjYZHfFx6VVbTD4zvYq6IODYYx8XPpXMADoGERAREQEREF9ZLlPabjHW04a5zchzXdDgekKd1BrOe5W59FDRimbIMSOMm8SOwcBhdURAREQEREBfHMa7paD6l9RBSdTsPRkKm6ncPNIKuUQWLmOb5zSFxUgqb4Y3dWD3ILNFWfTvHFvlBUiCDgjCD4iIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIuTGOecNGUHFc44nv6BgdpVxHA1vF3lH6FVQUmQMb0+Ue9VURAREQEREBF9a1znBrQST1AK9gtFxmGW0r2t/Gf5I+lBYopTxTHH903Oji7Q1++4eoL5zNlj8+tqpv8OIN/agjEUnz1kZ5tFVS/nyhv7E8OtrfMszP1p3OQRiKT8ZUo6LRSevJTxnT9dpovUD/NBGIpPxjRHzrPTkdz3BPC7Q7zrS9ne2od/FBGIpP/AHFJ1V8J/VcP5p4DbZPtF2a0/iyxFv0oIxFJmyVjhmnfT1Q/91KD+1WVRS1NOcT08sf5zSAgooiICIiAvj2NcMOAK+ogtpKcjiw57iqJBBwRgq/Xx7GvGHDKCwRVpYHN4t8ofSqKAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICL61pccAZKuoYQzieLv2IKcUBdxfwHYrhoDRgDAX1EBERARFXpKSoq5Obp4XSO68DgPSepBQXKNj5HhkbHPcegNGSVJ+B2+i411Vz8g/sac5x6XLjJeJmsMVDDFRRn+7HlH0u6UBlnmY0PrZoaNh4/ZHeUfQ0cV937NTcGRTVzx1vPNs9nSo173PeXvc5zj0knJK4oJN16qWgtpI4KRvZFGAT6SVYz1NROczzySfnOJVJEBERAREQEREBERAREQfQSDkHBV5T3a4wDDKqQt/Fed4fSrJEEp4xpJ/u22wuJ+PCSw+nHQU8Ct1T9xV/NOP9nUjd/wC4cFFogu623VlIN6aBwZ1PbxafWFaK6oq+soz/AFedzW9bTxafUeCu/CrbWcKyl8GkP9tTjh62/wAkEUikai0ziIz0j2VkA+NF0j0t6Qo5AREQFTlha/j0HtVREFi9jmHDguKv3NDhhwyFazQlnEcW/sQUkREBERAREQEREBERAREQEREBERAREQEREBERAREQFzjY57sD1lIozIewdZV4xoa3DRgIPkbGsGB6yuSIgIiIC5wxSzyCOGN0jz0NaMlXlDbXSw+FVMgpqUf2jhxd3NHWqk9yZDEaa1xmnjPB0p+2P9J6vQEHPwKjoBvXKXnZuqmidx/Wd1KhV3Somj5iENpqfqiiGAfSetWBJJyeJRAREQEREBERAREQEREBERAREQEREBERAREQEREFSnnmp5RJBK6N462nCkvDqOu8m5Q83Kf/AGiEYP6zetRKIL6utk9PHz7C2opj0Sx8R6+xWKuKGtqaKTfgkLc+c08WuHeFf81Q3TjT7tHVn+yJ+xvP5J6j3IIhFUqIZaeV0U0bo3t6QQqaAiIgt54fjMHpCt1IKjPDveU3p7O1BaoiICIiAiIgIiICIiAiIgIiICIiAiIgIiIC5xRmR3d1lI2F7sD1lXjGhrQ0DggNaGjAGAvqIgIiqU0EtTM2GBhe93QAg4Na5zg1rS5xOAAOJUs2mprYwS17RNUkZZTA8G97/wCS+vlgtDDFTObNXEYfN0tj7m9/eol7nPeXvcXOJySTklBWrqyorZecnfnHBrRwa0dgCt0RAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREEpT3CKoibS3RpljHBkw+2R/zCt7jQS0m7IHNlp3/a5mea7+R7lZq9t1wkpN6J7RNTP+2Qu6D3jsPegskUjX0DBD4bQOMtKTxz50R7HfzUcgIiIKM8W95TfO/arVSCoVEWfLaOPWEFsiIgIiICIiAiIgIiICIiAiIgIiIC+tBc4AdJXxXdPHuNyfOP0IOcbAxuB61yREBEVSmglqZ2QQtLnvOAEHKjppquobBAwue76O8qQqamG3wOoqB+9I4YnqB1/kt7krJ4rfTuoKJ4dI7hUTjr/Jb3KJQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERBc2+smop+ciIIIw9juLXjsIV1XUcM1ObhbweZ/tYvjRH+I71GK4oKuaiqBNER2Oaehw6wUFuikrjSQvg8YUAPg7jiSPrid2Hu7Co1AREQW1TFunfaOHWqCkCARgqzmj5t+Oo9CCmiIgIiICIiAiIgIiICIiAiLlG0veGhBVpo94756B0K5XxoDWgDoC+oCIiD61rnuDWgucTgAdZUvO5topnU0RBrpW/Znj+zH4o7+1fKNrbZRivlANTKCKZh+KPxz/BRL3Oe8ve4uc45JPSSg+IiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiILq2Vr6KffDQ+N43ZIz0Pb2KrdaNkO5U0ri+km4xu62nrae8KwUhaauOPfpKvLqSbg78g9Tggj0VxcaSSiqnQSYOOLXDoc3qIVugLjIwPYWn1LkiCwcCCQekL4rmqjyN8dI6VbICIiAiIgIiICIiAiIgK7pmbrN49JVvCzfkA6utXqAiIgK/tFLHK99TVZFLTjek/KPU0elWcET55mQxt3nvIa0d6kLzLHDGy10zsxQnMjh8eTrPq6EFpcKuStqnTycM8GtHQ0dQCt0RAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREEvREXOh8AkI8JhBdTOPxh1s/kokggkEEEdIK5RSPilbLG4te05aR1FSN3YypgZdYGgCQ7s7R8ST+R6UEWiIgHiMFWUrNx5HV1K9VKpZvR5HSEFoiIgIiICIiAiIgIi5xN35A32oLimZux5PSeKqoiAiKvQ0z6urjp4/Oe7GewdZQX9u/qFvkuTvt0mY6Ydh+M71KJPE5Kv73UsmqhFB9z07ebiHaB0n1qwQFnfkhbJtLbVbrqGm1PJcWMt8EMkPgczYzl7nA5y056AsELbb3N774NZfotN+/ItNKImZzylnqTMRu5x4u5615H+hm6Tub9LVd7F7ZTufRCpqmPjdIBkNcAwcDjGc8M5WjVRDLT1ElPPG6KWJ5ZIxwwWuBwQR25Xr1U3GiprjSW+eoZHU1geadjumTcALgO8A5x2Z7FoZy6Nm3vV2hs1dbafctWoCXy7o8mKrHnju3h5fp3uxea0zW8cp8f7/T0ViLVmOMevX1du5NnJ02f7Qtklu1Tfpr22vqZZ2PFNVMZHhkjmjALD1DtWFeU5oSybOdq1TpjT76t9DHSwytNVIHv3ntyeIA4epbmch/8HWy/pFV9c5aucu78IOt+b6X9xa+09HVrEetzP2fpUtM9v8mB0RSGm7RWX/UNvsdvYX1dfUx00Le1z3Bo/au1rNpiI4uTMVjMtoOTFybdL682Zx6r1hLdopa2pk8DZSTtjbzLTu7xy05JcHeoBdr2qckzRFr2d3u6aUmvj7zRUjqimZPUskY8s8pzS0MBJLQQOPThbG2ijtWgtntPRtcIbZYraA554Yjij8px7zgn1qx2Q6zpdomzi1aqhiZGK6IiaEHIjka4te32g+pNaIvNo0+HqJ/BpTNYrN+P+zDylRd95QWjToTa9f8AT7IyylZUmej4cDBJ5bMegHd9LSuhKKW26xZd67NphtRyVdgOhdpmzSXUWo5rwysZcZaYClqWsZuNawjgWHj5R61ln+h9sn/8TqX/AK6P/TWlWk9pGvNJ2w2zTeq7raqJ0hlMFNOWMLyAC7HbwHsWaeSltT2i6j27WCz33WN3uNvnFRztPPUFzH4ge4ZHcQD6lviL2xXdu8IY5mlZmWbP6H2yf/xOpf8Aro/9NP6H2yf/AMTqX/ro/wDTWUNvdzuFm2M6sutqq5aOupbbLJBPE7D43AcCD2rzx/237XP/ADCv/wD1RWEWzaa+uP6bTXoxZbbfNKWvQ+1u/aWszqh1BQSxshNQ8Pkw6NjjkgDPFx6lK8mPQlk2jbVqbTGoH1bKGSlmlcaWQMfvMbkcSDw9Sx/qC83XUF4qLxe6+evuFSQZqiZ2895AAGT6AB6lmnkI/hB0XzfVfuLT2avC2/dPhLPXtxru3x4sobe+TTs60Rskv2qbNPfXV9BEx8IqKtj48mRrTkBgzwcetabr0x5Xn4Ourf0eL6+NeZyxrMze0dzaYjYie/yfQCTgDJK2g2Jckm7altlPfdeXGosVFO0PioKdgNW9h6C8uy2PPYQ49oC6jyJ9DUmstscVXcoGz0FjgNe+Nwy18ocGxA/rHe/VW8G2/XUWzjZndtVviZNPTRhlLC48JJnndYD3ZOT3Are8xpU2p65/z7zLGsTqX2Y4evsx9Sck7Y5DTiOS13SpcBjnJbjIHHv8nA+hdP19yNNK1lFLNoq/XC11wBLIa5wnp3HqGQA9vpy70LVDUe1XaNf7zJdrjrS+eEPfvNENbJFHH3MYwhrR6AtseRHtpvusZq3RGrq59wuFJT+E0NbKcyyxAgPY8/GIyCCeJGc9CmtJvE83bXik9jTnXmkb/ofU1Tp3UtA+jr6c8Wni17T0PY7oc09RH7VArf7l46GpL9soOrIoGi52CRr+cA8p1O9wa9h7gS13dg9q0BWVLTOYnrhpesRiY6pZn5JOzLTm1LW90s2pX17KaltxqYzSTCN2/wA4xvElp4YcVsbf+R5s6fY61tlrr7DczA/wR89Wx8Ylx5O8AwEtzjPFYo9zq+FO/wDzKfro1vDXXGioqmip6qoZFLXTGCmDv7SQMc/dHfuscfUvRq1jYrjjHnLHTtO1bPCfKHkXdaCrtVzqrZXwPp6ukmdDPE8YLHtJDgfQQtn+SvsA0LtM2ZSaj1HNeGVrbhLTAUtS1jNxrWEcCw8fKPWqXL62beJtVUu0K2U+7RXciCvDRwZUtHBx/PaPa09qy97n/wDATL881H7kaj2edultrriPzmPXcrXjZvXHVP6lq5yr9nGntmG0SjsGm31z6Sa2x1TjVyiR++6SRpwQ0cMNCxAtkfdC/hotvzHF9bKtbllpTM13858Za6sRFt3KPCBbQclLk86d2j6GrNUavkukUUlWYKBtJM2PeaweW45ac+UcD80rWagpKivrqehpInS1FRK2KJjelznHAA9ZXqzs305R6C2bWjTzXMZDaqFrZ5OgFwG9I8+lxcfWvRGK6c2t64z67WE5m8Vj168mCda8kLQMOkbtPpuovxvEVJJJRNmq2PY6VrSWtIDBkEjHT1rRZzXNcWuBDgcEHqK9U9i+v6HaVoePUtCwRNdVTwPiByWFkhDc95ZuO/WWgHKt0Z7ydtt7ooYuboa9/jCjwMDclJJA7g/fb6l57ZrqRE9Ux/f5jwbRi1Jx1xP9fifFipS+i7dT3fWNltNWXinra+CnlLDh26+RrTg9RwVELsezD4StMfO9J9c1ejRiJ1KxPOGGrMxp2mOUt2qnkg7KI6aWRtTqTLWEjNczqH+GtBZmhkz2Doa4gL1/rvuKf/Dd+xeQNV90y/nn9q88TP8AyTHZ+3oxH/HntU0RFozFIWSojZM+kqD/AFapG4/8k9TvUVHogrVlPJS1UlPKPLY7B7+9UVK139ftMVcOM1PiKfvHxXfwUUgIiILKZm5IR1dIXBXVU3LN7rCtUBERAREQEREBXNI3AL+3gFbK/Y3dYG9gQfUREBStv/qVpqK48JZvsEP/AO4qMiY6SRsbBlziGgdpKkdQPa2ojoYzmOkYI+HW7pcfb+xBGIiIC229ze++DWX6LTfvyLUlbbe5vffBrL9Fpv35Fro9c90+DLV+GO+PGGQ+XDqW46Op9BamtMm5WW69PmZx4PAj8ph7nDIPcV3/AFzZ7Ht32DFtA9jo7rRtq7dK7pgqGjLQewh2WO7t4LEvujn3j6V+cpfql133P3aTzFZW7NLpUfY596stW8eh4H2WIekDfA7ndqw06xq11NOeeY+0f79G+padOaXjlifvP+fVmfkZUVVbdg1tt1dA+Cqpa2shmieMOY9s7w4H0ELVTl3fhB1vzfS/uL0Jo6SmoxK2lgZCJZXTSBgwHPccud6SeJXnty7vwg635vpf3FzXtNr0mfW6XdGsVreI9dKGB1sXyCdGeP8AazNqWpi3qTT9OZWkjgaiTLWD1Dfd6QFrovRrkV6M96exGgrKiLcrr483CbI4hjhiIejcAP6xXo0ujE35eM+pn6PPqb8V5+Hrd9Vty4tX+9rYlU2yCXcrL9O2hYAePN+dKfRujd/XWPfc6tX87a9QaHqJfKp5G3GkaT8V2GSAdwIYf1is/at19snpbpJadUam0s2uo3br6eumiL4SQDjDug4wrG1bSdiVLWNdbNWaMp6mT7G10E8LHOyejI7ThZaPRm09e16j872urvisdWPU/jcwZ7onoznaGw68pYvKhcbdWuA+K7L4ifQd8frBaZr1Z2zaRi11svv2mHNaZKykd4OT8WZvlRn/ADBq8qZ4pIJ5IJmOjljcWPa4YLSDggrGnRtNfr9/78Wt+lWLfT19PBwWZORf+Edpr0VP/wAvIsNrMnIv/CO016Kn/wCXkXr0Pj+k+Dy63wS3h5SnwCa0+aZv2Ly6XqLylPgE1p80zfsXl0vJX5tu6PN6p+XHfPkLPHIR/CDovm+q/cWB1njkI/hB0XzfVfuL2aHxfSfCXl1vh+seMNueV5+Drq39Hi+vjXmcvTHlefg66t/R4vr415nLyU+Zb6PVb5cd8+Tbn3N0xeONZtOOd8HpC3t3d6TP04WRvdA2zHYdTGLO4L1AZcfi83LjPrwtaeRxr6k0Jthp/Gk7YLZeITQVErjhsbnOBjeewBwAJ6g4lb47YNE0e0XZzddJ1Uoh8MiBgnxnmpWkOY/vAcBntGVt7VE206zHZ+Jz4eLH2eYrqWiePnGHlKs58hhsx5Q9rMWd0UdUZcfi80en17q6vqPYNtbsl5ktkuh7vWlr91lRQU7qiGQdTg9mQAfysHtAW1PIz2H3jZ8K3VuroGU16rofB6ajDg51NCSHOLyMjecQ3gOgDjxJAvRmIzbhifzGEasT8PbHiyhyljENgmtDNjd8VSgZ/Gx5P04Xl2t9OXvr+jsuzdmiKadrrpfJGOljaeMdMxwcXHs3nNaB24d2LQteWm+9rd0fb/XptupWO+fX2bO+51fCnf8A5lP10azTy6bxX6f2b6dvlrnMFdQakp6iCQdTmxSkeruWFvc6vhTv/wAyn66NZa90M+Bi1/PsX1My39ptNaacx2fzY+zxE3vE9v8AFkQeINvewT4jae90XpNJUt/iyQesDvXVuRHZq/T2ym6WO6wGCuoNQVdPOw9TmtjB9XYexYQ5Au0nxNqup2fXOo3aG8EzUBceDKlo4tH57R7WjtW71LSUtLJUSU8DIn1MvPTFoxvv3Wt3j34a0epXaIpabV6rR5x/cfZFZm1YrbrrPlPr7tD/AHQv4aLb8xxfWyrW5bI+6F/DRbfmOL62Va3LzaPwz3z4y9Ot8X0jwhnLkTaM99W2yjr6iLfobDGa+XI4GQcIh6d4736hW3HK/wBX+9DYXeXwy83WXQC202Dg5lzvkeiMP+hdX5BujPe9sifqGpi3azUFQZwSOIgZlsY9Z33frBZQ1nrrZfbrk6y6t1JpqGsp8PdS180e/HvDIO67oyD7Ct/aa5rGl9/P8bmHs9sWnU9bur872sfuder+YvOoND1EuGVUTbhSNJ+OzDJAO8tLD+qV2/3QrRnjLQ9q1rSxZntE/g1S4DjzEp4E+h4A/XKytQ7SthlFUtqKLVmiqaccGyQzQscM944rt2vtPUWtNB3fTtQ5j6e6UT4WvHEAub5Dx6Dgj0KPaJm9ItEb48v63L0MUvMTO6fP+97yWXY9mHwlaY+d6T65qhrvQVVqutXa66IxVVJO+CZh6WvY4tcPaFM7MPhK0x870n1zVt7NMTqUmOcMvaImKWieUvV+u+4p/wDDd+xeQNV90y/nn9q9gKtrn0srGjLnMcAO/C82Z+TdtrdPI5uhpyC4kf16m7f8ReWIn/kmeyPN6Mx/x47WIUWUrpye9sVrtlVcq/Rc8NJSQvnnkNZTncY0EuOBJk4APQsWrTMZwjE4yIiLriQsU7I6wwTfaKhpik7s9B9RVpVwPpqmSCQeVG4tKpKUvH9ao6W5Di57eam/Pb1+sIItERAIyCD1qwe3dcWnqV+rarbhwd2oKCIiAiIgIiIKlO3elHdxV4qFG3g53qVdAREQSen2tZUS1rwCyljMnHrd0NHtUc9znvc9xy5xyT2lSUn9W07GwcH1cpcfzG8B9Ki0BERAWbOSpthseyO532qvdsuNe25QwxxijDMtLHOJzvOH4ywmiqtpr1OWrFt0thOVTt207tb09ZrbZbPdaCSgq3zyOqxHhwLN3A3XHisG6Xvdw03qK3361TGGuoKhlRA8dTmnPHuPQR2FRqKadC21Xr63bdONmepvLTctLRJp4zU6U1C2bcHOCPmS0OxxwS8EjPctYeUbtAtu0zadUaqtNFV0dLLTQwiKq3d8FjcE+SSMetY4RctWLTE8nazNYmIXNsNG25UrriJnUYmYagQgb5jyN7dyQM4zjK3ZpOWTs+oqGKko9IaijigiEcMYEAa1rRho8/owAtHUVTaZrsp2YztJPVV6q9R6mud/r3F1VcKqSplOfjPcXY9AzhRoJBBBwQviKaxFYiI4KtM2mZlunozlj6ZoNJ2qhv2nr9U3SnpI4qqaDmiyWRrQC8ZeDxxniOtarbW73YtS7Rr1qDTlFVUVuuNSallPUhofG9/F48kkY3i4juK6qi7bpX256/25Xo12I6v0LvewPWtBs82qWnVtzpKmrpaIS78VPu847ficwY3iB0uHWuiIqraazmHLVi0YluDtY5V2jdX7Nr/pih05f6epudE+nikmEO41zh0nDycegLT5EWezG1tK2pxgWSOTltAtuzPadT6qu1FV1lLFTTQmKl3d8l7cA+UQMetY3RaVtNZzCbVi0Yltrtv5UekNebLb3pO3afvtLVXCJjI5agRc20tka7juvJ6G9i1KRFnFYiZlU2mYwLYnYnyqdU6JtlPYtS0PvltMDQyB7pubqoGDoaHkEPaOoOGerexwWuyLSLTG5M1id7fqk5Y+y6WnD57VqiCTHFhpIXce4iX+S6fr7lo0popKfQulanwh4IZV3ZzWiPv5qMne9bx61pqiiYyqJwldWaivWq7/AFV+1BcJrhcap29LNKeJ7AB0AAcABwA6FFIiRERGIJmZnMsu8lvapZ9k2srle7zbq+uhq6A0rGUm5vB3ONdk7xAxhpXduU/yhNM7VtBUenrNZbvQ1EFxZVukqxHuFrWSNwN1xOcvC1sRdv04iJ4fvPi5ToTMxx/WFzaq+rtdzpblQTvp6ulmbNBKw4LHtILSPQQt1rTy0dJC10oumlr6a8QsFSYOZMZkwN7dy8HdznGQtIEVbU7Oy5sxnLK3Kf2m2natr6k1DZqCuoaeG3spXR1e7vlzXvdkbpIx5QWLIub51nOlwj3hvloycdeO9cEUUiKdSrTNutu3p/le7OLHYqCzUGkdSR0lDTR08LQIODGNDR/adgWom0jU9TrPXd61TVhzZLlVvmDHHJYwnyGfqtAHqXX0S0bVtuev9uV6NdiOr9C3E2X8rrT1g2fWSx6isN8rLlb6VtNLPT80WSBnktPlPBzuhueHTlador2pxsubMZy7zt11Rp7Wu026ap03QVlBSXEtmkgqg0ObNugPI3SRgkb3pJXW9IXKKzastF4njfJFQ10NS9jMbzmseHEDPXgKKRTp+7xs8HdT3mdri3p/po6D/wCVdS+yD/UT+mjoP/lXUvsg/wBRaLIg3S1pyvdE3zR95stPpnUMU1fQT00b5BDutc+MtBOH5xkrS1EU7MbW07tTjAiIqcFKWb+s0dZbzxL2c7F+e3s9IUWri21BpK+GoHxHgn0df0ILdFeXmnFLc54W+YHZb6DxH7VZoCp1Dd6I93FVEIyMII9F9cMOI7CviAiIgIi+gZOEF5AMRN7+K5oBgYRAX1rS5wa0ZJOAvivrBEJrvTtd5rXb7vQ3j/BBV1E4NrWUjDltNE2Id5xkn2lRiq1cpnqpZz0yPLvaVSQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQSl3+z0FBW9bozE/0tP8lFqTpvs+naqLrp5myj0HySoxAREQWlSMSnv4qkrisHmu9St0BERAXOEZlaO9cFVpRmX0BBdoiICk7H9jir6n+7pi0HsLuAUYpOn+x6cqn/AN7OyP2DeQRiIt5/c5/g11H87j6liutdqJnl+4hNrYmI5/powi9ZtZ650ho00w1TqGgtHhW9zHhUm7zm7jex6Mj2rrv+3HZF/wCYNh/6kKM5Vh5dos3ctPU2n9WbYmXXTd2pbpQi1wRGenfvN3w5+W57Rke1YRU0ttRnHNVq7M4FfUFnu9wjMlBa66rYOl0NO94HrAW2HIk2JafvdgO0TVtvhuYfO6K2UlQ3ehAYcOlc08HHeyADwG6T04xsTr3bBsx2dVkdm1FqOlt9W2MFtHBTySvjaejLY2ncGOjOFpeIpiJnezpM3zMdTzAraKsopObrKSemf+LLGWH2FW69NLPtz2Kau/3e3WFnlEnDmblC6na49n2drWlaU8sC3We17d7xTWKhoaKgMFNJHFRxNji8qFpLgGjHEnOR05WczNZiJjraViLROODEKKT0nPFTaptNTUSNjhiroXyPceDWh4JJ9S9Mf9uOyL/zBsP/AFIWk16MTn1uZxbpTDy7ReuOk9Taf1ZajdNN3alulCJDEZ6d+83fGCW57Rke1R+sdoOidHVcFJqjU1utE88fORR1Mu6XtzjI7sqZ3dao39TyeRehPKC2ubM73sX1VabTrazVtdVUD44IIp8vkcSOAHavPZTFs2mMKmvRiRF6Mchv8He0fpVV9a5a8e6GfDPbPmOL62VVq+7tFef6ynS95Wbcv3hrciLfP3O/4Hrx8+yfUwq602omeX7iPNNrYmI5/poYi9bdeacotX6Nu2mbg0GmuNK+BxxndJHkuHeDgjvC8otSWit0/qC4WO5RGKsoKl9PM09TmOIPq4LHa6Wy02ejtI9Fuf7nloTmLbd9oVbDh9SfF9vLh8RpBlcPS7db+qVtyVteuzhnW21l47opjW/36Xz5xqPrHKHWOnbbrFubS9dm015CLsGznTVTrHXdl0xSA85caxkBcB5jCfLd6mgn1L1ftVDTWy2UttoohFTUsLIYWDoaxoAA9gW2ziu0z2ulsvIBFvZ7oLozxrs9tusaaLeqLLUc1UEDjzEpAyfQ8M/zFaJrGttqZjk0tXERPMRfW+cPSvTexbbdk0NkoIZdf2JkjKaNrmmoGQQ0Aha7PRzlntdLDzHRetejtY6X1jSz1Wl75RXeGneI5n00m8GOIyAfUqestb6S0aymfqnUFDaG1RcIDVSbvOFuN7HoyPapnd1qjf1PJhF6UbQts+yqu0FqCipNeWSaoqLZUxRRsqAXPe6JwAHeSV5rqYt0pjHreqa9GJEW+fud/wAD14+fZPqYV0L3SH/j+jP0Wq/fjVa3u7Vrzx4ZTpe8rM8v3hqSiLJ3J12SXDa1rM25kz6O0UTRLcqxrcljCeDG54b7sHGejBPHGDVazacQm1orGZYzjY+R4YxrnuccBrRkkqQmsF9gg8ImstyjhAzzj6V4bj0kYXp7pjR2zfZJpmSpt9vtdioqWLNTcagtEjh2yTO8o5PVnrwB1KAtfKO2MXG7ttlPranZM9+4189LPDET/iPYGgd5IC5umdmvW7vxmXmei9N9r2w/QO0y1SvqbbTUF2kZvU92oo2tlDiOBfjhK3udnh0EdK86doukLxoTWVw0tfIgysopN0ub5krDxa9p62uBBCz2ulsyvZzG1DryIvRjkN/g72j9KqvrXLatcxM8mdrYmI5vOdF6VcrLZuNouymrjooBJerTmtt5A8p5A8uIfnN6u0NXmu0ESAEEEHiCsqW2rzVpNcVizii9e7D/AMCoP0aP90Lya1n9+F6+cJ/rHLt52dTY7/wnT6ent935cNP+XLVU399TPaB3jiP2KMV/p+Tm7zSu7ZN328P4q1qo+aqpYvxHlvsK6KSIiClVDMWewq0V7MMxO9CskBERAVejHlOPcqCuKP43qQXCIiApOo8jTlK3+8ne/wBgAUYpO5+TZ7XH+TI72uQRi3n9zn+DXUfzuPqWLRhbz+5z/BrqP53H1LFrp/Dfu84ZanxV7/KXZuVtsX1Rtbdp06br7NSeLBUc94wmkZvc5zeN3cjfnzDnOOpYF/oZbUPl7R3/AFdT/oLYjlQbbbhsfdYRQ2GluvjQTl/PTuj5vm9zGMA5zv8A0LCn9NfUH/Ilr/62T/8AqvPTZxOzzl6L53Za47SNIXLQetLhpO8T0k9dQOa2WSle50RLmNcN0ua09Dh0gLrq7TtW1jPr/X901dVUUVDNcHsc6CN5c1m6xrOBPE+bldWXabWzG11uXxno9T0V5Dt9oLrsDtlvppWGqtU81PVRg8Wl0jpGkjsLXjj3HsXUuUNyXK/Xut6/WOmtS0tLV1+4Z6OvY7m99rA3LZGZIBDRwLT18epaf7Ntf6r2eX3xxpS6yUM7gGzRkB8U7fxXsPBw+kdRC2V0jy1atjI4tWaJhmPx6i21RZ7I3g/vrXViupbb4+ss9OZ04mvBiDWvJu2u6WgkqptNG6UsYJdNa5RUcO3cGJMfqrEkoka8slDg9nklruluOpepmx3appPapZai46YmqQ6le1lVS1Ue5NAXAluQCQQcHBBI4Faue6FaJtVpv1j1jbaaOmnu3OwVwjbgSyMDS2Qj8YhxBPXgLG8205iLdTSkReJw1TREVpegfIA+AiT55qP3I1iL3Rr4QdM/NT/rXLLvIA+AiT55qP3I1iL3Rr4QdM/NT/rXJ7X8yn0/iezfLv8AX+TVlERB6Mchv8He0fpVV9a5a8e6GfDPbPmOL62VbD8hv8He0fpVV9a5a8e6GfDPbPmOL62VPa/m19f9ZPZfl3+v8mty3z9zv+B68fPsn1MK0MW+fud/wPXj59k+phW2n8N+7zhlqfFXv8pZYrdcMt+3Wj0JWSNbFdLJ4XR5/vo5H77fWzj+otW+XDsxrDtbst8sVIZPfU9lI5rBw8MaQwZ7N5paf1XFSXLg1DXaT29aI1JbnFtVbqFlQzjje3Z35ae4jIPcVthaJbDrawWHUjKeGsp3CO5UD3jJie5hAcOxwD3BeelYvSupPC0xPdmfLwb3tNL2pziMd+I8/GXTdR1lt2GcnlwpSz/ctubT02Rjnqp3AHH5Uji492VkWySyT2WhnlcXSSU0b3uPWS0Elac+6Fa78Kvdn2fUU2Y6MCvrw09MrgRG0+hu879cLcPTv3vW79Ei/cCutp1K31J4z+/PyZ2rGnNKRynyx+PF5Pa3+/S+fONR9Y5Q6mNb/fpfPnGo+scodY6Hyq90N9b5lu+W0/uemjPGGs7vraqizDaoPBaVxHDnpfOI7wwEfrrM3LB2p1Gzui0nBbpi2rqbvHVzsaeLqaBwc9p7nEtHqK7TyV9Ge8jYnY7fNFzddWx+H1mRg85LhwB72t3W+pYR5VWxza3tM2qS3az2amls1JTR0tCX18TC5oG8926XZGXud6gF6NeZrqVrH/WY/G/x/DDQiLUtaePnu8N/e2f1RarZrvZ9XWp72y2+9W9zGSDiN2RmWvHoyCPQvKW+2yrst7rrRXxmOroqh9PMw9T2OLSPaF6d8nq1arsGyWy2DWdIymuttjNLhk7ZQ6Jp+xnLSR5uB+qtOuXfoz3ubYPH9NFuUeoIBUZA4CdmGyD1+Q79ZY6sRTW6PVO7zj8ZaaUzfSxbrjf5T5NfERFaW73ucf3j6q+covqlGe6R/cGif8Ws/ZCpP3OP7x9VfOUX1SjPdI/uDRP+LWfshT2z4q//AD4HsnVP/wBebTRERBvn7nf8D14+fZPqYV0L3SH/AI/oz9Fqv341333O/wCB68fPsn1MK6F7pD/x/Rn6LVfvxp7X8yn0/iezfBf6/wAmpK9DeQjYKe07CKa5tjAqLxWTVEr8cSGuMbR6AGE+srzyXo9yJrjDX8nexxROBfRS1FPKB8Vwlc4f9r2n1rWnwWx2fb/cMr/FX1wYX90T1bWuv1h0TBO9lFHSm4VDGnhJI5zmMz27oa7H5y1JWzvuiNpqafahZLy5jvBqy1CFj8cN+OR5cPY9p9a1iXl0eqZ7Z8Xq1euI7I8P29B+Qhq2t1HscfbLhO+aax1hpInvOTzJaHsGe7LmjuAWM/dHLBTxXHSmp4ow2eojmop3AecGFr2fvPXcvc77TU0my69XWZjmRV90xCSPObHG0Eju3iR6iuv+6QXGEWzR9oDgZ3TVNSW9YaGsaD6yT7Fr7Z8dZ47vzG/8ZZeyfDMcN/jOPzhpkvRjkN/g72j9KqvrXLznXoxyG/wd7R+lVX1rlrp/Bb1xZanxV7/KWQ7Trehq9qF70FNuRV9BSQVsHH7dDIMO9bXAZ7nBaLcsfZt7w9q0lxoKfm7LfnOq6bdHkxy5+yx+px3gOxw7F3TlM60rtn/LAodV0Bc40dHS89EDjnoSHCRh9LSfXg9S2J26aPtu2fYkX2Z8dTUPgZc7NOOt+7kN7t5pLT2E9y88b611o4TMT9/1v74l6YnF50p4xEx9v3+JZMsP/AqD9Gj/AHQvJrWf34Xr5wn+scvWeysfHZqKORpa9tPGHNI4gho4LyY1n9+F6+cJ/rHLut/+j7+MI0N3s/28JR1I/m6qGT8V7Xewq5v7ObvNU3tkLvbx/irFSep+N4kf+O1jv+0KnEYiIg+PGWkdysFIKPQEREBXNH5rvSrZXNH5rvSgroiICk7xworaP/h8/SVGKTvX3JbT/wDDAfSUEYt5/c5/g11H87j6li0YXetnO1zaHs8ttTbdH6g8WUtTNz80fgcE28/AbnMjHEcAOA4K6WiItHOPOEWrMzE8v1Lfrb5sUs+191oN1vNfbfFYl5vwZjDv85uZzvDq3B7Viz+hXo//AJxvv/5UX8lgD+k7tx/53/8ASqL/AEU/pO7cf+d//SqL/RWcViOppMzPWocp/ZVbNkur7bZbXdKy4xVdD4U6Spa0Oad9zcDd6vJWJF2jaLr/AFdtDulPdNYXbxnV08PMRSeDxQ7rMl2MRtaDxJ4kZXV1ykTEb+120xM7m6vJ72I7PNpXJ4sdXqC0OhuZkqWi40T+aqMCZ4G8cFr8Do3gcdStLryJaV9SXWraHNDATwjqbWJHAfnNlaD/AJQtWNH671lo+Tf0xqa62oE5dHT1DmxuP5TPNd6wsmW7lU7Z6SIRy6goqzHxp7dDn/sa1aX2bW2oZ0zEYluRye9jNn2P2evp6K51F1r7i9jquqljEYIYDutawE7oG848STx6Vrb7oNri23rVVm0fbKmOodZmyy1zo3ZDJpN0CPP4zWtyezewsbaq5R+2HUNHJRz6skoaeQYc23wR07iPz2jfHqcsSyPfI90kjnPe4kuc45JPaVnfa1LRM8GlMUicdcuKIipL0D5AHwESfPNR+5GsRe6NfCDpn5qf9a5Ya2f7aNpegrCbFpPUni63mZ05h8Bp5fLcACd6SNx6hwzhRG0baFrDaJcKW4axu/jOppYjDC/waKHdYTnGI2tB4nrTX95eto4Y8MGj0K2ieOfHLqqIiD0Y5Df4O9o/Sqr61y1490M+Ge2fMcX1sqxpobbhtR0Rp2HT2mNT+AWyF73xweAU0uC45cd58bncSe1dd2h661VtAvUV51ddPGVdFAKdkvg8UWIwSQMRtaOlx44zxTX95eto4frBo9Ctonj+8utrfP3O/wCB68fPsn1MK0MXfdne2HaNs+s81o0hqLxbQzTmoki8Cp5syFoaXZkjceho4ZxwWlLRWLRzjziUWrMzE8p8pZk90X+EzTvzP/8AzPXV9jvKc1Rs50PT6ThsdvutNSyPdBLUSva9jHHe3OHUCSfWsXbRtoGrtodzp7lrC7+M6qmh5iGTwaKHdZkuxiNrQeJPEjK6usdKJpWa85nxy11Ji9onljwwnNc6luOsdZXPU91INZcal07w0ndZk8GjPU0YA7gvV3Tv3vW79Ei/cC8hhwOVl6m5S+22np46eHWu7FEwMY3xXRnAAwBxiWlZiunsQztE21Nue3yY51v9+l8+caj6xy7PyetGnXe1+wWB8ZfSOqBUVnDgII/LeD6QN30uC6NXVU9bWz1lU/nJ55HSyuwBvOcck4HAcT1LsGzvXuq9n13mu2kLm221s8JgkmNLDMTGSCQOca4DiB0din2b3UVzw8Y/tXtHvJtji9QdouqKHQ+hLvqmtjD6e2UrpREHbnOOHBjAcHG84hucHpWrn9N7/wC7H/17/wDzrAGutt+1HXGnpNP6o1S6utkr2ySQCip4d4tOW5McbTgHjjOFjpTG1tTM9SujsxEN/dhvKjo9pO0Gm0jVaR8Rvq4pHU8/jPwgPkaN7c3eaZjLQ45z1YxxUty3tGe+nYrVXKni366wSiujIHExebKPRune/UXn5p283PT19or5Zqt1JcaGZs1PM1oJY9pyDggg+gggrJdw5SW2i4UFRQVusWz01TE6KaN1qo8PY4YcD9h6wSu6tYvTFev1MOaVprfM9XrLEiIi643e9zj+8fVXzlF9Uoz3SP7g0T/i1n7IVrZs42s7QNndDVUOjr/4sp6uUSzs8Dgm33AYBzIxxHDsVLaTtR13tGjoo9ZX3xo2hLzTDwSGHcL8b32tjc53R056E1/eTExwx+IwaHu857fzl0xERBvn7nf8D14+fZPqYV0L3SH/AI/oz9Fqv341gbZ3th2jbPrPNaNIai8W0M05qJIvAqebMhaGl2ZI3HoaOGccFYbSNpOtdotRRT6yvXjSShY9lOfBYYdwOILh9jY3OcDpymv7y9bRwx4YNHoVtE8c+OXUVsRyL9sdDs+1DVaY1LUinsF3ka9tQ4+TS1AGA53YxwwCerDT0ZWu6KqW2ZTau1D1S2s7O9L7V9HNs96y+EkT0VbTOG/C8jg9juIIIPEcQR6iNerXyJqGO7tkuWv6iotzX5dDBbRFK9vZvmRwB790+ha26A2v7SNCQNpdM6rraWjaeFJLuzwDtwyQODfUAu8zcq/bLJAY23i2xOxjnGW6Le+kEfQpxWs7VVZmYxLe6MaR2YaBjidLS2TT1op91pkdhrWjv6XOcfSXE9ZK84+UTtKm2o7Squ/tZJDbYWimt0L+lkDScE/lOJLj2Zx1Lr2ute6y1zVtqdWair7q5hzGyaTEUZ/JjbhjfUAutKJibW2rKiYrXZqL0Y5Df4O9o/Sqr61y851kbQ23DajojTsOntMan8AtkL3vjg8AppcFxy47z43O4k9q2raIrMc2VqzMxPL9O7cvP4f6j5tpv2FZb9z/ANpPh1mrNm10qM1FCHVVsLjxdCT9kjH5rjvDucexaj691lqTXWoHX/VVx8YXJ0TYjNzEcWWt6Buxta36FbaR1HetJaipNQadr30Fzo3F0E7Gtdukgg5DgQQQSMEEKND3dZrbqnPjmF63TmJr1xjww9civI3Wf34Xr5wn+scsmf0nduP/ADv/AOlUX+isR1tTNWVk9ZUv3555HSSOwBvOcck4HAcSs5rO3FuyfJcWjYmvbHmoqT1H92xHtp4z9CjFJ6i+64B2U0Y+haIRiIiArB3nH0q/Vg7zj6UHxERAVzR9DlbK4o+lw9CC4REQFJ3bjbrY7/3Lh7HKMUnX+VYba78V0rT/AJgUEYiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgKT1LwuQb+LCwf8AaFHMbvPa0dZwpDUzt6+VOOgED2NAQRqIiArA9JV+eAJUegIiICrUh+yEdyoqpTnEwQXiIiApP7ZpjviqvoLf5qMUnbPstnuMHWGslb6jx+goIxERAREQEREBFMjVGoAABdajA7wnvp1B8q1HtCCGRTPvp1B8q1HtCe+nUHyrUe0IIZFM++nUHyrUe0J76dQfKtR7QghkUz76dQfKtR7Qnvp1B8q1HtCCGRTPvp1B8q1HtCe+nUHyrUe0IIZFM++nUHyrUe0J76dQfKtR7QghkUz76dQfKtR7Qnvp1B8q1HtCCGRTPvp1B8q1HtCe+nUHyrUe0IIZFM++nUHyrUe0J76dQfKtR7QghkUz76dQfKtR7Qnvp1B8q1HtCCGRTPvp1B8q1HtCe+nUHyrUe0IIZFM++nUHyrUe0J76dQfKtR7QghkUz76dQfKtR7Qnvp1B8q1HtCCGRTPvp1B8q1HtCe+nUHyrUe0IIZFM++nUHyrUe0J76dQfKtR7QghkUz76dQfKtR7Qnvp1B8q1HtCCGRTPvp1B8q1HtCe+nUHyrUe0IIZFM++nUHyrUe0J76dQfKtR7QghkUz76dQfKtR7Qnvp1B8q1HtCCGRXlzulwuRjNfVSVBjzub/Vnp/YrNAREQEREF1ao+dudKzqMrc+jKXaTnbnVP6jK7HoyrjTTR41ZK7zYWOkPqBUc4lzi49JOSg+IiIOMhxG49ysVeVBxC5WaAiIgLkw4e09hXFEEgi+RneY09oX1AUnpsh1wdTk4FRC+L2j/wCijFWoZvB6yGf+7eHewoKRBBIIwR0r4r6+w8xdqhg80v3m+g8f4qxQEREBERAREQEREBF27Znoaq1tXVkMVbHQwUkIklnfHvAEnAHSOwn1Kz2h6UqtHajfZ6mdtQObbLHM1m6HtPXjJ6CCPUk7sZ4kb844OuoiyXrDZFcrBpH3xQ3OOvjY1kksLIC1zGOHnZyc4yM+1J3RkjfOGNEXbNmWi5db3WqoIrgyiNPBzxe6Iv3vKAxjI7V1aePmp5Iic7ji3PbgpO7cRvcEWWLTsbFXpy33uq1ZRUEVbCyVonhwGlzc7uS8ZKtNR7Gr9QWt9ztFfRXymY0ucKYkSEDpIHEO9Rz3Jbo5yV6XUxkiIgIpPStodftR0FmZOIHVkwiEhbvBuevHWsn1exKnpJ+Yqtc2uCXAO5LEGu49HAvXcbsuZ34YdRd22ibN71oyCKtnmgrrfM7dbUwZwHEZAcD0Z6ukKA0dZH6j1NQ2RlQ2ndVyFglLd4N4E9GRnoXK9KcQ7PRjMohFmGq2J01JOYKrXdqglHSySMNcPUXrruvtlt50raheGVlLc7bkB08GQWZ6CR2HtBK5MxDsRl0BEWTtN7Gr7cLUy6Xi4UVjppGhzBU5L8HoJHAN9BOe5VjdlOeDGKLJGrtkF+s1pfd7fWUl5oY2l73U2Q9rR0u3eIIHcSsbqcqwIpzRFhi1JqCK1TXOC2sexzvCJgN0YGccSOn0rtO0PZe/SWmor42/U9xhlmbE0RQ7oO8Cd4O3iCOC7PRjMuRvnEMdIpPStodftR0FmZOIHVkwiEhbvBuevHWsnVmxOmo5jDV67tdPKBksljDXY9Beu43Zczvww8i73r3QFHpiysuMGq7ddXumbFzEAG8AQTvcHHhw+ldHhjkmmZDEx0kj3BrGtGS4noAC5G+cQ7O5wRZWtmxK7m3R1l9vdusvOAYjlO85vc45AB7gSoTXuy++6Ut4uhmp7lbSRmopyfIz0FwPQD2gkJO7rI39ToiKc0RYodR3+O1z3SntjHsc7wifG6MDOOJHT6Vkmk2Gx1cUktJra3VEcfnuih3g30kP4LsxhyJyw0i7/rnZ5RaasLrnBq623R4kazmIAN456+Dz0ehdAUxOVYERF1wREQEREBERBJ2n7FbrjUn+6EQ7948f2KMUpP8AYNO08fQ6omdIfQ3gFFoCIiCjVnyAO0q1VarOXgdgVFAREQEREF3SnMWOwqqrWkdh5b2hXSAiIglLv9noKCtHEmMwv9Lej2hRalLb/WbTWUR4ujAqI/Vwd9Ci0BERAREQEREBEV5Y6JtxvFHQPmjgbPM2N0sjg1rATguJPAADiuxGZw5M4jLMenrdX6e5P9XUUFFUz3O/PwBBE572xu4A4AzjcDjn8pU9s1DVXzZlp3VtRSzQV1MwU9ayWMseM8MkHjjfbw/PV5tX2l1em7hb7Jou4UfglNSNEj42smbnoa0HiOAH0r7oTWp13pPUOn9YXShinkh/q0spZCDkHHYCWuDT61N524tavrG52nR2Yn1lgcLaq+6kp7Re9MWa57rrZeqB1NM1/mh+GBpPcd4tPp7lqvIwxyOY7GWkg4ORwWW+UNcqCtptLeAV9NUuhpXh/MTNeWHDOnB4HgqzGzGerPlLmJ2vp5w7Nsl0xNpLa3qC1ODjB4AZKV5+PEZG7vrHQe8LAdf93VH+K79pWzOy3WljvunaS5Xi40VLeqOF1HM6edsbpG+Sd7ieIO6D6crWWuIdWzkEEGRxBHpUzExaK8o81VnNZnnPkzZtY+ALSX/4b6hygOTTda6m154rjleaOrp5DLFnyd5oyHY7eGPWu73G3WPV+yXTllk1Za7ZNTwwSvMkrHEERFpaW74IPH6FHacZoPZTTVd1Oo6e/wB5kiMcUdMWnA6d0Bpdu5IGXE9A4d9zOzqXmec+CMbWnSI7PFizapRwUG0S+UtM0NibVuLWjoG95RHtK6yru83Cou12q7nVkGeqmdLJjoy454dytFnSMViJaXnNpmHadkvwlWD9NYssbXtl+pNV60ku9skt7aZ0McY56ZzXZaOPANKxHstnhptodjnqJo4YmVbC98jg1rR2knoXZNvt48I2hzSWy6c7TGmiAdT1G8zOOPFpwrvjZr3z4Ir8Vu6PF27apJSaU2O0GiKu4RVt2JZlrHZLGh++Tx4ho80ZxlY42LfCjYv0g/uOXUHOc5xc5xc48SSckrteyCogpdpNlqKqeKCFk5L5JHhrWjdd0k8Au6fzNqeMuany8RwhlPatsp1JqnWtVebdUWxlNMyNrRNK9rvJaAcgMPZ2qlqvwTQGxibR1yulPW3irB3YInE7gc8EkA8Q0AHiQMldK213uQbUqmus9032RiB8UtPNvNDgxvQQcdIXZtoNTZNoezmk1LFW0NLqCgjLZ6d8zWPkA89oBOT+M30kdJWX/lu6uPc1/wDSM/TvY52XUcFftDsdLUtDoX1jC5p6Djjj6F2/lL3StqNdttckrxSUlOwxR58necMl2O3q9SxtZrhUWm7UlzpSBPSzNljz0Zac8Vm/UbNBbVqakunvjg0/eY4hHKypLRkdO6Q4t3sEnDmnoPHu0tGa1xwnyZ1nFpzxhD8l651g1LcLKZHPoZqQzGMnLWva5oyB1ZDiD28FjXWtJBQawvFFTACGCumjjA6A0PIAWYLPV6H2TWmuqKC+w6hvtTHuM5gtLR2A7pIY3OCcnJxwWDayomq6uarqHl800jpJHHrcTkn2qbTm0Y4QqsYrOeMqKzltF/Bz01+dT/uPVnbLFs31loi3xUl0t2mbzAG+FGVw3nuAw7g9w3gekEHgqW2jUWn4dGWfRFhuMdyFEWGaojcHNAY0tAyOBJJJ4dGF3U3VmvbDlN9ot2S6Psl+EqwfprFl/axs+teotYSXOr1pbbTK6GNng04aXAAdPF46fQsObLZ4abaHY56iaOGJlWwvfI4Na0dpJ6Fl3aXoqwax1RJem6+s1EHxMj5ouY/G6OnPOD9i7b4a98+Dlfit3R4sX7Q9GW/S1NSTUWqaG9Goe5rmU4aDHgA5OHu6cq62CUkFXtRtbZ2hwiEkrQfxmsJHsPH1KtrXZ5atPafmulLra2XWWNzWimha3edkgZ4SHo6ehdU0bfJ9N6mob3Tt33Usm8WZxvtIw5vrBITTnFjUjNXZtvVzrK/aVcoKmV5io3NhgjJ4MbugnA7ySV1+h1dqKi03Uadprk9lsqCechLGuyD0gEgkDuGFlrVVm2f7SatmoKDWFLZa6VjRURVO6CSBjixzmnIHDIJBwFC6pZs50hoepsVvmo9S3uqP3WGtcITjG8HNJDQOpoJOelZxGzTE/wCrmdq0TH+MRLOXJ3+8PV35h+qcsGrM+wS526i0TqqGsuFJTSStPNslmaxz/sbhwBPFXb5d+5Nfjr3sMIiICIiAiIgIiIC+saXuDWjJJwAvikdPRtdcOfkH2KmYZn+ro+nCDlqJwbWspWHLKaJsQ7yBk/SVGLnPI6aZ8rzlz3Fx9JXBARFxkdusLuwILSY70rj3rgiICIiAiIg5MduvDuwq+UeryndvRDtHBBUREQXdoqfBLhFM7zAcPHa08Cvl1pvA7hNB8VrstPa08R9CtVK139cs9PWDjJT/AGCX0fFP8EEUiIgIiICIiAiIgItnNbaN05W7Pqmkt1jtlNdBbW1cMsFIxkjizBPlAZ49B9KxZsBsVvueoq+53ikgqbdbKR0kjJ4w9hceAyDwPAOPqXYjpTXl68nM9GLc/wCv2xsiylykLXbLVqy3w2u3UdBE+hD3MpoGxtJ33cSGgcVjCCKSeeOCJpfJI4MY0dJJOAFNZ2upVo2XBFs0NIaRFtGgjabb48Nl57wzwdnO7/m72/jezvcenoWtNRDJT1EkEzCySN5Y9p6QQcEJM9LHrrw5G+M+uamiyjycLXbLrq6vhulupK+JlCXtZUwNkaDvt4gOB4rht30nRWq40morDDEyzXNgLWwsDY45AOgAcACOOO0Fdt0cdpXpZ7GMUWWNpdptVJsc0jX0lsoqerqAznp4oGtkk+xk+U4DJ49qxOk7rTXkRvrE8xFsNszsGlrLo7Tseo7Nbqyv1BUO5t1TTMkcwOaSwAuBIGA31vWFdeWR2ndX3KzuBDaeciLPXGeLD/lIS3Rtsld9coNFLaQsdRqTUtDZaZwZJVSbpeRkMaOLneoAlZg1PWbNNnFQyxR6Uiv1wYxrqh9SGvxkZGS4EAkccNAHFdndGZcjfOIYJRZc1LNst1To6qulDFBpm9QZ3KcDAldjIbus4Fp6N4AYPTwWI1PHCuGRFsG1+ldLbHbDqGt0ZaLtNNHFHJzlPEHuLgTvFxY4k8F1aTahot0bmjZVZWkggECHh/8Aortt0zHJyu+InmxKi+uOXEgYyehbM2nSGmrrsvttELJbI7nX2cOiqRSsEvOCNp3t/Gc5IPtSd1JtyI+KKtZUXfNiWnob1tBihuNNHLSUMb56qOZgc0hvABwPA+URwPYuy8o+2WShh09UWa1UNAypjle7wanZFvjyCM7oGelLboiefryI3zMcmHkWUeTfa7ZddX10F0t1HXxNoS5rKmFsjQ7faMgOB48VK3jaLoq3Xest52WWWTwad8O/uQje3XEZxzXDoXZ3Y7XI357GGUXZ9oWo7RqOupqi0aapLDHDEWPipwzEhzneO61voXb+Tba7VdNQXZl2ttHXxRUYe1tTA2UNO+OIDgeK5WM/ktOGKUWSNumlKSzXmnvlliibZbqwSQiFoEcb8ZLQBwAIwQPT2KQ2t2m1UOzLRlZRWyipqmogaZ5oYGsfKeaacucBk8ePFcz0drtwrHSx2ZYnRdv2d6qsem2VjbxpKhv5nLDGagM+xYznG8x3TkdnQsyVVx0ZBsuh1x/s9sbmyEDwTwaHIzIWefzfdnoVTujPremN84a2Iu17Q9TWbUk9HJZ9L0dgbA1wkbThmJSSME7rG9GO/pXVFMKkREXXBERAREQFKD+p6fJ6Ja1+P1G//VWNHA+qqoqePzpHAejvVzfZ2TVxjh+0wNEUfoHX7UFgiIgKjVuw0N7VWVnUO3pT2DggpoiICIiAiIgKtSuxJu9RVFfQcHIQX6L4x28wOHWvqApGxTRtqX0s5xBVN5t/cfin1FRyIKlVC+nqJIJRh7HFpVNSty/r9uiuLeMseIqj0/Fd61FICIiAiIgIERIG0d7vItOutDxyuAgr6GWkkB6DvCPd/wC4D2rq98tjdn2z+7UTBuVF6vRhh7eYD+Hq3Qf8yx5tA2gy6plsc0FtNultLMMcKjnN8+Tg+aMeb3rltK2i1GtK21zPtwooqDJ5oT7/ADjyQS7O6MdAGOKrPSz2z9s59d6YjEY7I++Meu52PlSffjbPm8fvuUFsCsPjraHSyyM3qe3g1UmejLeDB/mIPqUdtS1r7+LzTXHxZ4v5inEO5z/O73lE5zutx0q52d6/GjbNdqWltHP1twbutqzUbnNANIb5O6c4JJ6RlRpzNImeO/8ApWp0sR3f2y86gsJ2rDWX+0myCQO5rwHnIvM3dzm97nenr83p6lirb9YfEu0OqmiZu09xaKqPHRk8Hj/MCfWsf5O9vZOc5yu7bQtfDWNjtNHVWjmK23t3TVio3udBaA7yd0YyQD0nC5Mboxw8JVE75zx8nZeS39+lx+b3fWMUjs0rabV1jv2zi7SAP35Z7dI7pYQ4nA9DuPoLl0DZbrT3j3qouXi3xhz1OYeb5/mt3Lgc53XdnQoS1XqrtepIr7QnmqiKo59gzw6clp7QQSCrnFpiJ6sY/KIzETMdec/hlnbPR1Fv2PaSoKuMx1FNI2KRp6nNjcD+xYq0ZZpNQapt1njB/rM7WPI6mdLj6mgldu2qbTvfzaaSgNk8X+Dz89v+Fc7veSRjG43HSoPZlqyDRuoHXiW0+MpBC6OJvhHNc2TjLs7rs8Mj1pSenNrc8/gtHQiteWGbtpdksd21DanO2gWiwvsrWiGjkMZcx4IcCcytI4BvDHUuocpm0wyzWnVdDJFPT1cXMSTRODmPI8pjgRwOQT7FibUN0nvV9rbtUjEtXO6VwzndyeA9Q4epdpfr8T7L26JrbR4QYnb0FZ4Tgx4fvDyN05xkjpHAqMTNI55z9+tcYi3Z1fpU2B1UFJtRtbp3BolEkTSfxnMIHtPD1rntzt9RQbUq+euhkNPVPZNG7HCRm6AQD3YIXQ4JZIJmTQyOjljcHMe04LSOIIPasrW/bVVS22Oj1Ppm3X4xjhJIQwk9pBa4Z9ACud+J4wiN2Y4S7JoiybJtZw3HxRpS4QOoog9zqmolAJIOMbsruzrWA3cHEd6yXqHbFeqy2PtditlDYKSQFrvBhl+D0gHAA9IGe9YzU/8AbKuGGyU1+oNObC9O3C42GmvcJjhj8GnLQ0Etd5XFrhkY7OtY+u+07TNbaqujh2Z2illnhfGydj496IuBAcMQjiM56QriybX6Cj0nb9PXHRdNdYaOJrMz1TS1xb8bdMRAK5/7WNKf+VVl/wA8X+iqv0rW5SmnRrXnDEa2Oqr14g0zsxr3O3YSYoZj1bj4Q0+zOfUtd6yVk9ZNOyJsLJJHPbG3oYCc4HoXbdX66OoNG2PTotfgptLWt5/wjf53DN3O7ujd7ekpFuj9Y/smvS+ksq6jtbNBWTXt+YBHJdZxBRHukbk49Be//Iuucoz/AIHo79Df+7Guq7QtpFZrDT1rtE1D4MKLDppOf3/CHhu6HY3Ru/GOMnpVvtG1z78KGzUvivwLxZCYt7wjnOcyGjON0Y83v6VExuxymPtESqJ3xM8pz3zj9O08lr79bj83n6xivb9tQ0xSXyvpZdmNnqZIamSN8znx5kIcQXH7CeJxnpK6Jst1p7x73UXLxb4w56nMPN8/zWPKBzndd2LuE21zTE0r5Zdltnkke4ue90sZLiekk8zxKu2/Zxy8013Z7/Jj/XV9odQ33xjb7HTWWHmms8GgLS3Izl3BrRk57FkTktffHef0EfvhdK2iartWpzRG2aVorD4Pv7/g7mnnd7GM7rG9GD29KqbLNb+8e5VtZ4s8YeFQCHd5/mt3jnOd12VNd0T9fMvvx9HeNmNVT630TddnlzkaKqnDp7bI/pbg5A/Vcf8AK49i+7dKeak2ZaLpKmMxzQRiORh6WubE0Ee0LFOnL3V2LUVLe6E7s9PLzgbng4dbT3EZHrXbdq20r39UNDTeJfF/gsrpN7wrnd7IxjG43CX313dc4z9OK4nF55b8fVj5Zyuv4LFH+ez/AOYKwau81W0Dn9lkOhvFG7zZB8L8JznEhf5m73485dtvpjthNd14nvdGREQEREBERARFVo6eSqqY6eIZe92B3d6CQtn9St09xdwkeDDB6T0u9QUUpC+TxvnZS05zT0zebZ3nrd6yo9AREQcZXbkZPsViq9W/LgwdXSqCAiIgIiICIiAiIgr0j8EsPX0K5VgCQQR0hXsbg9gcEHJERBfWaqZT1JZPxpphzcw7j1+pUbjSvoqt8D+OOLXdTm9RVupeD/els8GPGrpW5i7Xs62+kdSCIREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAUtTf7ttTqo8KmqBZD2tZ1u9atrRSNqZy+Y7tNCN+Z3YOz0lU7nVurat0xG63zWM6mtHQEFsiIgL49wa0uPUvqtqp+TuDoHSgouJJJPSV8REBERAREQEREBERAVamfuu3T0FUUQSCKnTyb7OPSOlVEBVKeaSnnZNE4tew5BVNEEpdoY6iFt0pW4jkOJmD+zf8AyKi1eWut8EmO+3nIJBuzR/jN/mvt1ovBJWvidzlNKN6GTtHZ6QgskREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBVKeGSonZDE0ue84ACpgEnA4lTB/3PRlv/wBoTt49sLD/ABKCndpo6aBtrpXAsjOZ3j+0f/IKLREBERBxleGMJ6+pWJOTkqpO/ffw6B0KmgIiICIiAiIgIiICIiAiIg5RuLHBwV61wc0OHQVYKrTybjsHzT9CC7REQFI2uriEbqGty6lkPT1xu/GCjkQXNwo5aKo5qTBBGWPHQ8dRCtlJ2+rhmp/F9eTzJP2KXrid/JWlfRzUU5imA7WuHEOHaCgt0REBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERARFLUlNFQQNr69gc93GCA9Lj+M7uQfaSJlrp211S0OqXjNNEer8s/wAFFzSPmldLK4ue45cT1lcquolqqh087y57jxP8FSQEREBUKmTA3B0npVSaQRt7z0KzJJOT0oPiIiAiIgIiICIiAiIgIiICIiAiIguKaX4jj6CrhR6uqeXe8l3T1d6CsiIgKSoK2J9OKC4Aup/iPHnRHtHd3KNRBdXGhlopAHEPieMxyt8147laq/t9w5mM0tTHz9I8+VGelp7WnqK+19vEcPhdHJ4RSE+cPOZ3OHUgj0REBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEAJOBxKqU8MtRM2KFjnvd0AKUzTWceSWVNw7elkJ/iUHyGnhtcbamvaJKkjMVMer8p38lG1dRNVTunneXvd0k/sXGaSSaV0sry97jkuJ4lcEBERAXF7gxu8V9c4NaSTwVnLIZHZPR1BB8e4vdvFcURAREQEREBERAREQEREBERAREQEREBERBdQTb3ku879qrKPVzBNnyX9PUUFdERAVxQ1lRRS85A/GeDmni1w7CFboglzTUlzG/QbsFV0upnHg78w/wUXLHJFIY5WOY9vAtcMELiCQQQcEdBUnFcoqmMQXWMzNAw2dv2xn8wgi0UhV2uRkRqaSRtXTfjs6W/nDpCj0BERAREQEREBERAREQEREBERAREQEREBERAREQERVaWmnqpRFTxOkeeoDo/kgpK9oLdNVMMz3Ngpm+dM/gB6O0q5EFvtvGreKypH9iw+Q0/lHr9Cs6+uqK14MzwGN8yNow1o7ggup7hFTROpbW10bDwfO77ZJ/IKLREBERAXxzg1uScBfHvaxuXFWkshkOT0dQQfZpDI7sHUFTREBERAREQEREBERAREQEREBERAREQEREBERAREQV4Z93yX8R2q5BBGQchR6qRSujPDiOxBeIuMb2vGWn1LkgIiIK1LUz0solp5XRvHWOv09qv/CrfX8K6HwaY/28I4H85v8AJRSIJCptNTHHz0BZVQf3kJ3sekdIUeqtNUT00nOQSvjd2tOFf+Mqaq4XKiY9x/tofIf6T1FBFopTxbTVPG318byeiKbyH+jsKs6uhq6Q4qKeSPvI4e3oQW6IiAiIgIiICIiAiIgIiICIiAi+ta57g1rS4noAGSpCGz1jmCScMpYvx53bv0dKCOValpqiqk5unhfI78kdHp7FfYs9H0mS4SjqHkRj+JVKqu1VNHzMZZTQf3cLd0evrKCt4DRUXlXGpD5B/wCzwHJ9Z6AqVVdZXxGnpY2UlP8AiR9J9J6So9EBERAREQFwllbGO09ipyz44M4ntVuSSck5KD697nuy4riiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiD6CQcg4KuIpweD+HerZEEgOIyEVlHI5h4Hh2K4jnY7gfJPegqoiICIiArykuddSjdhqXhn4rvKb7CrNEEp4ypJ/u22QuP48JMZ/kV85myz/aquopj2TR7w9oUYiCT8Tvk+5q2jn7A2XDvYVTls9zj86ilP5o3v2KwVWKoni+1Tys/NeQgSU9RH9sglZ+cwhUlfR3i5x+bWyn84737VV8e3I+fKyT86Jp/ggjEUn46quuGkd6YQnjqo6qajHogCCMXJrXOOGtLvQFI+O60eaKdvohavjr5dXcPCyB+Sxo/YEFvFb66XzKOdw7ebOFcNslwxvSxxwN/Glka3+Kt5bhXy+fWTkdnOHCtnOc45cST2koJLxdRRcam7QDuhaZPpX3nbLB9rpaiqcOuV+4PYFFogk3Xmqa0tpY4KRh6oYwCfX0qPmllmfvzSPkd2udkrgiAiIgIiICL49zWjLjhW8lQTwYMd6CvJI1g4nj2K1llc/h0DsVMkk5KICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIgIiICIiAiIg5slezoPDsKrsnYfO8kq1RBIAgjIOUVg1zmnLSQqzKhw84AoLlFTbNG7rx6VUBBGQcoCIiAiIgIiICIvpBHSCMoPiIiAiIgIiICIuLntb5zgEHJFRdUNHmglUXzPd14HcgunyMZ0n1Kg+oJ4MGO9UEQfSSTknJXxEQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAX0EjoJHoXxEFVs8g68+lc21P4zfYVboguxURnpyPUuQljPxwrJEF+HNPQ4H1r6o9EHcNAQUtRqSJlU1rgGOdG13QXDo/ifUu9a3gpZdN1T6lrMxs3o3EcWu6senoWGI5JI3tfHI5j2nIc04IKua663KuY1lZXVE7G9DXyEgd6CohIHSVYZPaviC+L2Dpe32riZox8bPqVmiC6NQ3qaSuDqhx6AAqCIObpHu6XFcERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREBERAREQEREH/9k="

import json

CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           ".takeoff_config.json")

def load_config():
    try:
        with open(CONFIG_FILE) as f:
            return json.load(f)
    except Exception:
        return {}

def save_config(data):
    try:
        existing = load_config()
        existing.update(data)
        with open(CONFIG_FILE, "w") as f:
            json.dump(existing, f, indent=2)
    except Exception:
        pass

# ══════════════════════════════════════════════════════════════════════════════
#  AUTO-UPDATER
# ══════════════════════════════════════════════════════════════════════════════

def _version_tuple(v):
    """Convert 'v1.2' or '1.2' to (1, 2) for comparison."""
    return tuple(int(x) for x in v.lstrip("v").split("."))


# Path used to pass changelog content to the newly launched exe after update
_UPDATE_FLAG_PATH = os.path.join(os.path.expanduser("~"), ".nepm_just_updated.txt")


def _show_post_update_changelog():
    """
    Called at startup. If a just-updated flag file exists, show the
    changelog in a dialog then delete the flag. Runs regardless of frozen state.
    """
    if not os.path.exists(_UPDATE_FLAG_PATH):
        return
    try:
        with open(_UPDATE_FLAG_PATH, encoding="utf-8") as f:
            content = f.read().strip()
        os.remove(_UPDATE_FLAG_PATH)

        lines     = content.split("\n", 1)
        version   = lines[0].strip() if lines else VERSION
        changelog = lines[1].strip() if len(lines) > 1 else ""

        _root = tk.Tk()
        _root.withdraw()
        msg = f"Successfully updated to {version}!\n\n"
        if changelog:
            msg += f"What's new in this version:\n\n{changelog}"
        messagebox.showinfo(
            f"Updated to {version} — NEPM Takeoff Wizard", msg
        )
        _root.destroy()
    except Exception:
        try:
            os.remove(_UPDATE_FLAG_PATH)
        except Exception:
            pass


def check_for_updates():
    """
    Checks GitHub for a newer version on startup. If one is found the update
    is FORCED — no option to skip. Changelog is fetched and shown after the
    new exe launches. Fails silently on any network or parsing error.
    Only performs the version check when packaged as a PyInstaller exe.
    """
    # Always check for the post-update flag first (runs as script or exe)
    _show_post_update_changelog()

    if not getattr(sys, "frozen", False):
        return   # Running as .py script — skip version check

    import urllib.request
    import threading
    import tempfile
    import subprocess

    try:
        # ── Version check ────────────────────────────────────────────────────
        with urllib.request.urlopen(GITHUB_VERSION_URL, timeout=5) as r:
            latest = r.read().decode().strip()

        if _version_tuple(latest) <= _version_tuple(VERSION):
            return   # Already up to date

        # ── Fetch changelog ──────────────────────────────────────────────────
        changelog = ""
        try:
            with urllib.request.urlopen(GITHUB_CHANGELOG_URL, timeout=5) as r:
                changelog = r.read().decode().strip()
        except Exception:
            pass   # Changelog unavailable — continue without it

        # ── Forced update notice (no skip option) ────────────────────────────
        _root = tk.Tk()
        _root.withdraw()
        msg = (
            f"A required update is available.\n\n"
            f"   Current : {VERSION}\n"
            f"   Latest  : v{latest}\n\n"
        )
        if changelog:
            msg += f"What's new:\n\n{changelog}\n\n"
        msg += "Click OK to install the update.\nThe wizard will restart automatically."
        messagebox.showinfo(
            "Required Update — NEPM Takeoff Wizard", msg
        )
        _root.destroy()

        # ── Download with progress window ────────────────────────────────────
        exe_path     = sys.executable
        new_exe_path = exe_path + ".update"

        prog = tk.Tk()
        prog.title("Downloading update…")
        prog.geometry("380x120")
        prog.resizable(False, False)
        tk.Label(
            prog,
            text=f"Downloading NEPM Takeoff Wizard v{latest}…",
            font=("Arial", 10), pady=16
        ).pack()
        try:
            from tkinter import ttk
            bar = ttk.Progressbar(prog, length=320, mode="indeterminate")
            bar.pack()
            bar.start(12)
        except Exception:
            pass
        tk.Label(
            prog,
            text="Please wait — do not close this window.",
            font=("Arial", 8), fg="#888888", pady=8
        ).pack()

        done  = [False]
        error = [None]

        def _download():
            try:
                urllib.request.urlretrieve(GITHUB_EXE_URL, new_exe_path)
                done[0] = True
            except Exception as exc:
                error[0] = str(exc)
            finally:
                try:
                    prog.destroy()
                except Exception:
                    pass

        threading.Thread(target=_download, daemon=True).start()
        prog.mainloop()

        if error[0]:
            raise RuntimeError(error[0])

        if not done[0]:
            # Window was closed — force a re-check next launch
            return

        # ── Write post-update flag so new exe shows changelog ────────────────
        try:
            with open(_UPDATE_FLAG_PATH, "w", encoding="utf-8") as f:
                f.write(f"v{latest}\n{changelog}")
        except Exception:
            pass

        # ── Swap exe via a tiny batch file ───────────────────────────────────
        bat = (
            "@echo off\n"
            "timeout /t 2 /nobreak > nul\n"
            f'move /y "{new_exe_path}" "{exe_path}"\n'
            f'start "" "{exe_path}"\n'
            "del \"%~f0\"\n"
        )
        bat_path = os.path.join(tempfile.gettempdir(), "nepm_wizard_update.bat")
        with open(bat_path, "w") as f:
            f.write(bat)

        subprocess.Popen(
            ["cmd", "/c", bat_path],
            creationflags=subprocess.CREATE_NO_WINDOW
        )
        sys.exit(0)   # Close current version — batch relaunches new one

    except Exception:
        pass   # Always fail silently — never block the app


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL PALETTE & STYLE HELPERS
# ══════════════════════════════════════════════════════════════════════════════
FONT_NAME = "Aptos"

C_DARK_BLUE  = "0D0D0D"
C_MID_BLUE   = "3D3D3D"
C_LIGHT_BLUE = "DCDCDC"
C_AMBER      = "888888"
C_WHITE      = "FFFFFF"
C_LT_GREY    = "F2F2F2"
C_DARK_GREY  = "595959"
C_GREEN      = "E2EFDA"
C_YELLOW     = "FFF2CC"
C_SOFT_GREEN = "EAF4EA"

def fill(c):
    return PatternFill("solid", fgColor=c)

def xfont(bold=False, c="000000", sz=10, italic=False):
    return Font(name=FONT_NAME, bold=bold, color=c, size=sz, italic=italic)

def xborder():
    t = Side(style="thin", color="BFBFBF")
    return Border(left=t, right=t, top=t, bottom=t)

def align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def hdr_cell(ws, row, col, text, bg=C_DARK_BLUE, fg=C_WHITE, sz=10, bold=True):
    c = ws.cell(row, col, text)
    c.font      = xfont(bold=bold, c=fg, sz=sz)
    c.fill      = fill(bg)
    c.alignment = align()
    c.border    = xborder()
    return c

def data_cell(ws, row, col, value=None, bg=C_WHITE, fmt=None,
              bold=False, fg="000000", h="left"):
    c = ws.cell(row, col)
    if value is not None:
        c.value = value
    c.font      = xfont(bold=bold, c=fg)
    c.fill      = fill(bg)
    c.alignment = align(h=h, wrap=False)
    c.border    = xborder()
    if fmt:
        c.number_format = fmt
    return c

def row_bg(r):
    return C_WHITE if r % 2 == 1 else C_LT_GREY

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def title_block(ws, job_name, sheet_label, n_cols=16):
    """
    Branded title block  (NEPM monochrome theme, v1.5):
      Row 1 — mega header (height 130pt): NEPM badge logo (A1:B1)
               job name text (C1:last, large white, centred)
      Row 2 — meta bar: date (left)  |  prepared by (right)
      Rows 3–5 — spacers (5pt each)
    sheet_label is accepted for API compatibility but not displayed;
    the tab name identifies the sheet type.
    Data always starts at row 7 — freeze_panes and DATA_START unchanged.
    """
    from datetime import date as _date

    last     = get_column_letter(n_cols)
    mid      = get_column_letter(max(1, n_cols // 2))
    nxt      = get_column_letter(max(2, n_cols // 2 + 1))
    LOGO_COLS = min(2, n_cols - 1)   # cols reserved for logo (A:B, or just A if narrow)

    # ── Row 1 — mega header ───────────────────────────────────────────────────
    ws.row_dimensions[1].height = 152

    # Logo area
    logo_end = get_column_letter(LOGO_COLS)
    ws.merge_cells(f"A1:{logo_end}1")
    lc = ws["A1"]
    lc.value     = ""
    lc.fill      = fill("0D0D0D")
    lc.alignment = align(h="center", v="center")
    try:
        from openpyxl.drawing.image import Image as _XLImg
        import io as _io
        _raw    = base64.b64decode(LOGO_B64)
        _xl_img = _XLImg(_io.BytesIO(_raw))
        _xl_img.width  = 197
        _xl_img.height = 197
        ws.add_image(_xl_img, "A1")
    except Exception:
        lc.value = "N"
        lc.font  = Font(name=FONT_NAME, bold=True, size=28, color=C_WHITE)

    # Job name area
    if n_cols > LOGO_COLS:
        job_start = get_column_letter(LOGO_COLS + 1)
        ws.merge_cells(f"{job_start}1:{last}1")
        jc = ws[f"{job_start}1"]
    else:
        jc = lc   # fallback for very narrow sheets
    jc.value     = job_name
    jc.font      = Font(name=FONT_NAME, bold=True, size=20, color=C_WHITE)
    jc.fill      = fill("0D0D0D")
    jc.alignment = align(h="center", v="center", wrap=False)

    # ── Row 2 — meta bar ──────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 18
    ws.merge_cells(f"A2:{mid}2")
    c = ws["A2"]
    c.value     = f"Date:  {_date.today().strftime('%d/%m/%Y')}"
    c.font      = xfont(italic=True, sz=9, c=C_DARK_GREY)
    c.fill      = fill(C_LT_GREY)
    c.alignment = align(h="left")

    ws.merge_cells(f"{nxt}2:{last}2")
    c = ws[f"{nxt}2"]
    _name   = get_display_name()
    c.value = (f"Prepared by:  {_name}" if _name
               else "Prepared by:  _________________________")
    c.font      = xfont(italic=True, sz=9, c=C_DARK_GREY)
    c.fill      = fill(C_LT_GREY)
    c.alignment = align(h="right")

    # ── Rows 3–5 — spacers ────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 5
    ws.row_dimensions[4].height = 5
    ws.row_dimensions[5].height = 5

def add_legend(ws, start_row, n_cols=12):
    last = get_column_letter(n_cols)
    ws.merge_cells(f"A{start_row}:{last}{start_row}")
    c = ws.cell(start_row, 1, "LEGEND")
    c.font      = xfont(bold=True, c=C_WHITE)
    c.fill      = fill(C_DARK_GREY)
    c.alignment = align()
    ws.row_dimensions[start_row].height = 18
    items = [
        (C_WHITE,      "White cell   —  manual input"),
        (C_LIGHT_BLUE, "Blue cell    —  auto-calculated formula (do not edit)"),
        (C_YELLOW,     "Yellow cell  —  required dropdown selection"),
        (C_SOFT_GREEN, "Green cell   —  dynamic calculator output (do not edit)"),
    ]
    for i, (colour, label) in enumerate(items):
        r = start_row + 1 + i
        ws.row_dimensions[r].height = 16
        ws.cell(r, 1).fill   = fill(colour)
        ws.cell(r, 1).border = xborder()
        ws.merge_cells(f"B{r}:{last}{r}")
        c2 = ws.cell(r, 2, label)
        c2.font      = xfont(sz=9)
        c2.fill      = fill(C_LT_GREY)
        c2.alignment = align(h="left")


# ══════════════════════════════════════════════════════════════════════════════
#  SHEET BUILDERS
# ══════════════════════════════════════════════════════════════════════════════
DATA_START = 7
DATA_ROWS  = 50

# Named ranges on Rates sheet that takeoff sheets reference
# These are the rows where bricks/blocks-per-m2 live for each type
# We store them as absolute references: 'Rates & Assumptions'!$D$row
RATES_SHEET = "Rates & Assumptions"


def build_brickwork(wb, job_name, brick_types, brick_rates_rows):
    """
    brick_rates_rows: list of Excel row numbers on the Rates sheet
                      where each brick type's bricks-per-m2 calc lives (col D)
    Columns (14 total):
      1=Desc  2=BrickType  3=LM  4=Height  5=M2  6=Ded  7=OpenCount
      8=OpenWidth  9=TotM2  10=#Brk  11=Wastage  12=Rate  13=Total  14=Notes
    """
    ws = wb.create_sheet("Brickwork")
    ws.sheet_properties.tabColor = C_AMBER
    ws.freeze_panes = "A7"

    n_cols = 14
    title_block(ws, job_name, "BRICKWORK TAKEOFF", n_cols)
    set_col_widths(ws, [30, 28, 7, 9, 8, 13, 10, 12, 10, 12, 12, 10, 12, 22])

    HEADERS = ["Description", "Brick Type", "LM", "Height\n(m)",
               "M2", "Deductions\nM2", "Opening\nCount", "Opening Width\n(LM)",
               "Total M2", "No. of\nBricks", "Wastage\n(+3%)",
               "Rate\n($ / brick)", "Total ($)", "Notes"]
    ws.row_dimensions[6].height = 46
    for ci, h in enumerate(HEADERS, 1):
        bg = C_AMBER if ci in (7, 8) else C_DARK_BLUE
        hdr_cell(ws, 6, ci, h, bg=bg)

    if brick_types:
        brick_ref = write_brick_list(wb, brick_types)
        dv = DataValidation(type="list", formula1=brick_ref, allow_blank=True,
                            showErrorMessage=False)
    else:
        dv = DataValidation(type="list", formula1='"Type"', allow_blank=True,
                            showErrorMessage=False)
    ws.add_data_validation(dv)
    dv.sqref = f"B{DATA_START}:B{DATA_START+DATA_ROWS-1}"

    for r in range(DATA_START, DATA_START + DATA_ROWS):
        bg = row_bg(r); rs = str(r)

        data_cell(ws, r,  1, bg=bg)                       # Description
        data_cell(ws, r,  2, bg=bg)                       # Brick Type
        data_cell(ws, r,  3, bg=bg, fmt="0.00")           # LM
        data_cell(ws, r,  4, bg=bg, fmt="0.00")           # Height

        # M2 = LM × Height
        c = data_cell(ws, r, 5, bg=C_LIGHT_BLUE, fmt="0.00")
        c.value = f'=IF(OR(C{rs}="",D{rs}=""),"",C{rs}*D{rs})'

        data_cell(ws, r,  6, bg=bg, fmt="0.00")           # Deductions M2
        data_cell(ws, r,  7, bg=bg, fmt="#,##0")          # Opening Count (NEW)
        data_cell(ws, r,  8, bg=bg, fmt="0.00")           # Opening Width LM (NEW)

        # Total M2 = M2 – Deductions
        c = data_cell(ws, r, 9, bg=C_LIGHT_BLUE, fmt="0.00")
        c.value = f'=IF(E{rs}="","",E{rs}-F{rs})'

        # No. of Bricks — looks up bricks/m2 from Rates sheet via MATCH on brick type
        c = data_cell(ws, r, 10, bg=C_LIGHT_BLUE, fmt="#,##0")
        c.value = (
            f'=IF(OR(I{rs}="",B{rs}=""),"",ROUND(I{rs}*'
            f'IFERROR(INDEX(\'{RATES_SHEET}\'!$E$3:$E$200,'
            f'MATCH(B{rs},\'{RATES_SHEET}\'!$A$3:$A$200,0)),50),0))'
        )

        # Wastage +3%
        c = data_cell(ws, r, 11, bg=C_LIGHT_BLUE, fmt="#,##0")
        c.value = f'=IF(J{rs}="","",ROUND(J{rs}*1.03,0))'

        data_cell(ws, r, 12, bg=bg, fmt="$#,##0.0000")   # Rate ($ per brick)

        # Total = Wastage qty × Rate per brick
        c = data_cell(ws, r, 13, bg=C_LIGHT_BLUE, fmt="$#,##0.00")
        c.value = f'=IF(K{rs}="","",K{rs}*L{rs})'

        data_cell(ws, r, 14, bg=bg)                       # Notes
        ws.row_dimensions[r].height = 18

    T = DATA_START + DATA_ROWS
    ws.row_dimensions[T].height = 22
    for ci in range(1, n_cols + 1):
        ws.cell(T, ci).fill = fill(C_DARK_BLUE)
    c = ws.cell(T, 1, "TOTAL")
    c.font = xfont(bold=True, c=C_WHITE, sz=11)
    c.alignment = align(h="right")
    for col, fmt in {7: "#,##0", 8: "0.00", 9: "0.00",
                     10: "#,##0", 11: "#,##0", 13: "$#,##0.00"}.items():
        cl = get_column_letter(col)
        c = ws.cell(T, col, f"=SUM({cl}{DATA_START}:{cl}{T-1})")
        c.font = xfont(bold=True, c=C_WHITE)
        c.fill = fill(C_DARK_BLUE)
        c.number_format = fmt

    ws.auto_filter.ref = f"A6:{get_column_letter(n_cols)}{T-1}"
    add_legend(ws, T + 2, n_cols)
    return ws


def build_blockwork(wb, job_name, block_data, block_rates_rows):
    """
    block_data: list of (name, size, corefill) tuples from wizard.
    Columns (20 total):
      1=Desc 2=BlkType 3=Size 4=LM 5=Ht 6=M2 7=Ded
      8=OpenCount 9=OpenWidth 10=TotM2 11=#Blk 12=Wast
      13=CF? 14=CorefM3 15=Steel 16=SBar 17=Cap 18=Rate 19=Total 20=Notes
    """
    block_types = [d[0] for d in block_data]
    ws = wb.create_sheet("Blockwork")
    ws.sheet_properties.tabColor = C_DARK_BLUE
    ws.freeze_panes = "A7"

    n_cols = 20
    title_block(ws, job_name, "BLOCKWORK TAKEOFF", n_cols)
    set_col_widths(ws, [26, 26, 10, 6, 9, 8, 13, 10, 12, 10, 10, 10,
                        11, 10, 10, 10, 14, 11, 12, 22])

    HEADERS = ["Description", "Block Type", "Block\nSize",
               "LM", "Height\n(m)", "M2", "Deductions\nM2",
               "Opening\nCount", "Opening Width\n(LM)",
               "Total M2", "No. of\nBlocks", "Wastage\n(+3%)",
               "Corefilled?\n(Yes / No)", "M3\nCorefill",
               "Steel\n(kg)", "Starter\nBars", "Capping\nBlocks",
               "Rate\n($ / block)", "Total ($)", "Notes"]
    ws.row_dimensions[6].height = 52
    for ci, h in enumerate(HEADERS, 1):
        if ci in (8, 9):
            bg = C_AMBER
        elif ci == 13:
            bg = C_MID_BLUE
        else:
            bg = C_DARK_BLUE
        hdr_cell(ws, 6, ci, h, bg=bg)

    if block_data:
        block_ref = write_block_lists(wb, block_data)
        dv_blk = DataValidation(type="list", formula1=block_ref,
                                allow_blank=True, showErrorMessage=False)
    else:
        dv_blk = DataValidation(type="list", formula1='"Type"',
                                allow_blank=True, showErrorMessage=False)
    dv_sz  = DataValidation(type="list",
                            formula1='"90mm,140mm,190mm,290mm"',
                            allow_blank=True, showErrorMessage=False)
    dv_cf  = DataValidation(type="list", formula1='"Yes,No"',
                            allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_blk)
    ws.add_data_validation(dv_sz)
    ws.add_data_validation(dv_cf)
    dv_blk.sqref = f"B{DATA_START}:B{DATA_START+DATA_ROWS-1}"
    dv_sz.sqref  = f"C{DATA_START}:C{DATA_START+DATA_ROWS-1}"
    dv_cf.sqref  = f"M{DATA_START}:M{DATA_START+DATA_ROWS-1}"

    for r in range(DATA_START, DATA_START + DATA_ROWS):
        bg = row_bg(r); rs = str(r)

        data_cell(ws, r,  1, bg=bg)                      # Description
        data_cell(ws, r,  2, bg=bg)                      # Block Type

        # Col 3: Block Size — auto-populated via VLOOKUP; manually overrideable
        c = data_cell(ws, r, 3, bg=C_LIGHT_BLUE)
        c.value = f'=IFERROR(VLOOKUP(B{rs},\'_ListsBlockwork\'!$A:$B,2,0),"")'
        c.alignment = align(h="center", wrap=False)

        data_cell(ws, r,  4, bg=bg, fmt="0.00")          # LM
        data_cell(ws, r,  5, bg=bg, fmt="0.00")          # Height

        # M2 = LM × Height
        c = data_cell(ws, r, 6, bg=C_LIGHT_BLUE, fmt="0.00")
        c.value = f'=IF(OR(D{rs}="",E{rs}=""),"",D{rs}*E{rs})'

        data_cell(ws, r,  7, bg=bg, fmt="0.00")          # Deductions M2
        data_cell(ws, r,  8, bg=bg, fmt="#,##0")         # Opening Count (NEW)
        data_cell(ws, r,  9, bg=bg, fmt="0.00")          # Opening Width LM (NEW)

        # Total M2 = M2 – Deductions
        c = data_cell(ws, r, 10, bg=C_LIGHT_BLUE, fmt="0.00")
        c.value = f'=IF(F{rs}="","",F{rs}-G{rs})'

        # No. of Blocks
        c = data_cell(ws, r, 11, bg=C_LIGHT_BLUE, fmt="#,##0")
        c.value = f'=IF(J{rs}="","",ROUND(J{rs}*12.5,0))'

        # Wastage +3%
        c = data_cell(ws, r, 12, bg=C_LIGHT_BLUE, fmt="#,##0")
        c.value = f'=IF(K{rs}="","",ROUND(K{rs}*1.03,0))'

        # Corefilled? — auto-populated from wizard; manually overrideable
        c = data_cell(ws, r, 13, bg=C_LIGHT_BLUE)
        c.value = f'=IFERROR(VLOOKUP(B{rs},\'_ListsBlockwork\'!$A:$C,3,0),"")'
        c.alignment = align(h="center", wrap=False)

        # M3 Corefill (references col 13=M for Corefilled, col 11=K for blocks)
        c = data_cell(ws, r, 14, bg=C_LIGHT_BLUE, fmt="0.00")
        c.value = (
            f'=IF(M{rs}<>"Yes","",'
            f'IF(C{rs}="290mm",ROUND(K{rs}/80,2),'
            f'IF(C{rs}="190mm",ROUND(K{rs}/110,2),'
            f'IF(C{rs}="140mm",ROUND(K{rs}/130,2),"N/A (90mm)"))))'
        )

        # Steel: 10 kg/m² corefilled (references col 13=M, col 10=J for Total M2)
        c = data_cell(ws, r, 15, bg=C_LIGHT_BLUE, fmt="#,##0.0")
        c.value = f'=IF(M{rs}<>"Yes","",IF(C{rs}="90mm","N/A",ROUND(J{rs}*10,1)))'

        # Starter bars: 2.5 per LM
        c = data_cell(ws, r, 16, bg=C_LIGHT_BLUE, fmt="#,##0")
        c.value = f'=IF(D{rs}="","",ROUND(D{rs}*2.5,0))'

        # Capping blocks: 2.5 per LM
        c = data_cell(ws, r, 17, bg=C_LIGHT_BLUE, fmt="#,##0.0")
        c.value = f'=IF(D{rs}="","",ROUND(D{rs}*2.5,1))'

        data_cell(ws, r, 18, bg=bg, fmt="$#,##0.0000")  # Rate ($ per block)

        # Total = Wastage qty × Rate per block (col 12=L, col 18=R)
        c = data_cell(ws, r, 19, bg=C_LIGHT_BLUE, fmt="$#,##0.00")
        c.value = f'=IF(L{rs}="","",L{rs}*R{rs})'

        data_cell(ws, r, 20, bg=bg)                      # Notes
        ws.row_dimensions[r].height = 18

    T = DATA_START + DATA_ROWS
    ws.row_dimensions[T].height = 22
    for ci in range(1, n_cols + 1):
        ws.cell(T, ci).fill = fill(C_DARK_BLUE)
    c = ws.cell(T, 1, "TOTAL")
    c.font = xfont(bold=True, c=C_WHITE, sz=11)
    c.alignment = align(h="right")
    sums = {8: "#,##0", 9: "0.00", 10: "0.00", 11: "#,##0", 12: "#,##0",
            14: "0.00", 15: "#,##0.0", 16: "#,##0",
            17: "#,##0.0", 19: "$#,##0.00"}
    for col, fmt in sums.items():
        cl = get_column_letter(col)
        c = ws.cell(T, col, f"=SUM({cl}{DATA_START}:{cl}{T-1})")
        c.font = xfont(bold=True, c=C_WHITE)
        c.fill = fill(C_DARK_BLUE)
        c.number_format = fmt

    ws.auto_filter.ref = f"A6:{get_column_letter(n_cols)}{T-1}"
    add_legend(ws, T + 2, n_cols)

    nr = T + 8
    ws.merge_cells(f"A{nr}:{get_column_letter(n_cols)}{nr}")
    note = ("NOTE — Corefill M3 & Steel (kg) only calculate when 'Corefilled?' = Yes.  "
            "90mm blocks cannot be corefilled (returns N/A).  "
            "Starter bars: 2.5 per LM regardless of corefill.  "
            "Capping blocks: 2.5 per LM assumes 400mm length — use 5.0 per LM if 200mm.  "
            "Block Size and Corefilled? are auto-populated from wizard selections — "
            "override per row by typing directly into those cells if needed.  "
            "Opening Count and Opening Width (LM) are for lintel reference only — "
            "they do not affect M2 or quantity calculations.")
    c = ws.cell(nr, 1, note)
    c.font      = xfont(sz=9, italic=True, c=C_DARK_GREY)
    c.fill      = fill("FFF9E6")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[nr].height = 56
    return ws


def build_demo(wb, job_name):
    ws = wb.create_sheet("Demo Takeoff")
    ws.sheet_properties.tabColor = C_MID_BLUE
    ws.freeze_panes = "A7"

    n_cols = 10
    title_block(ws, job_name, "DEMOLITION TAKEOFF", n_cols)
    set_col_widths(ws, [35, 8, 9, 8, 14, 10, 8, 9, 11, 28])

    HEADERS = ["Description", "LM", "Height\n(m)", "M2",
               "Deductions\nM2", "Total M2", "QTY",
               "Rate ($)", "Total ($)", "Notes"]
    ws.row_dimensions[6].height = 44
    for ci, h in enumerate(HEADERS, 1):
        hdr_cell(ws, 6, ci, h)

    for r in range(DATA_START, DATA_START + DATA_ROWS):
        bg = row_bg(r); rs = str(r)
        data_cell(ws, r, 1, bg=bg)
        data_cell(ws, r, 2, bg=bg, fmt="0.00")    # LM
        data_cell(ws, r, 3, bg=bg, fmt="0.00")    # Height

        # M2 = LM × Height
        c = data_cell(ws, r, 4, bg=C_LIGHT_BLUE, fmt="0.00")
        c.value = f'=IF(OR(B{rs}="",C{rs}=""),"",B{rs}*C{rs})'

        data_cell(ws, r, 5, bg=bg, fmt="0.00")    # Deductions M2

        # Total M2 = M2 – Deductions
        c = data_cell(ws, r, 6, bg=C_LIGHT_BLUE, fmt="0.00")
        c.value = f'=IF(D{rs}="","",D{rs}-E{rs})'

        data_cell(ws, r, 7, bg=bg, fmt="#,##0")   # QTY
        data_cell(ws, r, 8, bg=bg, fmt="$#,##0.00") # Rate

        # Total = QTY × Rate
        c = data_cell(ws, r, 9, bg=C_LIGHT_BLUE, fmt="$#,##0.00")
        c.value = f'=IF(G{rs}="","",G{rs}*H{rs})'

        data_cell(ws, r, 10, bg=bg)
        ws.row_dimensions[r].height = 18

    T = DATA_START + DATA_ROWS
    ws.row_dimensions[T].height = 22
    for ci in range(1, n_cols + 1):
        ws.cell(T, ci).fill = fill(C_DARK_BLUE)
    c = ws.cell(T, 1, "TOTAL")
    c.font = xfont(bold=True, c=C_WHITE, sz=11)
    c.alignment = align(h="right")
    for col, fmt in {6: "0.00", 9: "$#,##0.00"}.items():
        cl = get_column_letter(col)
        c = ws.cell(T, col, f"=SUM({cl}{DATA_START}:{cl}{T-1})")
        c.font = xfont(bold=True, c=C_WHITE)
        c.fill = fill(C_DARK_BLUE)
        c.number_format = fmt

    ws.auto_filter.ref = f"A6:{get_column_letter(n_cols)}{T-1}"
    add_legend(ws, T + 2, n_cols)
    return ws


def get_hidden_sheet(wb, name):
    """Return a hidden sheet by name, creating it if needed."""
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(name)
    ws.sheet_state = "hidden"
    return ws


def write_brick_list(wb, brick_types):
    """Write brick types to _ListsBrickwork col A. Returns range ref."""
    ws = get_hidden_sheet(wb, "_ListsBrickwork")
    for i, v in enumerate(brick_types, 1):
        ws.cell(i, 1, str(v))
    return f"'_ListsBrickwork'!$A$1:$A${len(brick_types)}"


def write_block_lists(wb, block_data):
    """
    Write block data to _ListsBlockwork.
    block_data: list of (name, size, corefill) tuples.
    Col A=name, Col B=size, Col C=corefill.
    Returns range ref for name column (for dropdown).
    """
    ws = get_hidden_sheet(wb, "_ListsBlockwork")
    for i, (name, size, corefill) in enumerate(block_data, 1):
        ws.cell(i, 1, str(name))
        ws.cell(i, 2, str(size))
        ws.cell(i, 3, str(corefill))
    return f"'_ListsBlockwork'!$A$1:$A${len(block_data)}"



def build_george_summary(wb, job_name, scope, brick_types, block_types):
    ws = wb.create_sheet("George Summary")
    ws.sheet_properties.tabColor = C_GREEN
    title_block(ws, job_name, "GEORGE SUMMARY", 14)
    set_col_widths(ws, [35, 11, 11, 11, 11, 11, 16, 11, 12, 16, 16, 11, 12, 20])

    r = 7
    brick_subtotal_row = None
    block_subtotal_row = None
    DS, DE = DATA_START, DATA_START + DATA_ROWS - 1

    # ── BRICKWORK ──────────────────────────────────────────────────────────
    if scope in ("brickwork", "both"):
        ws.merge_cells(f"A{r}:N{r}")
        c = ws.cell(r, 1, "BRICKWORK")
        c.font = Font(name=FONT_NAME, bold=True, size=13, color=C_DARK_BLUE)
        c.fill = fill(C_AMBER); c.alignment = align()
        ws.row_dimensions[r].height = 26; r += 1

        for ci, h in enumerate(["Brick Type", "Total M2", "No. of Bricks",
                                 "Wastage (+3%)", "Rate ($/brick)",
                                 "Total ($)", "Notes",
                                 "Total\nOpenings", "Opening Width\n(LM)",
                                 "M2 Check\n(Bricks÷50)", "Wastage Check\n(Blocks×1.03)"], 1):
            if ci in (10, 11):
                bg = "2E6B3E"
            elif ci in (8, 9):
                bg = C_AMBER
            else:
                bg = C_MID_BLUE
            hdr_cell(ws, r, ci, h, bg=bg)
        ws.row_dimensions[r].height = 30; r += 1

        sub_start = r
        for bt in brick_types:
            if not bt.strip(): continue
            ws.cell(r, 1, bt).border = xborder()
            ws.cell(r, 1).fill = fill(row_bg(r))
            ws.cell(r, 1).font = xfont()
            ws.cell(r, 1).alignment = align(h="left")
            # Brickwork cols after v1.2: I=TotM2(9), J=Bricks(10), K=Wastage(11)
            #                            G=OpenCount(7), H=OpenWidth(8)
            ws.cell(r, 2).value = f'=IFERROR(SUMIF(Brickwork!B{DS}:B{DE},A{r},Brickwork!I{DS}:I{DE}),0)'
            ws.cell(r, 3).value = f'=IFERROR(SUMIF(Brickwork!B{DS}:B{DE},A{r},Brickwork!J{DS}:J{DE}),0)'
            ws.cell(r, 4).value = f'=IFERROR(SUMIF(Brickwork!B{DS}:B{DE},A{r},Brickwork!K{DS}:K{DE}),0)'
            ws.cell(r, 6).value = f'=IF(E{r}="","",D{r}*E{r})'
            # Opening totals (cols 8 and 9 in Brickwork = G and H)
            ws.cell(r, 8).value = f'=IFERROR(SUMIF(Brickwork!B{DS}:B{DE},A{r},Brickwork!G{DS}:G{DE}),0)'
            ws.cell(r, 9).value = f'=IFERROR(SUMIF(Brickwork!B{DS}:B{DE},A{r},Brickwork!H{DS}:H{DE}),0)'
            fmts = {2:"0.00", 3:"#,##0", 4:"#,##0", 6:"$#,##0.00", 8:"#,##0", 9:"0.00"}
            for col, fmt in fmts.items():
                c = ws.cell(r, col)
                c.number_format = fmt
                c.fill = fill(row_bg(r)); c.border = xborder()
                c.font = xfont(); c.alignment = align(h="center", wrap=False)
            for col in [5, 7]:
                c = ws.cell(r, col)
                c.fill = fill(row_bg(r)); c.border = xborder(); c.font = xfont()
            # ── Sanity check cols 10 & 11 ─────────────────────────────────
            c = ws.cell(r, 10)
            c.value = (f'=IF(C{r}=0,"—",IF(ABS(ROUND(C{r}/50,2)-B{r})<=1,'
                       f'"✅","❌  "&TEXT(ABS(ROUND(C{r}/50,2)-B{r}),"0.00")&" m2 off"))')
            c.fill = fill(row_bg(r)); c.border = xborder()
            c.font = xfont(sz=9); c.alignment = align(h="center", wrap=False)
            c = ws.cell(r, 11)
            c.value = (f'=IF(D{r}=0,"—",IF(ABS(ROUND(D{r}/1.03,0)-C{r})<=1,'
                       f'"✅","❌  "&TEXT(ABS(ROUND(D{r}/1.03,0)-C{r}),"0")&" bricks off"))')
            c.fill = fill(row_bg(r)); c.border = xborder()
            c.font = xfont(sz=9); c.alignment = align(h="center", wrap=False)
            ws.row_dimensions[r].height = 18; r += 1

        sub_end = r - 1
        ws.merge_cells(f"A{r}:E{r}")
        c = ws.cell(r, 1, "Brickwork Subtotal")
        c.font = xfont(bold=True, c=C_WHITE); c.fill = fill(C_MID_BLUE)
        c.alignment = align(h="right")
        c = ws.cell(r, 6, f"=SUM(F{sub_start}:F{sub_end})")
        c.font = xfont(bold=True, c=C_WHITE); c.fill = fill(C_MID_BLUE)
        c.number_format = "$#,##0.00"
        for ci in range(2, 15): ws.cell(r, ci).fill = fill(C_MID_BLUE)
        ws.row_dimensions[r].height = 22
        brick_subtotal_row = r; r += 2

    # ── BLOCKWORK ──────────────────────────────────────────────────────────
    if scope in ("blockwork", "both"):
        ws.merge_cells(f"A{r}:N{r}")
        c = ws.cell(r, 1, "BLOCKWORK")
        c.font = Font(name=FONT_NAME, bold=True, size=13, color=C_WHITE)
        c.fill = fill(C_DARK_BLUE); c.alignment = align()
        ws.row_dimensions[r].height = 26; r += 1

        for ci, h in enumerate(["Block Type", "Total M2", "No. of Blocks",
                                 "Wastage (+3%)", "M3 Corefill", "Steel (kg)",
                                 "Starter Bars", "Rate ($/block)",
                                 "Total ($)", "Notes",
                                 "Total\nOpenings", "Opening Width\n(LM)",
                                 "M2 Check\n(Blocks÷12.5)", "Wastage Check\n(Blocks×1.03)"], 1):
            if ci in (13, 14):
                bg = "2E6B3E"
            elif ci in (11, 12):
                bg = C_AMBER
            else:
                bg = C_DARK_BLUE
            hdr_cell(ws, r, ci, h, bg=bg)
        ws.row_dimensions[r].height = 30; r += 1

        blk_sub_start = r
        for blk in block_types:
            if not blk.strip(): continue
            ws.cell(r, 1, blk).border = xborder()
            ws.cell(r, 1).fill = fill(row_bg(r))
            ws.cell(r, 1).font = xfont()
            ws.cell(r, 1).alignment = align(h="left")
            # Blockwork cols after v1.2:
            # J=TotM2(10), K=Blocks(11), L=Wastage(12),
            # N=CorefM3(14), O=Steel(15), P=StarterBars(16)
            # H=OpenCount(8), I=OpenWidth(9)
            ws.cell(r, 2).value = f'=IFERROR(SUMIF(Blockwork!B{DS}:B{DE},A{r},Blockwork!J{DS}:J{DE}),0)'
            ws.cell(r, 3).value = f'=IFERROR(SUMIF(Blockwork!B{DS}:B{DE},A{r},Blockwork!K{DS}:K{DE}),0)'
            ws.cell(r, 4).value = f'=IFERROR(SUMIF(Blockwork!B{DS}:B{DE},A{r},Blockwork!L{DS}:L{DE}),0)'
            ws.cell(r, 5).value = f'=IFERROR(SUMIF(Blockwork!B{DS}:B{DE},A{r},Blockwork!N{DS}:N{DE}),0)'
            ws.cell(r, 6).value = f'=IFERROR(SUMIF(Blockwork!B{DS}:B{DE},A{r},Blockwork!O{DS}:O{DE}),0)'
            ws.cell(r, 7).value = f'=IFERROR(SUMIF(Blockwork!B{DS}:B{DE},A{r},Blockwork!P{DS}:P{DE}),0)'
            ws.cell(r, 9).value = f'=IF(H{r}="","",D{r}*H{r})'
            # Opening totals (cols 8=H and 9=I in Blockwork)
            ws.cell(r, 11).value = f'=IFERROR(SUMIF(Blockwork!B{DS}:B{DE},A{r},Blockwork!H{DS}:H{DE}),0)'
            ws.cell(r, 12).value = f'=IFERROR(SUMIF(Blockwork!B{DS}:B{DE},A{r},Blockwork!I{DS}:I{DE}),0)'
            fmts = {2:"0.00", 3:"#,##0", 4:"#,##0", 5:"0.00",
                    6:"#,##0.0", 7:"#,##0", 9:"$#,##0.00", 11:"#,##0", 12:"0.00"}
            for col, fmt in fmts.items():
                c = ws.cell(r, col)
                c.number_format = fmt
                c.fill = fill(row_bg(r)); c.border = xborder()
                c.font = xfont(); c.alignment = align(h="center", wrap=False)
            for col in [8, 10]:
                c = ws.cell(r, col)
                c.fill = fill(row_bg(r)); c.border = xborder(); c.font = xfont()
            # ── Sanity check cols 13 & 14 ─────────────────────────────────
            c = ws.cell(r, 13)
            c.value = (f'=IF(C{r}=0,"—",IF(ABS(ROUND(C{r}/12.5,2)-B{r})<=1,'
                       f'"✅","❌  "&TEXT(ABS(ROUND(C{r}/12.5,2)-B{r}),"0.00")&" m2 off"))')
            c.fill = fill(row_bg(r)); c.border = xborder()
            c.font = xfont(sz=9); c.alignment = align(h="center", wrap=False)
            c = ws.cell(r, 14)
            c.value = (f'=IF(D{r}=0,"—",IF(ABS(ROUND(D{r}/1.03,0)-C{r})<=1,'
                       f'"✅","❌  "&TEXT(ABS(ROUND(D{r}/1.03,0)-C{r}),"0")&" blocks off"))')
            c.fill = fill(row_bg(r)); c.border = xborder()
            c.font = xfont(sz=9); c.alignment = align(h="center", wrap=False)
            ws.row_dimensions[r].height = 18; r += 1

        blk_sub_end = r - 1

        # ── Blockwork subtotal row ─────────────────────────────────────────
        ws.merge_cells(f"A{r}:H{r}")
        c = ws.cell(r, 1, "Blockwork Subtotal")
        c.font = xfont(bold=True, c=C_WHITE); c.fill = fill(C_DARK_BLUE)
        c.alignment = align(h="right")
        c = ws.cell(r, 9, f"=SUM(I{blk_sub_start}:I{blk_sub_end})")
        c.font = xfont(bold=True, c=C_WHITE); c.fill = fill(C_DARK_BLUE)
        c.number_format = "$#,##0.00"
        for ci in range(2, 15): ws.cell(r, ci).fill = fill(C_DARK_BLUE)
        ws.row_dimensions[r].height = 22
        block_subtotal_row = r; r += 1

        # ── Corefill subtotal ──────────────────────────────────────────────
        ws.merge_cells(f"A{r}:D{r}")
        c = ws.cell(r, 1, "Total Corefill")
        c.font = xfont(bold=True, c=C_WHITE); c.fill = fill(C_MID_BLUE)
        c.alignment = align(h="right")
        c = ws.cell(r, 5, f"=SUM(E{blk_sub_start}:E{blk_sub_end})")
        c.font = xfont(bold=True, c=C_WHITE); c.fill = fill(C_MID_BLUE)
        c.number_format = "0.00"
        c2 = ws.cell(r, 6, "M3")
        c2.font = xfont(italic=True, sz=9, c=C_LIGHT_BLUE); c2.fill = fill(C_MID_BLUE)
        c2.alignment = align(h="left")
        for ci in range(2, 15):
            ws.cell(r, ci).fill = fill(C_MID_BLUE)
        ws.row_dimensions[r].height = 18; r += 1

        # ── Steel subtotal ─────────────────────────────────────────────────
        ws.merge_cells(f"A{r}:E{r}")
        c = ws.cell(r, 1, "Total Steel")
        c.font = xfont(bold=True, c=C_WHITE); c.fill = fill(C_MID_BLUE)
        c.alignment = align(h="right")
        c = ws.cell(r, 6, f"=SUM(F{blk_sub_start}:F{blk_sub_end})")
        c.font = xfont(bold=True, c=C_WHITE); c.fill = fill(C_MID_BLUE)
        c.number_format = "#,##0.0"
        c2 = ws.cell(r, 7, "kg  (÷ 1000 = Tonnes)")
        c2.font = xfont(italic=True, sz=9, c=C_LIGHT_BLUE); c2.fill = fill(C_MID_BLUE)
        c2.alignment = align(h="left")
        for ci in range(2, 15):
            ws.cell(r, ci).fill = fill(C_MID_BLUE)
        ws.row_dimensions[r].height = 18; r += 1

        # ── Starter bars subtotal ──────────────────────────────────────────
        ws.merge_cells(f"A{r}:F{r}")
        c = ws.cell(r, 1, "Total Starter Bars")
        c.font = xfont(bold=True, c=C_WHITE); c.fill = fill(C_MID_BLUE)
        c.alignment = align(h="right")
        c = ws.cell(r, 7, f"=SUM(G{blk_sub_start}:G{blk_sub_end})")
        c.font = xfont(bold=True, c=C_WHITE); c.fill = fill(C_MID_BLUE)
        c.number_format = "#,##0"
        c2 = ws.cell(r, 8, "no.")
        c2.font = xfont(italic=True, sz=9, c=C_LIGHT_BLUE); c2.fill = fill(C_MID_BLUE)
        c2.alignment = align(h="left")
        for ci in range(2, 15):
            ws.cell(r, ci).fill = fill(C_MID_BLUE)
        ws.row_dimensions[r].height = 18; r += 2

    # ── Grand Total ────────────────────────────────────────────────────────
    if scope == "both" and brick_subtotal_row and block_subtotal_row:
        grand = f"=F{brick_subtotal_row}+I{block_subtotal_row}"
        grand_col = 9; label_end = 8
    elif scope == "brickwork" and brick_subtotal_row:
        grand = f"=F{brick_subtotal_row}"
        grand_col = 7; label_end = 6
    else:
        grand = f"=I{block_subtotal_row}"
        grand_col = 9; label_end = 8

    ws.merge_cells(f"A{r}:{get_column_letter(label_end)}{r}")
    c = ws.cell(r, 1, "GRAND TOTAL  (excl. GST)")
    c.font = Font(name=FONT_NAME, bold=True, size=12, color=C_WHITE)
    c.fill = fill(C_DARK_BLUE); c.alignment = align(h="right")
    c = ws.cell(r, grand_col, grand)
    c.font = Font(name=FONT_NAME, bold=True, size=13, color=C_WHITE)
    c.fill = fill(C_DARK_BLUE); c.number_format = "$#,##0.00"
    for ci in range(1, 15): ws.cell(r, ci).fill = fill(C_DARK_BLUE)
    ws.row_dimensions[r].height = 32
    return ws


def build_rates(wb, scope, brick_types, block_types):
    """
    Returns (ws, brick_rates_rows, block_rates_rows)
    brick_rates_rows: list of row numbers where each brick type's
                      bricks-per-m2 result lives (col D)
    """
    ws = wb.create_sheet(RATES_SHEET)
    ws.sheet_properties.tabColor = C_DARK_GREY
    title_block(ws, "For Internal Use Only — George & Estimator Eyes Only",
                "RATES & ASSUMPTIONS", 5)
    set_col_widths(ws, [42, 22, 14, 14, 16, 36])

    r = 7

    # ── DYNAMIC CALCULATOR ────────────────────────────────────────────────
    # Section heading
    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(r, 1, "BRICK CALCULATOR  —  Enter brick face dimensions to auto-calculate bricks per M2")
    c.font = Font(name=FONT_NAME, bold=True, size=11, color=C_WHITE)
    c.fill = fill(C_DARK_BLUE); c.alignment = align()
    ws.row_dimensions[r].height = 28; r += 1

    # Sub-heading explainer
    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(r, 1,
        "Enter brick face dimensions in mm — Length and Height (the two visible sides). "
        "Mortar joint adds to each face. "
        "Bricks per M2 = 1,000,000 ÷ ((Length + Joint) × (Height + Joint))  "
        "Blockwork is always 12.5 per M2 regardless of block width — see note below.")
    c.font      = xfont(sz=9, italic=True, c=C_DARK_GREY)
    c.fill      = fill("FFF9E6")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 36; r += 1

    # Mortar joint input row
    ws.merge_cells(f"A{r}:B{r}")
    c = ws.cell(r, 1, "Mortar joint thickness (mm)")
    c.font = xfont(bold=True, c=C_DARK_BLUE); c.fill = fill(C_LIGHT_BLUE)
    c.alignment = align(h="left"); c.border = xborder()
    mortar_cell = f"C{r}"
    mc = ws.cell(r, 3, 10)   # default 10mm joint
    mc.font = xfont(bold=True); mc.fill = fill(C_YELLOW)
    mc.alignment = align(h="center"); mc.border = xborder()
    mc.number_format = "0"
    ws.cell(r, 4).fill = fill(C_LT_GREY); ws.cell(r, 4).border = xborder()
    ws.merge_cells(f"E{r}:F{r}")
    note_c = ws.cell(r, 5, "Standard = 10mm  |  Fine joint = 3mm")
    note_c.font = xfont(sz=9, italic=True, c=C_DARK_GREY)
    note_c.fill = fill(C_LT_GREY); note_c.alignment = align(h="left")
    ws.row_dimensions[r].height = 20
    mortar_row = r; r += 1

    # Column headers for calculator table
    for ci, h in enumerate(["Type Name", "Length (mm)", "Width (mm)",
                             "Height (mm)", "Units per M2  ←  auto", "Notes"], 1):
        bg = C_SOFT_GREEN if ci == 5 else C_DARK_BLUE
        fg = C_DARK_BLUE if ci == 5 else C_WHITE
        hdr_cell(ws, r, ci, h, bg=bg, fg=fg)
    ws.row_dimensions[r].height = 26
    calc_hdr_row = r; r += 1

    brick_rates_rows  = []
    block_rates_rows  = []

    # ── Brick types rows ───────────────────────────────────────────────────
    if brick_types:
        ws.merge_cells(f"A{r}:F{r}")
        c = ws.cell(r, 1, "BRICKWORK")
        c.font = xfont(bold=True, c=C_DARK_BLUE)
        c.fill = fill(C_AMBER); c.alignment = align(h="left")
        ws.row_dimensions[r].height = 18; r += 1

        for bt in brick_types:
            if not bt.strip(): continue
            bg = row_bg(r)
            # Col A: type name (pre-filled, editable)
            c = ws.cell(r, 1, bt)
            c.font = xfont(); c.fill = fill(bg); c.border = xborder()
            c.alignment = align(h="left")
            # Col B: Length mm (default 230)
            c = ws.cell(r, 2, 230)
            c.font = xfont(); c.fill = fill(bg); c.border = xborder()
            c.number_format = "0"; c.alignment = align(h="center")
            # Col C: Width mm — bricks don't have a structural width like blocks
            # use 110mm (standard brick width) as reference only
            c = ws.cell(r, 3, 110)
            c.font = xfont(italic=True, c=C_DARK_GREY); c.fill = fill(bg); c.border = xborder()
            c.number_format = "0"; c.alignment = align(h="center")
            # Col D: Height mm (default 73mm → ~50/m2 with 10mm joint)
            c = ws.cell(r, 4, 73)
            c.font = xfont(); c.fill = fill(bg); c.border = xborder()
            c.number_format = "0"; c.alignment = align(h="center")
            # Col E: Units per M2 formula — uses Length(B) × Height(D) for face calc
            c = ws.cell(r, 5)
            c.value = (f'=IF(OR(B{r}=0,D{r}=0),"",ROUND('
                       f'1000000/((B{r}+${mortar_cell})*(D{r}+${mortar_cell})),1))')
            c.font = xfont(bold=True, c=C_DARK_BLUE)
            c.fill = fill(C_SOFT_GREEN); c.border = xborder()
            c.number_format = "0.0"; c.alignment = align(h="center")
            # Col F: Notes
            ws.cell(r, 6).fill = fill(bg); ws.cell(r, 6).border = xborder()
            ws.row_dimensions[r].height = 18
            brick_rates_rows.append(r)
            r += 1

    # ── Blockwork — fixed rate, no calculator needed ──────────────────────
    # Blocks per M2 is always 12.5 for ALL standard sizes (George's fixed rate)
    # Block width (90/140/190/290mm) only affects corefill rates, not units/m2
    if block_types:
        ws.merge_cells(f"A{r}:F{r}")
        c = ws.cell(r, 1,
            "BLOCKWORK  —  Fixed at 12.5 blocks per M2 for all standard sizes "
            "(90 / 140 / 190 / 290mm).  Block width affects corefill only, not units per M2.")
        c.font = xfont(sz=9, italic=True, c=C_DARK_GREY)
        c.fill = fill("FFF9E6")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 32; r += 1

    r += 1

    # ── RATES TABLE ───────────────────────────────────────────────────────
    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(r, 1, "RATES")
    c.font = Font(name=FONT_NAME, bold=True, size=11, color=C_WHITE)
    c.fill = fill(C_DARK_BLUE); c.alignment = align()
    ws.row_dimensions[r].height = 26; r += 1

    for ci, h in enumerate(["Item", "Unit", "Rate ($)", "Notes"], 1):
        hdr_cell(ws, r, ci, h)
    ws.row_dimensions[r].height = 26; r += 1

    rate_items = []
    if scope in ("brickwork", "both"):
        rate_items += [
            ("Brickwork – Supply & Lay",          "M2",        ""),
            ("Brickwork – Labour Only",            "M2",        ""),
            ("Brick – Mortar allowance",           "M2",        ""),
        ]
    if scope in ("blockwork", "both"):
        rate_items += [
            ("Blockwork – Supply & Lay (90mm)",       "M2",    ""),
            ("Blockwork – Supply & Lay (140mm)",      "M2",    ""),
            ("Blockwork – Supply & Lay (190mm)",      "M2",    ""),
            ("Blockwork – Supply & Lay (290mm)",      "M2",    ""),
            ("Corefill Concrete",                     "M3",    ""),
            ("N12 / N16 Steel (vert. & horiz.)",      "Tonne", ""),
            ("Starter Bars (incl. drilling & epoxy)", "each",  "Expensive — confirm with George"),
            ("Capping Blocks",                        "each",  ""),
        ]

    for item, unit, note in rate_items:
        bg = C_WHITE if r % 2 == 1 else C_LT_GREY
        data_cell(ws, r, 1, item, bg=bg)
        data_cell(ws, r, 2, unit, bg=bg, h="center")
        data_cell(ws, r, 3, bg=C_YELLOW, fmt="$#,##0.00")
        data_cell(ws, r, 4, note, bg=bg)
        ws.row_dimensions[r].height = 18; r += 1

    r += 1

    # ── ASSUMPTIONS ───────────────────────────────────────────────────────
    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(r, 1, "KEY ASSUMPTIONS  —  George's calculations (do not change)")
    c.font = xfont(bold=True, sz=11, c=C_DARK_BLUE)
    c.fill = fill(C_AMBER); c.alignment = align()
    ws.row_dimensions[r].height = 26; r += 1

    for ci, h in enumerate(["Assumption", "Value", "Notes"], 1):
        hdr_cell(ws, r, ci, h, bg=C_DARK_BLUE)
    ws.row_dimensions[r].height = 22; r += 1

    assumptions = [
        ("Standard bricks per M2 (230×73mm + 10mm joint)", "50",
         "Calculator above will override for non-standard sizes"),
        ("Standard blocks per M2 — ALL standard sizes (90/140/190/290mm)", "12.5 — FIXED",
         "Calculator above will override for non-standard sizes"),
        ("Wastage allowance",                               "3%",      "Applied to brick/block count"),
        ("Capping blocks per LM (400mm length)",            "2.5",     "If 200mm length: use 5.0 — check drawings"),
        ("Corefill — 290mm blocks per M3",                  "80 blks", "George"),
        ("Corefill — 190mm blocks per M3",                  "110 blks","George"),
        ("Corefill — 140mm blocks per M3",                  "~130 blks","George (approx)"),
        ("Corefill — 90mm blocks",                          "N/A",     "Cannot be corefilled"),
        ("Steel in corefilled blockwork",                   "10 kg/M2","N12 & N16 — changes if N20 or double bars"),
        ("Starter bars per LM",                             "2.5",     "Set at 400mm centres"),
    ]
    for assump, val, note in assumptions:
        bg = C_WHITE if r % 2 == 1 else C_LT_GREY
        data_cell(ws, r, 1, assump, bg=bg)
        c = data_cell(ws, r, 2, val, bg=bg, bold=True, h="center")
        data_cell(ws, r, 3, note, bg=bg)
        ws.row_dimensions[r].height = 18; r += 1

    return ws, brick_rates_rows, block_rates_rows


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL ORCHESTRATOR
# ══════════════════════════════════════════════════════════════════════════════
def build_excel(config):
    takeoff_type = config["type"]
    scope        = config.get("scope", "")
    job_name     = config["job_name"]
    brick_types  = config.get("brick_types", [])
    block_types  = config.get("block_types", [])
    save_folder  = config.get("save_folder",
                              os.path.join(os.path.expanduser("~"), "Downloads"))

    wb = Workbook()
    wb.remove(wb.active)

    if takeoff_type == "demo":
        build_demo(wb, job_name)
        type_label = "Demo"
    else:
        type_label = "Masonry"
        # Build rates sheet first so we get the row references back
        rates_ws, brick_rates_rows, block_rates_rows = build_rates(
            wb, scope, brick_types, block_types
        )
        # block_data: list of (name, size, corefill) tuples
        block_data = config.get("block_data", [])
        # Fallback: if block_data not set, build from block_types with empty defaults
        if not block_data and block_types:
            block_data = [(t, "", "") for t in block_types]

        if scope in ("brickwork", "both"):
            build_brickwork(wb, job_name, brick_types, brick_rates_rows)
        if scope in ("blockwork", "both"):
            build_blockwork(wb, job_name, block_data, block_rates_rows)
        build_george_summary(wb, job_name, scope, brick_types,
                             [d[0] for d in block_data])

        # Re-order sheets: Brickwork, Blockwork, George Summary, Rates
        desired = []
        if scope in ("brickwork", "both"):   desired.append("Brickwork")
        if scope in ("blockwork", "both"):   desired.append("Blockwork")
        desired += ["George Summary", RATES_SHEET]
        for i, name in enumerate(desired):
            while wb.sheetnames.index(name) != i:
                wb.move_sheet(name, offset=-1)
        # Keep list sheets hidden
        for hidden in ("_Lists", "_ListsBrickwork", "_ListsBlockwork"):
            if hidden in wb.sheetnames:
                wb[hidden].sheet_state = "hidden"

    os.makedirs(save_folder, exist_ok=True)
    filename  = f"{job_name} - {type_label} Takeoff.xlsx"
    save_path = os.path.join(save_folder, filename)
    base, ext = os.path.splitext(save_path)
    counter   = 1
    while os.path.exists(save_path):
        save_path = f"{base} ({counter}){ext}"
        counter  += 1

    wb.save(save_path)
    return save_path


def build_dynamic_save_path(job_name):
    """
    Construct H:\\My Drive\\Jobs\\[YYYY]\\[MM MMM]\\[job_name]\\
    Strips invalid Windows filename characters from job_name.
    Returns the full path string if H:\\ drive is available, else None.
    """
    if not os.path.isdir("H:\\"):
        return None
    import re
    from datetime import datetime as _dt
    safe_name = re.sub(r'[\\/:*?"<>|]', "", job_name).strip()
    if not safe_name:
        return None
    now   = _dt.now()
    year  = now.strftime("%Y")
    month = now.strftime("%m %B")   # e.g. "05 May"
    return os.path.join("H:\\", "My Drive", "Jobs", year, month, safe_name)


# ══════════════════════════════════════════════════════════════════════════════
#  THEME  —  single unified professional theme for National Estimation & PM
#  Palette: near-black navy / off-white / warm slate / amber gold accent
#  Aesthetic: architectural, minimal, premium
# ══════════════════════════════════════════════════════════════════════════════

# Single unified theme — no more masonry/demo split
T = {
    # ── chrome / structural ─────────────────────────────────────
    "bg_dark":         "#0D0D0D",   # near-black — header, nav
    "bg_mid":          "#1C1C1C",   # dark grey — secondary panels
    "bg_panel":        "#F5F5F5",   # off-white — content area
    "bg_card":         "#E8E8E8",   # light grey — summary cards
    "sep":             "#FFFFFF",   # white — rule lines (was gold)
    # ── text ────────────────────────────────────────────────────
    "text_light":      "#F5F5F5",   # near-white — on dark bg
    "text_dark":       "#0D0D0D",   # near-black — on light bg
    "text_muted":      "#888888",   # mid grey — secondary text
    "text_accent":     "#FFFFFF",   # white — logo mark, highlights (was gold)
    # ── interactive ─────────────────────────────────────────────
    "radio_bg":        "#E8E8E8",
    "radio_select":    "#FFFFFF",   # white selection dot — visible against dark indicator ring
    "entry_hl":        "#0D0D0D",   # black highlight border (was gold)
    "entry_hl_bg":     "#C0C0C0",
    # ── buttons ─────────────────────────────────────────────────
    "btn_back_bg":     "#1C1C1C",
    "btn_back_fg":     "#888888",
    "btn_next_bg":     "#1C1C1C",
    "btn_next_fg":     "#F5F5F5",
    "btn_create_bg":   "#FFFFFF",   # white create button (was gold)
    "btn_create_fg":   "#0D0D0D",   # black text on white
    # ── misc ────────────────────────────────────────────────────
    "progress_bar":    "#FFFFFF",   # white progress bar (was gold)
    "content_bg":      "#F5F5F5",
    "summary_bg":      "#E8E8E8",
    "step_title_fg":   "#0D0D0D",
    "subtitle_fg":     "#888888",
    "header_title_fg": "#FFFFFF",   # white (was gold)
    "bg_step_tag_bg":  "#0D0D0D",   # black pill (was gold)
    "bg_step_tag_fg":  "#FFFFFF",
}

# Keep THEMES dict for compatibility — all keys point to same theme
THEMES = {"masonry": T, "demo": T, "default": T}


# ══════════════════════════════════════════════════════════════════════════════
#  WIZARD APP
# ══════════════════════════════════════════════════════════════════════════════
class WizardApp(tk.Tk):
    W, H = 660, 760

    def __init__(self):
        super().__init__()
        self.title(f"National Estimation & Project Management — Takeoff Wizard  {VERSION}")
        self.resizable(False, False)
        self.protocol("WM_DELETE_WINDOW", self._on_close)

        self.config_data   = {}
        self.step          = 0
        self.steps         = []
        self.current_theme = T   # single unified professional theme

        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"{self.W}x{self.H}+{(sw-self.W)//2}+{(sh-self.H)//2}")

        self._build_chrome()
        self._show_step(0)

    def _build_chrome(self):
        t = self.current_theme

        # ── HEADER ── dark navy band with logo + title ───────────────────────
        self.hdr_frame = tk.Frame(self, bg=t["bg_dark"])
        self.hdr_frame.pack(fill="x")

        # Logo mark — NEPM badge image (Pillow required; falls back to text)
        try:
            from PIL import Image as _PILImage, ImageTk as _ImageTk
            import io as _io
            _raw = base64.b64decode(LOGO_B64)
            _pil = _PILImage.open(_io.BytesIO(_raw)).resize((80, 80), _PILImage.LANCZOS)
            self._logo_img = _ImageTk.PhotoImage(_pil)
            tk.Label(self.hdr_frame, image=self._logo_img,
                     bg=t["bg_dark"], bd=0
                     ).pack(side="left", padx=(16, 14), pady=4)
        except Exception:
            logo_box = tk.Frame(self.hdr_frame, bg=t["bg_mid"], width=80, height=80)
            logo_box.pack(side="left", padx=(16, 14), pady=4)
            logo_box.pack_propagate(False)
            tk.Label(logo_box, text="N", font=("Georgia", 26, "bold"),
                     fg=t["text_accent"], bg=t["bg_mid"]
                     ).place(relx=0.5, rely=0.5, anchor="center")

        # Title stack
        title_stack = tk.Frame(self.hdr_frame, bg=t["bg_dark"])
        title_stack.pack(side="left", pady=16)
        tk.Label(title_stack, text="NATIONAL ESTIMATION",
                 font=("Arial", 15, "bold"),
                 fg=t["text_accent"], bg=t["bg_dark"],
                 anchor="w").pack(anchor="w")
        tk.Label(title_stack, text="AND PROJECT MANAGEMENT",
                 font=("Arial", 7),
                 fg=t["text_muted"], bg=t["bg_dark"],
                 anchor="w").pack(anchor="w")

        # Step indicator — right-aligned
        # Version label
        tk.Label(self.hdr_frame, text=VERSION,
                 font=("Arial", 8), fg=t["text_accent"],
                 bg=t["bg_dark"]).pack(side="right", padx=(0,6))
        self.lbl_step = tk.Label(self.hdr_frame, text="",
                                  font=("Arial", 9),
                                  fg=t["text_muted"], bg=t["bg_dark"])
        self.lbl_step.pack(side="right", padx=(24,4))

        # ── GOLD RULE ────────────────────────────────────────────────────────
        tk.Frame(self, bg=t["sep"], height=2).pack(fill="x")

        # ── STEP LABEL BAR ───────────────────────────────────────────────────
        self.step_bar = tk.Frame(self, bg=t["bg_panel"], height=36)
        self.step_bar.pack(fill="x")
        self.step_bar.pack_propagate(False)

        self.lbl_step_title = tk.Label(self.step_bar, text="",
                                        font=("Arial", 10, "bold"),
                                        fg=t["text_dark"], bg=t["bg_panel"],
                                        anchor="w")
        self.lbl_step_title.pack(side="left", padx=24, pady=8)

        # ── PROGRESS BAR (thin, gold) ─────────────────────────────────────────
        self.prog_track = tk.Frame(self, bg="#D8DCE3", height=3)
        self.prog_track.pack(fill="x")
        self.prog_bar = tk.Frame(self.prog_track, bg=t["progress_bar"], height=3)
        self.prog_bar.place(x=0, y=0, width=0, relheight=1)

        # ── CONTENT AREA ─────────────────────────────────────────────────────
        content_h = self.H - 56 - 88 - 2 - 36 - 3 - 72
        self.content = tk.Frame(self, bg=t["content_bg"], height=content_h)
        self.content.pack(fill="x")
        self.content.pack_propagate(False)

        # ── NAV BAR ──────────────────────────────────────────────────────────
        self.nav_frame = tk.Frame(self, bg=t["bg_dark"], height=72)
        self.nav_frame.pack(fill="x", side="bottom")
        self.nav_frame.pack_propagate(False)

        self.btn_back = tk.Button(self.nav_frame, text="← BACK",
                                   font=("Arial", 9, "bold"),
                                   bg=t["btn_back_bg"], fg=t["btn_back_fg"],
                                   relief="flat", padx=20, pady=10,
                                   cursor="hand2", bd=0,
                                   command=self._go_back)
        self.btn_back.pack(side="left", padx=20, pady=16)

        self.btn_next = tk.Button(self.nav_frame, text="NEXT →",
                                   font=("Arial", 9, "bold"),
                                   bg=t["bg_mid"], fg=t["text_light"],
                                   relief="flat", padx=20, pady=10,
                                   cursor="hand2", bd=0,
                                   command=self._go_next)
        self.btn_next.pack(side="right", padx=20, pady=16)

        self.btn_create = tk.Button(self.nav_frame, text="✓  CREATE TAKEOFF",
                                     font=("Arial", 10, "bold"),
                                     bg=t["btn_create_bg"], fg=t["btn_create_fg"],
                                     relief="flat", padx=24, pady=10,
                                     cursor="hand2", bd=0,
                                     command=self._create)
        self.btn_create.pack(side="right", padx=8, pady=16)

    def apply_theme(self, key):
        # Single unified theme — key ignored, always uses T
        pass

    def _update_nav(self):
        total   = len(self.steps)
        is_last = self.step == total - 1
        t = self.current_theme

        self.btn_back.config(
            state="normal" if self.step > 0 else "disabled",
            bg=t["btn_back_bg"] if self.step > 0 else "#0D1B2A",
            fg=t["btn_back_fg"] if self.step > 0 else "#2A3D52"
        )
        if is_last:
            self.btn_next.pack_forget()
            self.btn_create.pack(side="right", padx=8, pady=16)
        else:
            self.btn_create.pack_forget()
            self.btn_next.pack(side="right", padx=20, pady=16)

        self.lbl_step.config(text=f"{self.step + 1} / {total}")

        # Step title label in the step bar
        step_class  = self.steps[self.step]
        step_title  = getattr(step_class, "TITLE", "")
        self.lbl_step_title.config(text=step_title)

        pct = (self.step + 1) / total
        self.prog_track.update_idletasks()
        self.prog_bar.place(width=int(self.prog_track.winfo_width() * pct))

    def _show_step(self, idx):
        for w in self.content.winfo_children():
            w.destroy()
        self.steps = self._compute_steps()
        self.step  = idx
        frame = self.steps[idx](self.content, self.config_data, self)
        frame.pack(fill="both", expand=True)
        self.current_frame = frame
        self._update_nav()

    def _compute_steps(self):
        steps = [StepType]
        t  = self.config_data.get("type", "")
        sc = self.config_data.get("scope", "")
        if t == "masonry":
            steps.append(StepScope)
            if sc in ("brickwork", "both"):  steps.append(StepBrickTypes)
            if sc in ("blockwork", "both"):  steps.append(StepBlockTypes)
        steps.append(StepJobName)
        return steps

    def _go_next(self):
        if not self.current_frame.validate(): return
        self.current_frame.collect(self.config_data)
        self._show_step(self.step + 1)

    def _go_back(self):
        if self.step > 0:
            self._show_step(self.step - 1)

    def _create(self):
        if not self.current_frame.validate(): return
        self.current_frame.collect(self.config_data)
        try:
            path = build_excel(self.config_data)
            # Remember this save folder for next time
            save_config({"last_save_folder": self.config_data.get("save_folder", "")})
            messagebox.showinfo("Takeoff Created ✓",
                                f"Saved to:\n\n{path}\n\nOpening now…")
            try: os.startfile(path)
            except Exception: pass
            self.destroy()
        except Exception as e:
            messagebox.showerror("Error", f"Something went wrong:\n\n{e}")

    def _on_close(self):
        if messagebox.askokcancel("Quit", "Cancel and exit the wizard?"):
            self.destroy()


# ══════════════════════════════════════════════════════════════════════════════
#  BASE STEP
# ══════════════════════════════════════════════════════════════════════════════
class BaseStep(tk.Frame):
    TITLE    = ""
    SUBTITLE = ""
    N_TYPES  = 20   # number of type entry slots

    def __init__(self, parent, config, app):
        t = app.current_theme
        super().__init__(parent, bg=t["content_bg"])
        self.config_data = config
        self.app         = app
        self.theme       = t
        self._build_header()
        self._build_body()

    def _build_header(self):
        t = self.theme
        tk.Label(self, text=self.TITLE, font=("Arial", 14, "bold"),
                 fg=t["step_title_fg"], bg=t["content_bg"]
                 ).pack(pady=(18, 3), padx=30, anchor="w")
        if self.SUBTITLE:
            tk.Label(self, text=self.SUBTITLE, font=("Arial", 9),
                     fg=t["subtitle_fg"], bg=t["content_bg"],
                     wraplength=550, justify="left"
                     ).pack(padx=30, anchor="w", pady=(0, 8))
        tk.Frame(self, bg=t["sep"], height=2).pack(fill="x", padx=30, pady=(0, 12))

    def _build_body(self): pass
    def validate(self):    return True
    def collect(self, config): pass

    def _radio_group(self, parent, var, options):
        t = self.theme
        for text, value in options:
            tk.Radiobutton(parent, text=text, variable=var, value=value,
                           font=("Arial", 11), bg=t["radio_bg"], fg=t["text_dark"],
                           activebackground=t["radio_bg"],
                           selectcolor=t["radio_select"],
                           pady=7, padx=6, anchor="w"
                           ).pack(fill="x", padx=8, pady=2)

    def _make_entry(self, parent, font_size=11):
        t = self.theme
        return tk.Entry(parent, font=("Arial", font_size), relief="solid",
                        highlightthickness=1, highlightcolor=t["entry_hl"],
                        highlightbackground=t["entry_hl_bg"],
                        bg=t["content_bg"], fg=t["text_dark"],
                        insertbackground=t["text_dark"], bd=0)

    def _add_placeholder(self, entry, placeholder):
        t = self.theme
        entry.insert(0, placeholder)
        entry.config(fg=t["text_muted"])
        def _fi(ev, en=entry, ph=placeholder):
            if en.get() == ph: en.delete(0, tk.END); en.config(fg=t["text_dark"])
        def _fo(ev, en=entry, ph=placeholder):
            if en.get() == "": en.insert(0, ph); en.config(fg=t["text_muted"])
        entry.bind("<FocusIn>",  _fi)
        entry.bind("<FocusOut>", _fo)

    def _type_entry_grid(self, parent, existing, placeholders):
        """Scrollable two-column grid of N_TYPES entries."""
        t = self.theme

        # ── scrollable canvas container ───────────────────────────────────
        CANVAS_H = 380   # visible height in pixels
        canvas = tk.Canvas(parent, bg=t["content_bg"],
                           highlightthickness=0, height=CANVAS_H)
        scrollbar = tk.Scrollbar(parent, orient="vertical",
                                 command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        # inner frame inside the canvas
        inner = tk.Frame(canvas, bg=t["content_bg"])
        window_id = canvas.create_window((0, 0), window=inner, anchor="nw")

        # make inner frame fill canvas width
        def _on_canvas_resize(event):
            canvas.itemconfig(window_id, width=event.width)
        canvas.bind("<Configure>", _on_canvas_resize)

        # update scroll region when inner frame changes size
        def _on_frame_resize(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        inner.bind("<Configure>", _on_frame_resize)

        # mousewheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # ── two-column grid ───────────────────────────────────────────────
        inner.columnconfigure(0, weight=1)
        inner.columnconfigure(1, weight=1)

        entries = []
        for i in range(self.N_TYPES):
            col_frame = tk.Frame(inner, bg=t["content_bg"])
            col_frame.grid(row=i % 10, column=i // 10, sticky="ew",
                           padx=(4 if i // 10 == 1 else 0, 4), pady=3)
            col_frame.columnconfigure(1, weight=1)

            tk.Label(col_frame, text=f"{i+1}.", font=("Arial", 10),
                     fg=t["subtitle_fg"], bg=t["content_bg"],
                     width=3, anchor="e").pack(side="left")

            e = self._make_entry(col_frame)
            e.pack(side="left", fill="x", expand=True, ipady=5, padx=(4, 0))

            ph = placeholders[i] if i < len(placeholders) else f"Type {i+1}"
            val = existing[i] if i < len(existing) else ""
            if val:
                e.insert(0, val); e.config(fg=t["text_dark"])
            else:
                self._add_placeholder(e, ph)
            entries.append((e, ph))

        return entries


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 1 — TYPE
# ══════════════════════════════════════════════════════════════════════════════
class StepType(BaseStep):
    TITLE    = "What type of takeoff is this?"
    SUBTITLE = "Your selection themes the wizard and determines which sheets are created."

    def _build_body(self):
        t = self.theme
        self.var = tk.StringVar(value=self.config_data.get("type", "masonry"))
        box = tk.Frame(self, bg=t["radio_bg"])
        box.pack(padx=30, fill="x", pady=4)
        self._radio_group(box, self.var, [
            ("Masonry Takeoff  —  brickwork, blockwork, or both", "masonry"),
            ("Demolition Takeoff",                                 "demo"),
        ])

    def collect(self, config):
        config["type"] = self.var.get()
        if config["type"] == "demo":
            config.pop("scope", None)
            config.pop("brick_types", None)
            config.pop("block_types", None)


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 2 — SCOPE
# ══════════════════════════════════════════════════════════════════════════════
class StepScope(BaseStep):
    TITLE    = "What is the masonry scope?"
    SUBTITLE = "Determines which worksheets are created in your takeoff."

    def _build_body(self):
        t = self.theme
        self.var = tk.StringVar(value=self.config_data.get("scope", "brickwork"))
        box = tk.Frame(self, bg=t["radio_bg"])
        box.pack(padx=30, fill="x", pady=4)
        self._radio_group(box, self.var, [
            ("Brickwork only",             "brickwork"),
            ("Blockwork only",             "blockwork"),
            ("Both brickwork & blockwork", "both"),
        ])

    def collect(self, config):
        config["scope"] = self.var.get()


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 3a — BRICK TYPES  (20 slots)
# ══════════════════════════════════════════════════════════════════════════════
class StepBrickTypes(BaseStep):
    TITLE    = "What brick types are on this job?"
    SUBTITLE = ("Up to 20 types — include size and finish in one field "
                "(e.g. '230mm Standard Brick - Painted'). "
                "These become your dropdowns and flow into the George Summary. "
                "Dimensions can be refined on the Rates sheet after creation.")

    PLACEHOLDERS = [
        "e.g. 230mm Standard Brick",
        "e.g. 230mm Standard Brick - Painted",
        "e.g. 230mm Recycled Brick",
        "e.g. Brick Veneer",
        "e.g. Double Skin Brick",
        "e.g. 50mm Brick",
        "e.g. 50mm Brick - Painted",
        "e.g. Face Brick",
        "e.g. Engineer Brick",
        "e.g. Clinker Brick",
        "e.g. 230mm Standard Brick - Bagged",
        "e.g. 230mm Standard Brick - Rendered",
        "e.g. 110mm Brick",
        "e.g. Glazed Brick",
        "e.g. Perforated Brick",
        "e.g. Split Face Brick",
        "e.g. Pressed Brick",
        "e.g. Handmade Brick",
        "e.g. Limestone Block",
        "e.g. Custom Brick Type",
    ]

    def _build_body(self):
        existing = self.config_data.get("brick_types", [])
        # hint above the scrollable grid
        tk.Label(self, text="Leave blank to skip — only filled entries appear in dropdowns.",
                 font=("Arial", 9), fg=self.theme["text_muted"],
                 bg=self.theme["content_bg"], wraplength=580, justify="left"
                 ).pack(padx=30, pady=(0, 6), anchor="w")
        container = tk.Frame(self, bg=self.theme["content_bg"])
        container.pack(padx=30, fill="both", expand=True)
        self.entries = self._type_entry_grid(container, existing, self.PLACEHOLDERS)

    def validate(self):
        vals = [e.get().strip() for e, ph in self.entries
                if e.get().strip() not in ("", ph)]
        if not vals:
            messagebox.showwarning("No brick types entered",
                                   "Please enter at least one brick type.")
            return False
        return True

    def collect(self, config):
        config["brick_types"] = [
            e.get().strip() for e, ph in self.entries
            if e.get().strip() not in ("", ph)
        ]


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 3b — BLOCK TYPES  (20 slots)
# ══════════════════════════════════════════════════════════════════════════════
class StepBlockTypes(BaseStep):
    TITLE    = "What block types are on this job?"
    SUBTITLE = ("Up to 20 types. Enter the name, select block size and whether it's corefilled. "
                "Size and Corefilled? will auto-populate in the spreadsheet — "
                "you can still override them per row if needed.")

    N_TYPES = 20
    SIZES   = ["", "90mm", "140mm", "190mm", "290mm"]
    PNAMES  = [
        "e.g. 190mm Besser Block", "e.g. 290mm Besser Block",
        "e.g. 140mm Besser Block", "e.g. 90mm Besser Block",
        "e.g. 190mm Besser Block - Painted", "e.g. 190mm Besser Block - Bagged",
        "e.g. 190mm Besser Block - Rendered", "e.g. 290mm Besser Block - Painted",
        "e.g. 290mm Besser Block - Bagged", "e.g. 140mm Besser Block - Painted",
        "e.g. 90mm Besser Block - Painted", "e.g. 190mm Feature Block",
        "e.g. 190mm H Block", "e.g. 290mm H Block",
        "e.g. 190mm Lintel Block", "e.g. 290mm Lintel Block",
        "e.g. 140mm Feature Block", "e.g. 190mm Besser Block - Rendered",
        "e.g. 290mm Besser Block - Rendered", "e.g. Custom Block Type",
    ]

    def _guess_size(self, name):
        """Auto-detect size from type name for convenience."""
        for sz in ["290", "190", "140", "90"]:
            if sz in name:
                return sz + "mm"
        return ""

    def _build_body(self):
        t = self.theme
        existing_data = self.config_data.get("block_data", [])

        tk.Label(self,
                 text="Leave blank to skip. Size and Corefilled? auto-populate in the spreadsheet — override per row if needed.",
                 font=("Arial", 9), fg=t["text_muted"], bg=t["content_bg"],
                 wraplength=600, justify="left"
                 ).pack(padx=24, pady=(0, 4), anchor="w")

        # Column headers
        hdr_row = tk.Frame(self, bg=t["bg_mid"])
        hdr_row.pack(padx=24, fill="x", pady=(0, 2))
        tk.Label(hdr_row, text="#",      width=3,  font=("Arial", 8, "bold"),
                 fg=t["text_light"], bg=t["bg_mid"], anchor="e").pack(side="left", padx=(4,4))
        tk.Label(hdr_row, text="Block Type Name", font=("Arial", 8, "bold"),
                 fg=t["text_light"], bg=t["bg_mid"], anchor="w").pack(side="left", fill="x", expand=True)
        tk.Label(hdr_row, text="Size",   width=9,  font=("Arial", 8, "bold"),
                 fg=t["text_light"], bg=t["bg_mid"], anchor="c").pack(side="left", padx=(4,0))
        tk.Label(hdr_row, text="Corefill?", width=9, font=("Arial", 8, "bold"),
                 fg=t["text_light"], bg=t["bg_mid"], anchor="c").pack(side="left", padx=(4,4))

        # Scrollable canvas
        canvas = tk.Canvas(self, bg=t["content_bg"], highlightthickness=0, height=400)
        sb = tk.Scrollbar(self, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y", padx=(0,4))
        canvas.pack(padx=(24,0), fill="both", expand=True)

        inner = tk.Frame(canvas, bg=t["content_bg"])
        wid = canvas.create_window((0,0), window=inner, anchor="nw")
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(wid, width=e.width))
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        self.block_rows = []  # list of (name_entry, size_var, corefill_var)

        for i in range(self.N_TYPES):
            ex_name    = existing_data[i][0] if i < len(existing_data) else ""
            ex_size    = existing_data[i][1] if i < len(existing_data) else ""
            ex_corefill= existing_data[i][2] if i < len(existing_data) else ""
            ph = self.PNAMES[i] if i < len(self.PNAMES) else f"Block type {i+1}"

            row = tk.Frame(inner, bg=t["content_bg"])
            row.pack(fill="x", pady=2, padx=2)

            # Row number
            tk.Label(row, text=f"{i+1}.", width=3, font=("Arial", 9),
                     fg=t["subtitle_fg"], bg=t["content_bg"], anchor="e").pack(side="left", padx=(0,4))

            # Name entry
            e = self._make_entry(row, font_size=10)
            e.pack(side="left", fill="x", expand=True, ipady=4)
            if ex_name:
                e.insert(0, ex_name); e.config(fg=t["text_dark"])
            else:
                self._add_placeholder(e, ph)

            # Size dropdown
            size_var = tk.StringVar(value=ex_size)
            size_om  = tk.OptionMenu(row, size_var, *self.SIZES)
            size_om.config(width=6, font=("Arial", 9),
                           bg=t["bg_mid"], fg=t["text_light"],
                           activebackground=t["sep"], relief="flat",
                           highlightthickness=0)
            size_om["menu"].config(bg=t["bg_mid"], fg=t["text_light"],
                                   font=("Arial", 9))
            size_om.pack(side="left", padx=(6,0))

            # Auto-detect size when name is typed (if size not already set)
            def _auto_size(event, entry=e, svar=size_var, ph_=ph):
                name = entry.get().strip()
                if name and name != ph_ and not svar.get():
                    svar.set(self._guess_size(name))
            e.bind("<FocusOut>", _auto_size)

            # Corefill dropdown
            cf_var = tk.StringVar(value=ex_corefill)
            cf_om  = tk.OptionMenu(row, cf_var, "", "Yes", "No")
            cf_om.config(width=5, font=("Arial", 9),
                         bg=t["bg_mid"], fg=t["text_light"],
                         activebackground=t["sep"], relief="flat",
                         highlightthickness=0)
            cf_om["menu"].config(bg=t["bg_mid"], fg=t["text_light"],
                                 font=("Arial", 9))
            cf_om.pack(side="left", padx=(4,0))

            self.block_rows.append((e, size_var, cf_var, ph))

    def validate(self):
        vals = [e.get().strip() for e, sv, cv, ph in self.block_rows
                if e.get().strip() not in ("", ph)]
        if not vals:
            messagebox.showwarning("No block types entered",
                                   "Please enter at least one block type.")
            return False
        return True

    def collect(self, config):
        block_data  = []
        block_types = []
        for e, size_var, cf_var, ph in self.block_rows:
            name = e.get().strip()
            if name and name not in ("", ph):
                block_data.append((name, size_var.get(), cf_var.get()))
                block_types.append(name)
        config["block_data"]  = block_data
        config["block_types"] = block_types


# ══════════════════════════════════════════════════════════════════════════════
#  STEP FINAL — JOB NAME
# ══════════════════════════════════════════════════════════════════════════════
class StepJobName(BaseStep):
    TITLE    = "What is the job name?"
    SUBTITLE = "Used as the sheet title and file name. Saves to your chosen folder."

    def _build_body(self):
        t     = self.theme
        inner = tk.Frame(self, bg=t["content_bg"])
        inner.pack(padx=30, fill="x")

        tk.Label(inner, text="Job name", font=("Arial", 10, "bold"),
                 fg=t["step_title_fg"], bg=t["content_bg"]
                 ).pack(anchor="w", pady=(0, 5))

        self.job_entry = self._make_entry(inner, font_size=13)
        self.job_entry.config(highlightthickness=2)
        self.job_entry.pack(fill="x", ipady=9)
        existing = self.config_data.get("job_name", "")
        if existing:
            self.job_entry.insert(0, existing)
        self.job_entry.focus_set()

        self.preview_lbl = tk.Label(inner, text="", font=("Arial", 9, "italic"),
                                     fg=t["text_muted"], bg=t["content_bg"], anchor="w")
        self.preview_lbl.pack(anchor="w", pady=(5, 0))

        # save folder
        tk.Label(inner, text="\nSave location", font=("Arial", 10, "bold"),
                 fg=t["step_title_fg"], bg=t["content_bg"]
                 ).pack(anchor="w", pady=(10, 4))

        folder_row = tk.Frame(inner, bg=t["content_bg"])
        folder_row.pack(fill="x")
        cfg_saved = load_config().get("last_save_folder", "")
        default = self.config_data.get("save_folder",
                   cfg_saved if cfg_saved and os.path.isdir(cfg_saved)
                   else os.path.join(os.path.expanduser("~"), "Downloads"))
        self.folder_var = tk.StringVar(value=default)
        self._folder_manually_set = False   # flips True if user clicks Browse
        fe = self._make_entry(folder_row, font_size=10)
        fe.config(textvariable=self.folder_var)
        fe.pack(side="left", fill="x", expand=True, ipady=5)

        def browse():
            from tkinter import filedialog
            d = filedialog.askdirectory(initialdir=self.folder_var.get())
            if d:
                self.folder_var.set(d)
                self._folder_manually_set = True
                self._refresh()

        tk.Button(folder_row, text="Browse…", font=("Arial", 10),
                  bg=t["btn_back_bg"], fg=t["btn_back_fg"],
                  relief="flat", padx=10, cursor="hand2", command=browse
                  ).pack(side="left", padx=(8, 0), ipady=4)

        # summary
        tk.Label(inner, text="\nJob summary", font=("Arial", 10, "bold"),
                 fg=t["step_title_fg"], bg=t["content_bg"]
                 ).pack(anchor="w")

        self.summary_lbl = tk.Label(inner, text="", font=("Arial", 9),
                                     fg=t["text_dark"], bg=t["summary_bg"],
                                     wraplength=540, justify="left",
                                     anchor="nw", padx=12, pady=10)
        self.summary_lbl.pack(fill="x")

        self.job_entry.bind("<KeyRelease>", lambda e: self._refresh())
        self._refresh()

    def _refresh(self):
        job = self.job_entry.get().strip()
        typ = self.config_data.get("type", "masonry").capitalize()
        self.preview_lbl.config(
            text=f"→  Will save as:  {job} - {typ} Takeoff.xlsx" if job
            else "→  Enter a job name above"
        )
        # Auto-update save path from job name unless user has manually browsed
        if not self._folder_manually_set and job:
            dynamic = build_dynamic_save_path(job)
            if dynamic:
                self.folder_var.set(dynamic)
        cfg  = self.config_data
        sc   = cfg.get("scope", "—").capitalize()
        bt   = ", ".join(cfg.get("brick_types", [])) or "—"
        blkt = ", ".join(cfg.get("block_types", [])) or "—"
        lines = [f"Type:       {typ}"]
        if cfg.get("type") == "masonry":
            lines.append(f"Scope:      {sc}")
            if cfg.get("scope") in ("brickwork", "both"): lines.append(f"Bricks:     {bt}")
            if cfg.get("scope") in ("blockwork", "both"): lines.append(f"Blocks:     {blkt}")
        lines.append(f"Job name:   {job or '—'}")
        self.summary_lbl.config(text="\n".join(lines))

    def validate(self):
        if not self.job_entry.get().strip():
            messagebox.showwarning("Job name required",
                                   "Please enter a job name before creating the takeoff.")
            return False
        return True

    def collect(self, config):
        config["job_name"]    = self.job_entry.get().strip()
        config["save_folder"] = self.folder_var.get()


# ══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    check_for_updates()   # Silent version check — no-op if offline or up to date
    app = WizardApp()
    app.mainloop()
