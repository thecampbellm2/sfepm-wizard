"""
SFEPM Takeoff Wizard  v1.2
==================
Double-click to run. Requires Python 3.8+ with openpyxl.
If openpyxl is not installed, run:  pip install openpyxl

Changes in v1.2:
  - Brickwork tab: Opening Count and Opening Width (LM) columns added
    next to Deductions M2 for lintel reference
  - Blockwork tab: Same opening columns added
  - George Summary: Total Openings and Total Opening Width (LM) per
    brick/block type, summed from takeoff sheets

Changes in v1.1:
  - Auto-updater: wizard checks GitHub for updates on launch
  - Forced update with changelog shown after restart

Changes in v1.0 (internal v5):
  - Font changed to Aptos throughout spreadsheet
  - Total ($) now calculates as Wastage qty × Rate (rate per brick/block)
  - Dynamic brick/block calculator on Rates & Assumptions sheet
  - 10 brick/block type slots in wizard (up from 5)
"""

import sys
import os
import tkinter as tk
from tkinter import messagebox

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    import subprocess
    root = tk.Tk(); root.withdraw()
    ans = messagebox.askyesno(
        "Missing dependency",
        "openpyxl is required but not installed.\n\nInstall it now? (requires internet)"
    )
    if ans:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    else:
        sys.exit(1)



# ══════════════════════════════════════════════════════════════════════════════
#  VERSION & USER LOOKUP
# ══════════════════════════════════════════════════════════════════════════════
VERSION = "v1.2"

# ── Auto-updater config ───────────────────────────────────────────────────────
# Set these to your GitHub repo's raw file URLs after first publish.
# Replace YOUR_USERNAME and YOUR_REPO with your actual GitHub details.
GITHUB_VERSION_URL   = "https://raw.githubusercontent.com/thecampbellm2/sfepm-wizard/main/version.txt"
GITHUB_EXE_URL       = "https://raw.githubusercontent.com/thecampbellm2/sfepm-wizard/main/SFEPM_Takeoff_Wizard.exe"
GITHUB_CHANGELOG_URL = "https://raw.githubusercontent.com/YOUR_USERNAME/YOUR_REPO/main/changelog.txt"
# ─────────────────────────────────────────────────────────────────────────────

# Map Windows usernames to display names for "Prepared by" field
# Add new users here: "windows_username": "Display Name"
USER_DISPLAY_NAMES = {
    "theca": "Mike Campbell",
    # "george": "Georgina Murison",   # ← uncomment and set correct username
}

def get_display_name():
    """Return display name for current Windows user, fallback to username."""
    import os
    try:
        username = os.environ.get("USERNAME", os.environ.get("USER", "")).lower()
        return USER_DISPLAY_NAMES.get(username, username.title())
    except Exception:
        return ""


# ══════════════════════════════════════════════════════════════════════════════
#  PERSISTENT CONFIG  (remembers save folder between sessions)
# ══════════════════════════════════════════════════════════════════════════════
import json

CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           ".takeoff_config.json")

def load_config():
    try:
        with open(CONFIG_FILE) as f:
            return json.load(f)
    except Exception:
        return {}

def save_config(data):
    try:
        existing = load_config()
        existing.update(data)
        with open(CONFIG_FILE, "w") as f:
            json.dump(existing, f, indent=2)
    except Exception:
        pass

# ══════════════════════════════════════════════════════════════════════════════
#  AUTO-UPDATER
# ══════════════════════════════════════════════════════════════════════════════

def _version_tuple(v):
    """Convert 'v1.2' or '1.2' to (1, 2) for comparison."""
    return tuple(int(x) for x in v.lstrip("v").split("."))


# Path used to pass changelog content to the newly launched exe after update
_UPDATE_FLAG_PATH = os.path.join(os.path.expanduser("~"), ".sfepm_just_updated.txt")


def _show_post_update_changelog():
    """
    Called at startup. If a just-updated flag file exists, show the
    changelog in a dialog then delete the flag. Runs regardless of frozen state.
    """
    if not os.path.exists(_UPDATE_FLAG_PATH):
        return
    try:
        with open(_UPDATE_FLAG_PATH, encoding="utf-8") as f:
            content = f.read().strip()
        os.remove(_UPDATE_FLAG_PATH)

        lines     = content.split("\n", 1)
        version   = lines[0].strip() if lines else VERSION
        changelog = lines[1].strip() if len(lines) > 1 else ""

        _root = tk.Tk()
        _root.withdraw()
        msg = f"Successfully updated to {version}!\n\n"
        if changelog:
            msg += f"What's new in this version:\n\n{changelog}"
        messagebox.showinfo(
            f"Updated to {version} — SFEPM Takeoff Wizard", msg
        )
        _root.destroy()
    except Exception:
        try:
            os.remove(_UPDATE_FLAG_PATH)
        except Exception:
            pass


def check_for_updates():
    """
    Checks GitHub for a newer version on startup. If one is found the update
    is FORCED — no option to skip. Changelog is fetched and shown after the
    new exe launches. Fails silently on any network or parsing error.
    Only performs the version check when packaged as a PyInstaller exe.
    """
    # Always check for the post-update flag first (runs as script or exe)
    _show_post_update_changelog()

    if not getattr(sys, "frozen", False):
        return   # Running as .py script — skip version check

    if "YOUR_USERNAME" in GITHUB_VERSION_URL:
        return   # Placeholder URLs not yet configured

    import urllib.request
    import threading
    import tempfile
    import subprocess

    try:
        # ── Version check ────────────────────────────────────────────────────
        with urllib.request.urlopen(GITHUB_VERSION_URL, timeout=5) as r:
            latest = r.read().decode().strip()

        if _version_tuple(latest) <= _version_tuple(VERSION):
            return   # Already up to date

        # ── Fetch changelog ──────────────────────────────────────────────────
        changelog = ""
        try:
            with urllib.request.urlopen(GITHUB_CHANGELOG_URL, timeout=5) as r:
                changelog = r.read().decode().strip()
        except Exception:
            pass   # Changelog unavailable — continue without it

        # ── Forced update notice (no skip option) ────────────────────────────
        _root = tk.Tk()
        _root.withdraw()
        msg = (
            f"A required update is available.\n\n"
            f"   Current : {VERSION}\n"
            f"   Latest  : v{latest}\n\n"
        )
        if changelog:
            msg += f"What's new:\n\n{changelog}\n\n"
        msg += "Click OK to install the update.\nThe wizard will restart automatically."
        messagebox.showinfo(
            "Required Update — SFEPM Takeoff Wizard", msg
        )
        _root.destroy()

        # ── Download with progress window ────────────────────────────────────
        exe_path     = sys.executable
        new_exe_path = exe_path + ".update"

        prog = tk.Tk()
        prog.title("Downloading update…")
        prog.geometry("380x120")
        prog.resizable(False, False)
        tk.Label(
            prog,
            text=f"Downloading SFEPM Takeoff Wizard v{latest}…",
            font=("Arial", 10), pady=16
        ).pack()
        try:
            from tkinter import ttk
            bar = ttk.Progressbar(prog, length=320, mode="indeterminate")
            bar.pack()
            bar.start(12)
        except Exception:
            pass
        tk.Label(
            prog,
            text="Please wait — do not close this window.",
            font=("Arial", 8), fg="#888888", pady=8
        ).pack()

        done  = [False]
        error = [None]

        def _download():
            try:
                urllib.request.urlretrieve(GITHUB_EXE_URL, new_exe_path)
                done[0] = True
            except Exception as exc:
                error[0] = str(exc)
            finally:
                try:
                    prog.destroy()
                except Exception:
                    pass

        threading.Thread(target=_download, daemon=True).start()
        prog.mainloop()

        if error[0]:
            raise RuntimeError(error[0])

        if not done[0]:
            # Window was closed — force a re-check next launch
            return

        # ── Write post-update flag so new exe shows changelog ────────────────
        try:
            with open(_UPDATE_FLAG_PATH, "w", encoding="utf-8") as f:
                f.write(f"v{latest}\n{changelog}")
        except Exception:
            pass

        # ── Swap exe via a tiny batch file ───────────────────────────────────
        bat = (
            "@echo off\n"
            "timeout /t 2 /nobreak > nul\n"
            f'move /y "{new_exe_path}" "{exe_path}"\n'
            f'start "" "{exe_path}"\n'
            "del \"%~f0\"\n"
        )
        bat_path = os.path.join(tempfile.gettempdir(), "sfepm_wizard_update.bat")
        with open(bat_path, "w") as f:
            f.write(bat)

        subprocess.Popen(
            ["cmd", "/c", bat_path],
            creationflags=subprocess.CREATE_NO_WINDOW
        )
        sys.exit(0)   # Close current version — batch relaunches new one

    except Exception:
        pass   # Always fail silently — never block the app


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL PALETTE & STYLE HELPERS
# ══════════════════════════════════════════════════════════════════════════════
FONT_NAME = "Aptos"

C_DARK_BLUE  = "1F3864"
C_MID_BLUE   = "2E75B6"
C_LIGHT_BLUE = "D6E4F0"
C_AMBER      = "F4B942"
C_WHITE      = "FFFFFF"
C_LT_GREY    = "F2F2F2"
C_DARK_GREY  = "595959"
C_GREEN      = "E2EFDA"
C_YELLOW     = "FFF2CC"
C_SOFT_GREEN = "EAF4EA"

def fill(c):
    return PatternFill("solid", fgColor=c)

def xfont(bold=False, c="000000", sz=10, italic=False):
    return Font(name=FONT_NAME, bold=bold, color=c, size=sz, italic=italic)

def xborder():
    t = Side(style="thin", color="BFBFBF")
    return Border(left=t, right=t, top=t, bottom=t)

def align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def hdr_cell(ws, row, col, text, bg=C_DARK_BLUE, fg=C_WHITE, sz=10, bold=True):
    c = ws.cell(row, col, text)
    c.font      = xfont(bold=bold, c=fg, sz=sz)
    c.fill      = fill(bg)
    c.alignment = align()
    c.border    = xborder()
    return c

def data_cell(ws, row, col, value=None, bg=C_WHITE, fmt=None,
              bold=False, fg="000000", h="left"):
    c = ws.cell(row, col)
    if value is not None:
        c.value = value
    c.font      = xfont(bold=bold, c=fg)
    c.fill      = fill(bg)
    c.alignment = align(h=h, wrap=False)
    c.border    = xborder()
    if fmt:
        c.number_format = fmt
    return c

def row_bg(r):
    return C_WHITE if r % 2 == 1 else C_LT_GREY

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def title_block(ws, job_name, sheet_label, n_cols=16):
    """
    Branded title block:
      Row 1 — dark bar: logo mark left  |  sheet label centre  |  business name right
      Row 2 — accent bar: job name
      Row 3 — meta bar: date left  |  prepared by right
      Rows 4-5 — spacers
    """
    last = get_column_letter(n_cols)
    mid  = get_column_letter(max(1, n_cols // 2))
    nxt  = get_column_letter(max(2, n_cols // 2 + 1))
    logo_cols = 2  # columns reserved for logo mark

    # ── Row 1 — primary header bar ──────────────────────────────────────────
    ws.row_dimensions[1].height = 52
    wide = n_cols >= 8   # wide enough for split layout with business name

    # Logo mark cell (cols 1-2 if wide, else col 1 only)
    logo_cols_actual = min(logo_cols, n_cols - 1) if wide else 1
    logo_col = get_column_letter(logo_cols_actual)
    ws.merge_cells(f"A1:{logo_col}1")
    c = ws["A1"]
    c.value     = "SF"
    c.font      = Font(name=FONT_NAME, bold=True, size=16, color=C_AMBER)
    c.fill      = fill("0D1B2A")
    c.alignment = align(h="center")

    if wide:
        # Sheet label (centre span, leaving 4 cols right for business name)
        mid_start = get_column_letter(logo_cols_actual + 1)
        mid_end   = get_column_letter(n_cols - 4)
        ws.merge_cells(f"{mid_start}1:{mid_end}1")
        c = ws[f"{mid_start}1"]
        c.value     = sheet_label
        c.font      = Font(name=FONT_NAME, bold=True, size=16, color=C_WHITE)
        c.fill      = fill(C_DARK_BLUE)
        c.alignment = align(h="center")

        # Business name (right 4 cols)
        biz_start = get_column_letter(n_cols - 3)
        ws.merge_cells(f"{biz_start}1:{last}1")
        c = ws[f"{biz_start}1"]
        c.value     = "SYDNEY FITOUT"
        c.font      = Font(name=FONT_NAME, bold=True, size=11, color=C_AMBER)
        c.fill      = fill(C_DARK_BLUE)
        c.alignment = align(h="right")
    else:
        # Narrow sheet: sheet label fills remaining cols
        if logo_cols_actual < n_cols:
            label_start = get_column_letter(logo_cols_actual + 1)
            ws.merge_cells(f"{label_start}1:{last}1")
            c = ws[f"{label_start}1"]
        else:
            c = ws["A1"]
        c.value     = sheet_label
        c.font      = Font(name=FONT_NAME, bold=True, size=13, color=C_WHITE)
        c.fill      = fill(C_DARK_BLUE)
        c.alignment = align(h="center")

    # ── Row 2 — job name / accent bar ───────────────────────────────────────
    ws.row_dimensions[2].height = 28

    if wide:
        # Left portion: job name
        tag_end = get_column_letter(n_cols - 4)
        ws.merge_cells(f"A2:{tag_end}2")
        c = ws["A2"]
        c.value     = job_name
        c.font      = Font(name=FONT_NAME, bold=True, size=13, color=C_WHITE)
        c.fill      = fill("0D1B2A")
        c.alignment = Alignment(horizontal="left", vertical="center",
                                indent=2, wrap_text=False)
        # Right portion: tagline
        tag_start = get_column_letter(n_cols - 3)
        ws.merge_cells(f"{tag_start}2:{last}2")
        c = ws[f"{tag_start}2"]
        c.value     = "ESTIMATION & PROJECT MANAGEMENT"
        c.font      = Font(name=FONT_NAME, size=7, color="8899AA", italic=True)
        c.fill      = fill("0D1B2A")
        c.alignment = align(h="right")
    else:
        # Narrow: full-width job name
        ws.merge_cells(f"A2:{last}2")
        c = ws["A2"]
        c.value     = job_name
        c.font      = Font(name=FONT_NAME, bold=True, size=11, color=C_WHITE)
        c.fill      = fill("0D1B2A")
        c.alignment = Alignment(horizontal="left", vertical="center",
                                indent=1, wrap_text=False)

    # ── Row 3 — meta bar ────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 18
    from datetime import date as _date
    ws.merge_cells(f"A3:{mid}3")
    c = ws["A3"]
    c.value     = f"Date:  {_date.today().strftime('%d/%m/%Y')}"
    c.font      = xfont(italic=True, sz=9, c=C_DARK_GREY)
    c.fill      = fill(C_LT_GREY)
    c.alignment = align(h="left")

    ws.merge_cells(f"{nxt}3:{last}3")
    c = ws[f"{nxt}3"]
    _name = get_display_name()
    c.value     = f"Prepared by:  {_name}" if _name else "Prepared by:  _________________________"
    c.font      = xfont(italic=True, sz=9, c=C_DARK_GREY)
    c.fill      = fill(C_LT_GREY)
    c.alignment = align(h="right")

    # ── Spacers ──────────────────────────────────────────────────────────────
    ws.row_dimensions[4].height = 5
    ws.row_dimensions[5].height = 5

def add_legend(ws, start_row, n_cols=12):
    last = get_column_letter(n_cols)
    ws.merge_cells(f"A{start_row}:{last}{start_row}")
    c = ws.cell(start_row, 1, "LEGEND")
    c.font      = xfont(bold=True, c=C_WHITE)
    c.fill      = fill(C_DARK_GREY)
    c.alignment = align()
    ws.row_dimensions[start_row].height = 18
    items = [
        (C_WHITE,      "White cell   —  manual input"),
        (C_LIGHT_BLUE, "Blue cell    —  auto-calculated formula (do not edit)"),
        (C_YELLOW,     "Yellow cell  —  required dropdown selection"),
        (C_SOFT_GREEN, "Green cell   —  dynamic calculator output (do not edit)"),
    ]
    for i, (colour, label) in enumerate(items):
        r = start_row + 1 + i
        ws.row_dimensions[r].height = 16
        ws.cell(r, 1).fill   = fill(colour)
        ws.cell(r, 1).border = xborder()
        ws.merge_cells(f"B{r}:{last}{r}")
        c2 = ws.cell(r, 2, label)
        c2.font      = xfont(sz=9)
        c2.fill      = fill(C_LT_GREY)
        c2.alignment = align(h="left")


# ══════════════════════════════════════════════════════════════════════════════
#  SHEET BUILDERS
# ══════════════════════════════════════════════════════════════════════════════
DATA_START = 7
DATA_ROWS  = 50

# Named ranges on Rates sheet that takeoff sheets reference
# These are the rows where bricks/blocks-per-m2 live for each type
# We store them as absolute references: 'Rates & Assumptions'!$D$row
RATES_SHEET = "Rates & Assumptions"


def build_brickwork(wb, job_name, brick_types, brick_rates_rows):
    """
    brick_rates_rows: list of Excel row numbers on the Rates sheet
                      where each brick type's bricks-per-m2 calc lives (col D)
    Columns (14 total):
      1=Desc  2=BrickType  3=LM  4=Height  5=M2  6=Ded  7=OpenCount
      8=OpenWidth  9=TotM2  10=#Brk  11=Wastage  12=Rate  13=Total  14=Notes
    """
    ws = wb.create_sheet("Brickwork")
    ws.sheet_properties.tabColor = C_AMBER
    ws.freeze_panes = "A7"

    n_cols = 14
    title_block(ws, job_name, "BRICKWORK TAKEOFF", n_cols)
    set_col_widths(ws, [30, 28, 7, 9, 8, 13, 10, 12, 10, 12, 12, 10, 12, 22])

    HEADERS = ["Description", "Brick Type", "LM", "Height\n(m)",
               "M2", "Deductions\nM2", "Opening\nCount", "Opening Width\n(LM)",
               "Total M2", "No. of\nBricks", "Wastage\n(+3%)",
               "Rate\n($ / brick)", "Total ($)", "Notes"]
    ws.row_dimensions[6].height = 46
    for ci, h in enumerate(HEADERS, 1):
        bg = C_AMBER if ci in (7, 8) else C_DARK_BLUE
        hdr_cell(ws, 6, ci, h, bg=bg)

    if brick_types:
        brick_ref = write_brick_list(wb, brick_types)
        dv = DataValidation(type="list", formula1=brick_ref, allow_blank=True,
                            showErrorMessage=False)
    else:
        dv = DataValidation(type="list", formula1='"Type"', allow_blank=True,
                            showErrorMessage=False)
    ws.add_data_validation(dv)
    dv.sqref = f"B{DATA_START}:B{DATA_START+DATA_ROWS-1}"

    for r in range(DATA_START, DATA_START + DATA_ROWS):
        bg = row_bg(r); rs = str(r)

        data_cell(ws, r,  1, bg=bg)                       # Description
        data_cell(ws, r,  2, bg=bg)                       # Brick Type
        data_cell(ws, r,  3, bg=bg, fmt="0.00")           # LM
        data_cell(ws, r,  4, bg=bg, fmt="0.00")           # Height

        # M2 = LM × Height
        c = data_cell(ws, r, 5, bg=C_LIGHT_BLUE, fmt="0.00")
        c.value = f'=IF(OR(C{rs}="",D{rs}=""),"",C{rs}*D{rs})'

        data_cell(ws, r,  6, bg=bg, fmt="0.00")           # Deductions M2
        data_cell(ws, r,  7, bg=bg, fmt="#,##0")          # Opening Count (NEW)
        data_cell(ws, r,  8, bg=bg, fmt="0.00")           # Opening Width LM (NEW)

        # Total M2 = M2 – Deductions
        c = data_cell(ws, r, 9, bg=C_LIGHT_BLUE, fmt="0.00")
        c.value = f'=IF(E{rs}="","",E{rs}-F{rs})'

        # No. of Bricks — looks up bricks/m2 from Rates sheet via MATCH on brick type
        c = data_cell(ws, r, 10, bg=C_LIGHT_BLUE, fmt="#,##0")
        c.value = (
            f'=IF(OR(I{rs}="",B{rs}=""),"",ROUND(I{rs}*'
            f'IFERROR(INDEX(\'{RATES_SHEET}\'!$E$3:$E$200,'
            f'MATCH(B{rs},\'{RATES_SHEET}\'!$A$3:$A$200,0)),50),0))'
        )

        # Wastage +3%
        c = data_cell(ws, r, 11, bg=C_LIGHT_BLUE, fmt="#,##0")
        c.value = f'=IF(J{rs}="","",ROUND(J{rs}*1.03,0))'

        data_cell(ws, r, 12, bg=bg, fmt="$#,##0.0000")   # Rate ($ per brick)

        # Total = Wastage qty × Rate per brick
        c = data_cell(ws, r, 13, bg=C_LIGHT_BLUE, fmt="$#,##0.00")
        c.value = f'=IF(K{rs}="","",K{rs}*L{rs})'

        data_cell(ws, r, 14, bg=bg)                       # Notes
        ws.row_dimensions[r].height = 18

    T = DATA_START + DATA_ROWS
    ws.row_dimensions[T].height = 22
    for ci in range(1, n_cols + 1):
        ws.cell(T, ci).fill = fill(C_DARK_BLUE)
    c = ws.cell(T, 1, "TOTAL")
    c.font = xfont(bold=True, c=C_WHITE, sz=11)
    c.alignment = align(h="right")
    for col, fmt in {7: "#,##0", 8: "0.00", 9: "0.00",
                     10: "#,##0", 11: "#,##0", 13: "$#,##0.00"}.items():
        cl = get_column_letter(col)
        c = ws.cell(T, col, f"=SUM({cl}{DATA_START}:{cl}{T-1})")
        c.font = xfont(bold=True, c=C_AMBER)
        c.fill = fill(C_DARK_BLUE)
        c.number_format = fmt

    ws.auto_filter.ref = f"A6:{get_column_letter(n_cols)}{T-1}"
    add_legend(ws, T + 2, n_cols)
    return ws


def build_blockwork(wb, job_name, block_data, block_rates_rows):
    """
    block_data: list of (name, size, corefill) tuples from wizard.
    Columns (20 total):
      1=Desc 2=BlkType 3=Size 4=LM 5=Ht 6=M2 7=Ded
      8=OpenCount 9=OpenWidth 10=TotM2 11=#Blk 12=Wast
      13=CF? 14=CorefM3 15=Steel 16=SBar 17=Cap 18=Rate 19=Total 20=Notes
    """
    block_types = [d[0] for d in block_data]
    ws = wb.create_sheet("Blockwork")
    ws.sheet_properties.tabColor = C_DARK_BLUE
    ws.freeze_panes = "A7"

    n_cols = 20
    title_block(ws, job_name, "BLOCKWORK TAKEOFF", n_cols)
    set_col_widths(ws, [26, 26, 10, 6, 9, 8, 13, 10, 12, 10, 10, 10,
                        11, 10, 10, 10, 14, 11, 12, 22])

    HEADERS = ["Description", "Block Type", "Block\nSize",
               "LM", "Height\n(m)", "M2", "Deductions\nM2",
               "Opening\nCount", "Opening Width\n(LM)",
               "Total M2", "No. of\nBlocks", "Wastage\n(+3%)",
               "Corefilled?\n(Yes / No)", "M3\nCorefill",
               "Steel\n(kg)", "Starter\nBars", "Capping\nBlocks",
               "Rate\n($ / block)", "Total ($)", "Notes"]
    ws.row_dimensions[6].height = 52
    for ci, h in enumerate(HEADERS, 1):
        if ci in (8, 9):
            bg = C_AMBER
        elif ci == 13:
            bg = C_MID_BLUE
        else:
            bg = C_DARK_BLUE
        hdr_cell(ws, 6, ci, h, bg=bg)

    if block_data:
        block_ref = write_block_lists(wb, block_data)
        dv_blk = DataValidation(type="list", formula1=block_ref,
                                allow_blank=True, showErrorMessage=False)
    else:
        dv_blk = DataValidation(type="list", formula1='"Type"',
                                allow_blank=True, showErrorMessage=False)
    dv_sz  = DataValidation(type="list",
                            formula1='"90mm,140mm,190mm,290mm"',
                            allow_blank=True, showErrorMessage=False)
    dv_cf  = DataValidation(type="list", formula1='"Yes,No"',
                            allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_blk)
    ws.add_data_validation(dv_sz)
    ws.add_data_validation(dv_cf)
    dv_blk.sqref = f"B{DATA_START}:B{DATA_START+DATA_ROWS-1}"
    dv_sz.sqref  = f"C{DATA_START}:C{DATA_START+DATA_ROWS-1}"
    dv_cf.sqref  = f"M{DATA_START}:M{DATA_START+DATA_ROWS-1}"

    for r in range(DATA_START, DATA_START + DATA_ROWS):
        bg = row_bg(r); rs = str(r)

        data_cell(ws, r,  1, bg=bg)                      # Description
        data_cell(ws, r,  2, bg=bg)                      # Block Type

        # Col 3: Block Size — auto-populated via VLOOKUP; manually overrideable
        c = data_cell(ws, r, 3, bg=C_LIGHT_BLUE)
        c.value = f'=IFERROR(VLOOKUP(B{rs},\'_ListsBlockwork\'!$A:$B,2,0),"")'
        c.alignment = align(h="center", wrap=False)

        data_cell(ws, r,  4, bg=bg, fmt="0.00")          # LM
        data_cell(ws, r,  5, bg=bg, fmt="0.00")          # Height

        # M2 = LM × Height
        c = data_cell(ws, r, 6, bg=C_LIGHT_BLUE, fmt="0.00")
        c.value = f'=IF(OR(D{rs}="",E{rs}=""),"",D{rs}*E{rs})'

        data_cell(ws, r,  7, bg=bg, fmt="0.00")          # Deductions M2
        data_cell(ws, r,  8, bg=bg, fmt="#,##0")         # Opening Count (NEW)
        data_cell(ws, r,  9, bg=bg, fmt="0.00")          # Opening Width LM (NEW)

        # Total M2 = M2 – Deductions
        c = data_cell(ws, r, 10, bg=C_LIGHT_BLUE, fmt="0.00")
        c.value = f'=IF(F{rs}="","",F{rs}-G{rs})'

        # No. of Blocks
        c = data_cell(ws, r, 11, bg=C_LIGHT_BLUE, fmt="#,##0")
        c.value = f'=IF(J{rs}="","",ROUND(J{rs}*12.5,0))'

        # Wastage +3%
        c = data_cell(ws, r, 12, bg=C_LIGHT_BLUE, fmt="#,##0")
        c.value = f'=IF(K{rs}="","",ROUND(K{rs}*1.03,0))'

        # Corefilled? — auto-populated from wizard; manually overrideable
        c = data_cell(ws, r, 13, bg=C_LIGHT_BLUE)
        c.value = f'=IFERROR(VLOOKUP(B{rs},\'_ListsBlockwork\'!$A:$C,3,0),"")'
        c.alignment = align(h="center", wrap=False)

        # M3 Corefill (references col 13=M for Corefilled, col 11=K for blocks)
        c = data_cell(ws, r, 14, bg=C_LIGHT_BLUE, fmt="0.00")
        c.value = (
            f'=IF(M{rs}<>"Yes","",'
            f'IF(C{rs}="290mm",ROUND(K{rs}/80,2),'
            f'IF(C{rs}="190mm",ROUND(K{rs}/110,2),'
            f'IF(C{rs}="140mm",ROUND(K{rs}/130,2),"N/A (90mm)"))))'
        )

        # Steel: 10 kg/m² corefilled (references col 13=M, col 10=J for Total M2)
        c = data_cell(ws, r, 15, bg=C_LIGHT_BLUE, fmt="#,##0.0")
        c.value = f'=IF(M{rs}<>"Yes","",IF(C{rs}="90mm","N/A",ROUND(J{rs}*10,1)))'

        # Starter bars: 2.5 per LM
        c = data_cell(ws, r, 16, bg=C_LIGHT_BLUE, fmt="#,##0")
        c.value = f'=IF(D{rs}="","",ROUND(D{rs}*2.5,0))'

        # Capping blocks: 2.5 per LM
        c = data_cell(ws, r, 17, bg=C_LIGHT_BLUE, fmt="#,##0.0")
        c.value = f'=IF(D{rs}="","",ROUND(D{rs}*2.5,1))'

        data_cell(ws, r, 18, bg=bg, fmt="$#,##0.0000")  # Rate ($ per block)

        # Total = Wastage qty × Rate per block (col 12=L, col 18=R)
        c = data_cell(ws, r, 19, bg=C_LIGHT_BLUE, fmt="$#,##0.00")
        c.value = f'=IF(L{rs}="","",L{rs}*R{rs})'

        data_cell(ws, r, 20, bg=bg)                      # Notes
        ws.row_dimensions[r].height = 18

    T = DATA_START + DATA_ROWS
    ws.row_dimensions[T].height = 22
    for ci in range(1, n_cols + 1):
        ws.cell(T, ci).fill = fill(C_DARK_BLUE)
    c = ws.cell(T, 1, "TOTAL")
    c.font = xfont(bold=True, c=C_WHITE, sz=11)
    c.alignment = align(h="right")
    sums = {8: "#,##0", 9: "0.00", 10: "0.00", 11: "#,##0", 12: "#,##0",
            14: "0.00", 15: "#,##0.0", 16: "#,##0",
            17: "#,##0.0", 19: "$#,##0.00"}
    for col, fmt in sums.items():
        cl = get_column_letter(col)
        c = ws.cell(T, col, f"=SUM({cl}{DATA_START}:{cl}{T-1})")
        c.font = xfont(bold=True, c=C_AMBER)
        c.fill = fill(C_DARK_BLUE)
        c.number_format = fmt

    ws.auto_filter.ref = f"A6:{get_column_letter(n_cols)}{T-1}"
    add_legend(ws, T + 2, n_cols)

    nr = T + 8
    ws.merge_cells(f"A{nr}:{get_column_letter(n_cols)}{nr}")
    note = ("NOTE — Corefill M3 & Steel (kg) only calculate when 'Corefilled?' = Yes.  "
            "90mm blocks cannot be corefilled (returns N/A).  "
            "Starter bars: 2.5 per LM regardless of corefill.  "
            "Capping blocks: 2.5 per LM assumes 400mm length — use 5.0 per LM if 200mm.  "
            "Block Size and Corefilled? are auto-populated from wizard selections — "
            "override per row by typing directly into those cells if needed.  "
            "Opening Count and Opening Width (LM) are for lintel reference only — "
            "they do not affect M2 or quantity calculations.")
    c = ws.cell(nr, 1, note)
    c.font      = xfont(sz=9, italic=True, c=C_DARK_GREY)
    c.fill      = fill("FFF9E6")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[nr].height = 56
    return ws


def build_demo(wb, job_name):
    ws = wb.create_sheet("Demo Takeoff")
    ws.sheet_properties.tabColor = C_MID_BLUE
    ws.freeze_panes = "A7"

    n_cols = 10
    title_block(ws, job_name, "DEMOLITION TAKEOFF", n_cols)
    set_col_widths(ws, [35, 8, 9, 8, 14, 10, 8, 9, 11, 28])

    HEADERS = ["Description", "LM", "Height\n(m)", "M2",
               "Deductions\nM2", "Total M2", "QTY",
               "Rate ($)", "Total ($)", "Notes"]
    ws.row_dimensions[6].height = 44
    for ci, h in enumerate(HEADERS, 1):
        hdr_cell(ws, 6, ci, h)

    for r in range(DATA_START, DATA_START + DATA_ROWS):
        bg = row_bg(r); rs = str(r)
        data_cell(ws, r, 1, bg=bg)
        data_cell(ws, r, 2, bg=bg, fmt="0.00")    # LM
        data_cell(ws, r, 3, bg=bg, fmt="0.00")    # Height

        # M2 = LM × Height
        c = data_cell(ws, r, 4, bg=C_LIGHT_BLUE, fmt="0.00")
        c.value = f'=IF(OR(B{rs}="",C{rs}=""),"",B{rs}*C{rs})'

        data_cell(ws, r, 5, bg=bg, fmt="0.00")    # Deductions M2

        # Total M2 = M2 – Deductions
        c = data_cell(ws, r, 6, bg=C_LIGHT_BLUE, fmt="0.00")
        c.value = f'=IF(D{rs}="","",D{rs}-E{rs})'

        data_cell(ws, r, 7, bg=bg, fmt="#,##0")   # QTY
        data_cell(ws, r, 8, bg=bg, fmt="$#,##0.00") # Rate

        # Total = QTY × Rate
        c = data_cell(ws, r, 9, bg=C_LIGHT_BLUE, fmt="$#,##0.00")
        c.value = f'=IF(G{rs}="","",G{rs}*H{rs})'

        data_cell(ws, r, 10, bg=bg)
        ws.row_dimensions[r].height = 18

    T = DATA_START + DATA_ROWS
    ws.row_dimensions[T].height = 22
    for ci in range(1, n_cols + 1):
        ws.cell(T, ci).fill = fill(C_DARK_BLUE)
    c = ws.cell(T, 1, "TOTAL")
    c.font = xfont(bold=True, c=C_WHITE, sz=11)
    c.alignment = align(h="right")
    for col, fmt in {6: "0.00", 9: "$#,##0.00"}.items():
        cl = get_column_letter(col)
        c = ws.cell(T, col, f"=SUM({cl}{DATA_START}:{cl}{T-1})")
        c.font = xfont(bold=True, c=C_AMBER)
        c.fill = fill(C_DARK_BLUE)
        c.number_format = fmt

    ws.auto_filter.ref = f"A6:{get_column_letter(n_cols)}{T-1}"
    add_legend(ws, T + 2, n_cols)
    return ws


def get_hidden_sheet(wb, name):
    """Return a hidden sheet by name, creating it if needed."""
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(name)
    ws.sheet_state = "hidden"
    return ws


def write_brick_list(wb, brick_types):
    """Write brick types to _ListsBrickwork col A. Returns range ref."""
    ws = get_hidden_sheet(wb, "_ListsBrickwork")
    for i, v in enumerate(brick_types, 1):
        ws.cell(i, 1, str(v))
    return f"'_ListsBrickwork'!$A$1:$A${len(brick_types)}"


def write_block_lists(wb, block_data):
    """
    Write block data to _ListsBlockwork.
    block_data: list of (name, size, corefill) tuples.
    Col A=name, Col B=size, Col C=corefill.
    Returns range ref for name column (for dropdown).
    """
    ws = get_hidden_sheet(wb, "_ListsBlockwork")
    for i, (name, size, corefill) in enumerate(block_data, 1):
        ws.cell(i, 1, str(name))
        ws.cell(i, 2, str(size))
        ws.cell(i, 3, str(corefill))
    return f"'_ListsBlockwork'!$A$1:$A${len(block_data)}"



def build_george_summary(wb, job_name, scope, brick_types, block_types):
    ws = wb.create_sheet("George Summary")
    ws.sheet_properties.tabColor = C_GREEN
    title_block(ws, job_name, "GEORGE SUMMARY", 14)
    set_col_widths(ws, [35, 11, 11, 11, 11, 11, 16, 11, 12, 16, 16, 11, 12, 20])

    r = 7
    brick_subtotal_row = None
    block_subtotal_row = None
    DS, DE = DATA_START, DATA_START + DATA_ROWS - 1

    # ── BRICKWORK ──────────────────────────────────────────────────────────
    if scope in ("brickwork", "both"):
        ws.merge_cells(f"A{r}:N{r}")
        c = ws.cell(r, 1, "BRICKWORK")
        c.font = Font(name=FONT_NAME, bold=True, size=13, color=C_DARK_BLUE)
        c.fill = fill(C_AMBER); c.alignment = align()
        ws.row_dimensions[r].height = 26; r += 1

        for ci, h in enumerate(["Brick Type", "Total M2", "No. of Bricks",
                                 "Wastage (+3%)", "Rate ($/brick)",
                                 "Total ($)", "Notes",
                                 "Total\nOpenings", "Opening Width\n(LM)",
                                 "M2 Check\n(Bricks÷50)", "Wastage Check\n(Blocks×1.03)"], 1):
            if ci in (10, 11):
                bg = "2E6B3E"
            elif ci in (8, 9):
                bg = C_AMBER
            else:
                bg = C_MID_BLUE
            hdr_cell(ws, r, ci, h, bg=bg)
        ws.row_dimensions[r].height = 30; r += 1

        sub_start = r
        for bt in brick_types:
            if not bt.strip(): continue
            ws.cell(r, 1, bt).border = xborder()
            ws.cell(r, 1).fill = fill(row_bg(r))
            ws.cell(r, 1).font = xfont()
            ws.cell(r, 1).alignment = align(h="left")
            # Brickwork cols after v1.2: I=TotM2(9), J=Bricks(10), K=Wastage(11)
            #                            G=OpenCount(7), H=OpenWidth(8)
            ws.cell(r, 2).value = f'=IFERROR(SUMIF(Brickwork!B{DS}:B{DE},A{r},Brickwork!I{DS}:I{DE}),0)'
            ws.cell(r, 3).value = f'=IFERROR(SUMIF(Brickwork!B{DS}:B{DE},A{r},Brickwork!J{DS}:J{DE}),0)'
            ws.cell(r, 4).value = f'=IFERROR(SUMIF(Brickwork!B{DS}:B{DE},A{r},Brickwork!K{DS}:K{DE}),0)'
            ws.cell(r, 6).value = f'=IF(E{r}="","",D{r}*E{r})'
            # Opening totals (cols 8 and 9 in Brickwork = G and H)
            ws.cell(r, 8).value = f'=IFERROR(SUMIF(Brickwork!B{DS}:B{DE},A{r},Brickwork!G{DS}:G{DE}),0)'
            ws.cell(r, 9).value = f'=IFERROR(SUMIF(Brickwork!B{DS}:B{DE},A{r},Brickwork!H{DS}:H{DE}),0)'
            fmts = {2:"0.00", 3:"#,##0", 4:"#,##0", 6:"$#,##0.00", 8:"#,##0", 9:"0.00"}
            for col, fmt in fmts.items():
                c = ws.cell(r, col)
                c.number_format = fmt
                c.fill = fill(row_bg(r)); c.border = xborder()
                c.font = xfont(); c.alignment = align(h="center", wrap=False)
            for col in [5, 7]:
                c = ws.cell(r, col)
                c.fill = fill(row_bg(r)); c.border = xborder(); c.font = xfont()
            # ── Sanity check cols 10 & 11 ─────────────────────────────────
            c = ws.cell(r, 10)
            c.value = (f'=IF(C{r}=0,"—",IF(ABS(ROUND(C{r}/50,2)-B{r})<=1,'
                       f'"✅","❌  "&TEXT(ABS(ROUND(C{r}/50,2)-B{r}),"0.00")&" m2 off"))')
            c.fill = fill(row_bg(r)); c.border = xborder()
            c.font = xfont(sz=9); c.alignment = align(h="center", wrap=False)
            c = ws.cell(r, 11)
            c.value = (f'=IF(D{r}=0,"—",IF(ABS(ROUND(D{r}/1.03,0)-C{r})<=1,'
                       f'"✅","❌  "&TEXT(ABS(ROUND(D{r}/1.03,0)-C{r}),"0")&" bricks off"))')
            c.fill = fill(row_bg(r)); c.border = xborder()
            c.font = xfont(sz=9); c.alignment = align(h="center", wrap=False)
            ws.row_dimensions[r].height = 18; r += 1

        sub_end = r - 1
        ws.merge_cells(f"A{r}:E{r}")
        c = ws.cell(r, 1, "Brickwork Subtotal")
        c.font = xfont(bold=True, c=C_WHITE); c.fill = fill(C_MID_BLUE)
        c.alignment = align(h="right")
        c = ws.cell(r, 6, f"=SUM(F{sub_start}:F{sub_end})")
        c.font = xfont(bold=True, c=C_WHITE); c.fill = fill(C_MID_BLUE)
        c.number_format = "$#,##0.00"
        for ci in range(2, 15): ws.cell(r, ci).fill = fill(C_MID_BLUE)
        ws.row_dimensions[r].height = 22
        brick_subtotal_row = r; r += 2

    # ── BLOCKWORK ──────────────────────────────────────────────────────────
    if scope in ("blockwork", "both"):
        ws.merge_cells(f"A{r}:N{r}")
        c = ws.cell(r, 1, "BLOCKWORK")
        c.font = Font(name=FONT_NAME, bold=True, size=13, color=C_AMBER)
        c.fill = fill(C_DARK_BLUE); c.alignment = align()
        ws.row_dimensions[r].height = 26; r += 1

        for ci, h in enumerate(["Block Type", "Total M2", "No. of Blocks",
                                 "Wastage (+3%)", "M3 Corefill", "Steel (kg)",
                                 "Starter Bars", "Rate ($/block)",
                                 "Total ($)", "Notes",
                                 "Total\nOpenings", "Opening Width\n(LM)",
                                 "M2 Check\n(Blocks÷12.5)", "Wastage Check\n(Blocks×1.03)"], 1):
            if ci in (13, 14):
                bg = "2E6B3E"
            elif ci in (11, 12):
                bg = C_AMBER
            else:
                bg = C_DARK_BLUE
            hdr_cell(ws, r, ci, h, bg=bg)
        ws.row_dimensions[r].height = 30; r += 1

        blk_sub_start = r
        for blk in block_types:
            if not blk.strip(): continue
            ws.cell(r, 1, blk).border = xborder()
            ws.cell(r, 1).fill = fill(row_bg(r))
            ws.cell(r, 1).font = xfont()
            ws.cell(r, 1).alignment = align(h="left")
            # Blockwork cols after v1.2:
            # J=TotM2(10), K=Blocks(11), L=Wastage(12),
            # N=CorefM3(14), O=Steel(15), P=StarterBars(16)
            # H=OpenCount(8), I=OpenWidth(9)
            ws.cell(r, 2).value = f'=IFERROR(SUMIF(Blockwork!B{DS}:B{DE},A{r},Blockwork!J{DS}:J{DE}),0)'
            ws.cell(r, 3).value = f'=IFERROR(SUMIF(Blockwork!B{DS}:B{DE},A{r},Blockwork!K{DS}:K{DE}),0)'
            ws.cell(r, 4).value = f'=IFERROR(SUMIF(Blockwork!B{DS}:B{DE},A{r},Blockwork!L{DS}:L{DE}),0)'
            ws.cell(r, 5).value = f'=IFERROR(SUMIF(Blockwork!B{DS}:B{DE},A{r},Blockwork!N{DS}:N{DE}),0)'
            ws.cell(r, 6).value = f'=IFERROR(SUMIF(Blockwork!B{DS}:B{DE},A{r},Blockwork!O{DS}:O{DE}),0)'
            ws.cell(r, 7).value = f'=IFERROR(SUMIF(Blockwork!B{DS}:B{DE},A{r},Blockwork!P{DS}:P{DE}),0)'
            ws.cell(r, 9).value = f'=IF(H{r}="","",D{r}*H{r})'
            # Opening totals (cols 8=H and 9=I in Blockwork)
            ws.cell(r, 11).value = f'=IFERROR(SUMIF(Blockwork!B{DS}:B{DE},A{r},Blockwork!H{DS}:H{DE}),0)'
            ws.cell(r, 12).value = f'=IFERROR(SUMIF(Blockwork!B{DS}:B{DE},A{r},Blockwork!I{DS}:I{DE}),0)'
            fmts = {2:"0.00", 3:"#,##0", 4:"#,##0", 5:"0.00",
                    6:"#,##0.0", 7:"#,##0", 9:"$#,##0.00", 11:"#,##0", 12:"0.00"}
            for col, fmt in fmts.items():
                c = ws.cell(r, col)
                c.number_format = fmt
                c.fill = fill(row_bg(r)); c.border = xborder()
                c.font = xfont(); c.alignment = align(h="center", wrap=False)
            for col in [8, 10]:
                c = ws.cell(r, col)
                c.fill = fill(row_bg(r)); c.border = xborder(); c.font = xfont()
            # ── Sanity check cols 13 & 14 ─────────────────────────────────
            c = ws.cell(r, 13)
            c.value = (f'=IF(C{r}=0,"—",IF(ABS(ROUND(C{r}/12.5,2)-B{r})<=1,'
                       f'"✅","❌  "&TEXT(ABS(ROUND(C{r}/12.5,2)-B{r}),"0.00")&" m2 off"))')
            c.fill = fill(row_bg(r)); c.border = xborder()
            c.font = xfont(sz=9); c.alignment = align(h="center", wrap=False)
            c = ws.cell(r, 14)
            c.value = (f'=IF(D{r}=0,"—",IF(ABS(ROUND(D{r}/1.03,0)-C{r})<=1,'
                       f'"✅","❌  "&TEXT(ABS(ROUND(D{r}/1.03,0)-C{r}),"0")&" blocks off"))')
            c.fill = fill(row_bg(r)); c.border = xborder()
            c.font = xfont(sz=9); c.alignment = align(h="center", wrap=False)
            ws.row_dimensions[r].height = 18; r += 1

        blk_sub_end = r - 1

        # ── Blockwork subtotal row ─────────────────────────────────────────
        ws.merge_cells(f"A{r}:H{r}")
        c = ws.cell(r, 1, "Blockwork Subtotal")
        c.font = xfont(bold=True, c=C_AMBER); c.fill = fill(C_DARK_BLUE)
        c.alignment = align(h="right")
        c = ws.cell(r, 9, f"=SUM(I{blk_sub_start}:I{blk_sub_end})")
        c.font = xfont(bold=True, c=C_AMBER); c.fill = fill(C_DARK_BLUE)
        c.number_format = "$#,##0.00"
        for ci in range(2, 15): ws.cell(r, ci).fill = fill(C_DARK_BLUE)
        ws.row_dimensions[r].height = 22
        block_subtotal_row = r; r += 1

        # ── Corefill subtotal ──────────────────────────────────────────────
        ws.merge_cells(f"A{r}:D{r}")
        c = ws.cell(r, 1, "Total Corefill")
        c.font = xfont(bold=True, c=C_WHITE); c.fill = fill(C_MID_BLUE)
        c.alignment = align(h="right")
        c = ws.cell(r, 5, f"=SUM(E{blk_sub_start}:E{blk_sub_end})")
        c.font = xfont(bold=True, c=C_WHITE); c.fill = fill(C_MID_BLUE)
        c.number_format = "0.00"
        c2 = ws.cell(r, 6, "M3")
        c2.font = xfont(italic=True, sz=9, c=C_LIGHT_BLUE); c2.fill = fill(C_MID_BLUE)
        c2.alignment = align(h="left")
        for ci in range(2, 15):
            ws.cell(r, ci).fill = fill(C_MID_BLUE)
        ws.row_dimensions[r].height = 18; r += 1

        # ── Steel subtotal ─────────────────────────────────────────────────
        ws.merge_cells(f"A{r}:E{r}")
        c = ws.cell(r, 1, "Total Steel")
        c.font = xfont(bold=True, c=C_WHITE); c.fill = fill(C_MID_BLUE)
        c.alignment = align(h="right")
        c = ws.cell(r, 6, f"=SUM(F{blk_sub_start}:F{blk_sub_end})")
        c.font = xfont(bold=True, c=C_WHITE); c.fill = fill(C_MID_BLUE)
        c.number_format = "#,##0.0"
        c2 = ws.cell(r, 7, "kg  (÷ 1000 = Tonnes)")
        c2.font = xfont(italic=True, sz=9, c=C_LIGHT_BLUE); c2.fill = fill(C_MID_BLUE)
        c2.alignment = align(h="left")
        for ci in range(2, 15):
            ws.cell(r, ci).fill = fill(C_MID_BLUE)
        ws.row_dimensions[r].height = 18; r += 1

        # ── Starter bars subtotal ──────────────────────────────────────────
        ws.merge_cells(f"A{r}:F{r}")
        c = ws.cell(r, 1, "Total Starter Bars")
        c.font = xfont(bold=True, c=C_WHITE); c.fill = fill(C_MID_BLUE)
        c.alignment = align(h="right")
        c = ws.cell(r, 7, f"=SUM(G{blk_sub_start}:G{blk_sub_end})")
        c.font = xfont(bold=True, c=C_WHITE); c.fill = fill(C_MID_BLUE)
        c.number_format = "#,##0"
        c2 = ws.cell(r, 8, "no.")
        c2.font = xfont(italic=True, sz=9, c=C_LIGHT_BLUE); c2.fill = fill(C_MID_BLUE)
        c2.alignment = align(h="left")
        for ci in range(2, 15):
            ws.cell(r, ci).fill = fill(C_MID_BLUE)
        ws.row_dimensions[r].height = 18; r += 2

    # ── Grand Total ────────────────────────────────────────────────────────
    if scope == "both" and brick_subtotal_row and block_subtotal_row:
        grand = f"=F{brick_subtotal_row}+I{block_subtotal_row}"
        grand_col = 9; label_end = 8
    elif scope == "brickwork" and brick_subtotal_row:
        grand = f"=F{brick_subtotal_row}"
        grand_col = 7; label_end = 6
    else:
        grand = f"=I{block_subtotal_row}"
        grand_col = 9; label_end = 8

    ws.merge_cells(f"A{r}:{get_column_letter(label_end)}{r}")
    c = ws.cell(r, 1, "GRAND TOTAL  (excl. GST)")
    c.font = Font(name=FONT_NAME, bold=True, size=12, color=C_AMBER)
    c.fill = fill(C_DARK_BLUE); c.alignment = align(h="right")
    c = ws.cell(r, grand_col, grand)
    c.font = Font(name=FONT_NAME, bold=True, size=13, color=C_AMBER)
    c.fill = fill(C_DARK_BLUE); c.number_format = "$#,##0.00"
    for ci in range(1, 15): ws.cell(r, ci).fill = fill(C_DARK_BLUE)
    ws.row_dimensions[r].height = 32
    return ws


def build_rates(wb, scope, brick_types, block_types):
    """
    Returns (ws, brick_rates_rows, block_rates_rows)
    brick_rates_rows: list of row numbers where each brick type's
                      bricks-per-m2 result lives (col D)
    """
    ws = wb.create_sheet(RATES_SHEET)
    ws.sheet_properties.tabColor = C_DARK_GREY
    title_block(ws, "For Internal Use Only — George & Estimator Eyes Only",
                "RATES & ASSUMPTIONS", 5)
    set_col_widths(ws, [42, 22, 14, 14, 16, 36])

    r = 7

    # ── DYNAMIC CALCULATOR ────────────────────────────────────────────────
    # Section heading
    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(r, 1, "BRICK CALCULATOR  —  Enter brick face dimensions to auto-calculate bricks per M2")
    c.font = Font(name=FONT_NAME, bold=True, size=11, color=C_WHITE)
    c.fill = fill(C_DARK_BLUE); c.alignment = align()
    ws.row_dimensions[r].height = 28; r += 1

    # Sub-heading explainer
    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(r, 1,
        "Enter brick face dimensions in mm — Length and Height (the two visible sides). "
        "Mortar joint adds to each face. "
        "Bricks per M2 = 1,000,000 ÷ ((Length + Joint) × (Height + Joint))  "
        "Blockwork is always 12.5 per M2 regardless of block width — see note below.")
    c.font      = xfont(sz=9, italic=True, c=C_DARK_GREY)
    c.fill      = fill("FFF9E6")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 36; r += 1

    # Mortar joint input row
    ws.merge_cells(f"A{r}:B{r}")
    c = ws.cell(r, 1, "Mortar joint thickness (mm)")
    c.font = xfont(bold=True, c=C_DARK_BLUE); c.fill = fill(C_LIGHT_BLUE)
    c.alignment = align(h="left"); c.border = xborder()
    mortar_cell = f"C{r}"
    mc = ws.cell(r, 3, 10)   # default 10mm joint
    mc.font = xfont(bold=True); mc.fill = fill(C_YELLOW)
    mc.alignment = align(h="center"); mc.border = xborder()
    mc.number_format = "0"
    ws.cell(r, 4).fill = fill(C_LT_GREY); ws.cell(r, 4).border = xborder()
    ws.merge_cells(f"E{r}:F{r}")
    note_c = ws.cell(r, 5, "Standard = 10mm  |  Fine joint = 3mm")
    note_c.font = xfont(sz=9, italic=True, c=C_DARK_GREY)
    note_c.fill = fill(C_LT_GREY); note_c.alignment = align(h="left")
    ws.row_dimensions[r].height = 20
    mortar_row = r; r += 1

    # Column headers for calculator table
    for ci, h in enumerate(["Type Name", "Length (mm)", "Width (mm)",
                             "Height (mm)", "Units per M2  ←  auto", "Notes"], 1):
        bg = C_SOFT_GREEN if ci == 5 else C_DARK_BLUE
        fg = C_DARK_BLUE if ci == 5 else C_WHITE
        hdr_cell(ws, r, ci, h, bg=bg, fg=fg)
    ws.row_dimensions[r].height = 26
    calc_hdr_row = r; r += 1

    brick_rates_rows  = []
    block_rates_rows  = []

    # ── Brick types rows ───────────────────────────────────────────────────
    if brick_types:
        ws.merge_cells(f"A{r}:F{r}")
        c = ws.cell(r, 1, "BRICKWORK")
        c.font = xfont(bold=True, c=C_DARK_BLUE)
        c.fill = fill(C_AMBER); c.alignment = align(h="left")
        ws.row_dimensions[r].height = 18; r += 1

        for bt in brick_types:
            if not bt.strip(): continue
            bg = row_bg(r)
            # Col A: type name (pre-filled, editable)
            c = ws.cell(r, 1, bt)
            c.font = xfont(); c.fill = fill(bg); c.border = xborder()
            c.alignment = align(h="left")
            # Col B: Length mm (default 230)
            c = ws.cell(r, 2, 230)
            c.font = xfont(); c.fill = fill(bg); c.border = xborder()
            c.number_format = "0"; c.alignment = align(h="center")
            # Col C: Width mm — bricks don't have a structural width like blocks
            # use 110mm (standard brick width) as reference only
            c = ws.cell(r, 3, 110)
            c.font = xfont(italic=True, c=C_DARK_GREY); c.fill = fill(bg); c.border = xborder()
            c.number_format = "0"; c.alignment = align(h="center")
            # Col D: Height mm (default 73mm → ~50/m2 with 10mm joint)
            c = ws.cell(r, 4, 73)
            c.font = xfont(); c.fill = fill(bg); c.border = xborder()
            c.number_format = "0"; c.alignment = align(h="center")
            # Col E: Units per M2 formula — uses Length(B) × Height(D) for face calc
            c = ws.cell(r, 5)
            c.value = (f'=IF(OR(B{r}=0,D{r}=0),"",ROUND('
                       f'1000000/((B{r}+${mortar_cell})*(D{r}+${mortar_cell})),1))')
            c.font = xfont(bold=True, c=C_DARK_BLUE)
            c.fill = fill(C_SOFT_GREEN); c.border = xborder()
            c.number_format = "0.0"; c.alignment = align(h="center")
            # Col F: Notes
            ws.cell(r, 6).fill = fill(bg); ws.cell(r, 6).border = xborder()
            ws.row_dimensions[r].height = 18
            brick_rates_rows.append(r)
            r += 1

    # ── Blockwork — fixed rate, no calculator needed ──────────────────────
    # Blocks per M2 is always 12.5 for ALL standard sizes (George's fixed rate)
    # Block width (90/140/190/290mm) only affects corefill rates, not units/m2
    if block_types:
        ws.merge_cells(f"A{r}:F{r}")
        c = ws.cell(r, 1,
            "BLOCKWORK  —  Fixed at 12.5 blocks per M2 for all standard sizes "
            "(90 / 140 / 190 / 290mm).  Block width affects corefill only, not units per M2.")
        c.font = xfont(sz=9, italic=True, c=C_DARK_GREY)
        c.fill = fill("FFF9E6")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 32; r += 1

    r += 1

    # ── RATES TABLE ───────────────────────────────────────────────────────
    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(r, 1, "RATES")
    c.font = Font(name=FONT_NAME, bold=True, size=11, color=C_WHITE)
    c.fill = fill(C_DARK_BLUE); c.alignment = align()
    ws.row_dimensions[r].height = 26; r += 1

    for ci, h in enumerate(["Item", "Unit", "Rate ($)", "Notes"], 1):
        hdr_cell(ws, r, ci, h)
    ws.row_dimensions[r].height = 26; r += 1

    rate_items = []
    if scope in ("brickwork", "both"):
        rate_items += [
            ("Brickwork – Supply & Lay",          "M2",        ""),
            ("Brickwork – Labour Only",            "M2",        ""),
            ("Brick – Mortar allowance",           "M2",        ""),
        ]
    if scope in ("blockwork", "both"):
        rate_items += [
            ("Blockwork – Supply & Lay (90mm)",       "M2",    ""),
            ("Blockwork – Supply & Lay (140mm)",      "M2",    ""),
            ("Blockwork – Supply & Lay (190mm)",      "M2",    ""),
            ("Blockwork – Supply & Lay (290mm)",      "M2",    ""),
            ("Corefill Concrete",                     "M3",    ""),
            ("N12 / N16 Steel (vert. & horiz.)",      "Tonne", ""),
            ("Starter Bars (incl. drilling & epoxy)", "each",  "Expensive — confirm with George"),
            ("Capping Blocks",                        "each",  ""),
        ]

    for item, unit, note in rate_items:
        bg = C_WHITE if r % 2 == 1 else C_LT_GREY
        data_cell(ws, r, 1, item, bg=bg)
        data_cell(ws, r, 2, unit, bg=bg, h="center")
        data_cell(ws, r, 3, bg=C_YELLOW, fmt="$#,##0.00")
        data_cell(ws, r, 4, note, bg=bg)
        ws.row_dimensions[r].height = 18; r += 1

    r += 1

    # ── ASSUMPTIONS ───────────────────────────────────────────────────────
    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(r, 1, "KEY ASSUMPTIONS  —  George's calculations (do not change)")
    c.font = xfont(bold=True, sz=11, c=C_DARK_BLUE)
    c.fill = fill(C_AMBER); c.alignment = align()
    ws.row_dimensions[r].height = 26; r += 1

    for ci, h in enumerate(["Assumption", "Value", "Notes"], 1):
        hdr_cell(ws, r, ci, h, bg=C_DARK_BLUE)
    ws.row_dimensions[r].height = 22; r += 1

    assumptions = [
        ("Standard bricks per M2 (230×73mm + 10mm joint)", "50",
         "Calculator above will override for non-standard sizes"),
        ("Standard blocks per M2 — ALL standard sizes (90/140/190/290mm)", "12.5 — FIXED",
         "Calculator above will override for non-standard sizes"),
        ("Wastage allowance",                               "3%",      "Applied to brick/block count"),
        ("Capping blocks per LM (400mm length)",            "2.5",     "If 200mm length: use 5.0 — check drawings"),
        ("Corefill — 290mm blocks per M3",                  "80 blks", "George"),
        ("Corefill — 190mm blocks per M3",                  "110 blks","George"),
        ("Corefill — 140mm blocks per M3",                  "~130 blks","George (approx)"),
        ("Corefill — 90mm blocks",                          "N/A",     "Cannot be corefilled"),
        ("Steel in corefilled blockwork",                   "10 kg/M2","N12 & N16 — changes if N20 or double bars"),
        ("Starter bars per LM",                             "2.5",     "Set at 400mm centres"),
    ]
    for assump, val, note in assumptions:
        bg = C_WHITE if r % 2 == 1 else C_LT_GREY
        data_cell(ws, r, 1, assump, bg=bg)
        c = data_cell(ws, r, 2, val, bg=bg, bold=True, h="center")
        data_cell(ws, r, 3, note, bg=bg)
        ws.row_dimensions[r].height = 18; r += 1

    return ws, brick_rates_rows, block_rates_rows


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL ORCHESTRATOR
# ══════════════════════════════════════════════════════════════════════════════
def build_excel(config):
    takeoff_type = config["type"]
    scope        = config.get("scope", "")
    job_name     = config["job_name"]
    brick_types  = config.get("brick_types", [])
    block_types  = config.get("block_types", [])
    save_folder  = config.get("save_folder",
                              os.path.join(os.path.expanduser("~"), "Downloads"))

    wb = Workbook()
    wb.remove(wb.active)

    if takeoff_type == "demo":
        build_demo(wb, job_name)
        type_label = "Demo"
    else:
        type_label = "Masonry"
        # Build rates sheet first so we get the row references back
        rates_ws, brick_rates_rows, block_rates_rows = build_rates(
            wb, scope, brick_types, block_types
        )
        # block_data: list of (name, size, corefill) tuples
        block_data = config.get("block_data", [])
        # Fallback: if block_data not set, build from block_types with empty defaults
        if not block_data and block_types:
            block_data = [(t, "", "") for t in block_types]

        if scope in ("brickwork", "both"):
            build_brickwork(wb, job_name, brick_types, brick_rates_rows)
        if scope in ("blockwork", "both"):
            build_blockwork(wb, job_name, block_data, block_rates_rows)
        build_george_summary(wb, job_name, scope, brick_types,
                             [d[0] for d in block_data])

        # Re-order sheets: Brickwork, Blockwork, George Summary, Rates
        desired = []
        if scope in ("brickwork", "both"):   desired.append("Brickwork")
        if scope in ("blockwork", "both"):   desired.append("Blockwork")
        desired += ["George Summary", RATES_SHEET]
        for i, name in enumerate(desired):
            while wb.sheetnames.index(name) != i:
                wb.move_sheet(name, offset=-1)
        # Keep list sheets hidden
        for hidden in ("_Lists", "_ListsBrickwork", "_ListsBlockwork"):
            if hidden in wb.sheetnames:
                wb[hidden].sheet_state = "hidden"

    os.makedirs(save_folder, exist_ok=True)
    filename  = f"{job_name} - {type_label} Takeoff.xlsx"
    save_path = os.path.join(save_folder, filename)
    base, ext = os.path.splitext(save_path)
    counter   = 1
    while os.path.exists(save_path):
        save_path = f"{base} ({counter}){ext}"
        counter  += 1

    wb.save(save_path)
    return save_path


# ══════════════════════════════════════════════════════════════════════════════
#  THEME  —  single unified professional theme for Sydney Fitout
#  Palette: near-black navy / off-white / warm slate / amber gold accent
#  Aesthetic: architectural, minimal, premium
# ══════════════════════════════════════════════════════════════════════════════

# Single unified theme — no more masonry/demo split
T = {
    # ── chrome / structural ─────────────────────────────────────
    "bg_dark":         "#0D1B2A",   # near-black navy — header, nav
    "bg_mid":          "#1A2E44",   # deep navy — secondary panels
    "bg_panel":        "#F7F8FA",   # off-white — content area
    "bg_card":         "#ECEEF2",   # cool light grey — summary cards
    "sep":             "#C8A951",   # warm gold — rule lines
    # ── text ────────────────────────────────────────────────────
    "text_light":      "#F7F8FA",   # near-white — on dark bg
    "text_dark":       "#0D1B2A",   # navy — on light bg
    "text_muted":      "#7A8A99",   # slate — secondary text
    "text_accent":     "#C8A951",   # gold — logo, highlights
    # ── interactive ─────────────────────────────────────────────
    "radio_bg":        "#ECEEF2",
    "radio_select":    "#C8A951",
    "entry_hl":        "#C8A951",
    "entry_hl_bg":     "#C8D0DA",
    # ── buttons ─────────────────────────────────────────────────
    "btn_back_bg":     "#1A2E44",
    "btn_back_fg":     "#7A8A99",
    "btn_next_bg":     "#1A2E44",
    "btn_next_fg":     "#F7F8FA",
    "btn_create_bg":   "#C8A951",
    "btn_create_fg":   "#0D1B2A",
    # ── misc ────────────────────────────────────────────────────
    "progress_bar":    "#C8A951",
    "content_bg":      "#F7F8FA",
    "summary_bg":      "#ECEEF2",
    "step_title_fg":   "#0D1B2A",
    "subtitle_fg":     "#7A8A99",
    "header_title_fg": "#C8A951",
    "bg_step_tag_bg":  "#C8A951",
    "bg_step_tag_fg":  "#0D1B2A",
}

# Keep THEMES dict for compatibility — all keys point to same theme
THEMES = {"masonry": T, "demo": T, "default": T}


# ══════════════════════════════════════════════════════════════════════════════
#  WIZARD APP
# ══════════════════════════════════════════════════════════════════════════════
class WizardApp(tk.Tk):
    W, H = 660, 760

    def __init__(self):
        super().__init__()
        self.title(f"Sydney Fitout Estimation and Project Management Takeoff Wizard  {VERSION}")
        self.resizable(False, False)
        self.protocol("WM_DELETE_WINDOW", self._on_close)

        self.config_data   = {}
        self.step          = 0
        self.steps         = []
        self.current_theme = T   # single unified professional theme

        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"{self.W}x{self.H}+{(sw-self.W)//2}+{(sh-self.H)//2}")

        self._build_chrome()
        self._show_step(0)

    def _build_chrome(self):
        t = self.current_theme

        # ── HEADER ── dark navy band with logo + title ───────────────────────
        self.hdr_frame = tk.Frame(self, bg=t["bg_dark"])
        self.hdr_frame.pack(fill="x")

        # Logo mark — gold "SF" monogram in a tight box
        logo_box = tk.Frame(self.hdr_frame, bg=t["bg_mid"],
                            width=56, height=56)
        logo_box.pack(side="left", padx=(20, 16), pady=16)
        logo_box.pack_propagate(False)
        tk.Label(logo_box, text="SF", font=("Georgia", 20, "bold"),
                 fg=t["text_accent"], bg=t["bg_mid"]
                 ).place(relx=0.5, rely=0.5, anchor="center")

        # Title stack
        title_stack = tk.Frame(self.hdr_frame, bg=t["bg_dark"])
        title_stack.pack(side="left", pady=16)
        tk.Label(title_stack, text="SYDNEY FITOUT",
                 font=("Arial", 15, "bold"),
                 fg=t["text_accent"], bg=t["bg_dark"],
                 anchor="w").pack(anchor="w")
        tk.Label(title_stack, text="ESTIMATION & PROJECT MANAGEMENT",
                 font=("Arial", 7),
                 fg=t["text_muted"], bg=t["bg_dark"],
                 anchor="w").pack(anchor="w")

        # Step indicator — right-aligned
        # Version label
        tk.Label(self.hdr_frame, text=VERSION,
                 font=("Arial", 8), fg=t["text_accent"],
                 bg=t["bg_dark"]).pack(side="right", padx=(0,6))
        self.lbl_step = tk.Label(self.hdr_frame, text="",
                                  font=("Arial", 9),
                                  fg=t["text_muted"], bg=t["bg_dark"])
        self.lbl_step.pack(side="right", padx=(24,4))

        # ── GOLD RULE ────────────────────────────────────────────────────────
        tk.Frame(self, bg=t["sep"], height=2).pack(fill="x")

        # ── STEP LABEL BAR ───────────────────────────────────────────────────
        self.step_bar = tk.Frame(self, bg=t["bg_panel"], height=36)
        self.step_bar.pack(fill="x")
        self.step_bar.pack_propagate(False)

        self.lbl_step_title = tk.Label(self.step_bar, text="",
                                        font=("Arial", 10, "bold"),
                                        fg=t["text_dark"], bg=t["bg_panel"],
                                        anchor="w")
        self.lbl_step_title.pack(side="left", padx=24, pady=8)

        # ── PROGRESS BAR (thin, gold) ─────────────────────────────────────────
        self.prog_track = tk.Frame(self, bg="#D8DCE3", height=3)
        self.prog_track.pack(fill="x")
        self.prog_bar = tk.Frame(self.prog_track, bg=t["progress_bar"], height=3)
        self.prog_bar.place(x=0, y=0, width=0, relheight=1)

        # ── CONTENT AREA ─────────────────────────────────────────────────────
        content_h = self.H - 56 - 88 - 2 - 36 - 3 - 72
        self.content = tk.Frame(self, bg=t["content_bg"], height=content_h)
        self.content.pack(fill="x")
        self.content.pack_propagate(False)

        # ── NAV BAR ──────────────────────────────────────────────────────────
        self.nav_frame = tk.Frame(self, bg=t["bg_dark"], height=72)
        self.nav_frame.pack(fill="x", side="bottom")
        self.nav_frame.pack_propagate(False)

        self.btn_back = tk.Button(self.nav_frame, text="← BACK",
                                   font=("Arial", 9, "bold"),
                                   bg=t["btn_back_bg"], fg=t["btn_back_fg"],
                                   relief="flat", padx=20, pady=10,
                                   cursor="hand2", bd=0,
                                   command=self._go_back)
        self.btn_back.pack(side="left", padx=20, pady=16)

        self.btn_next = tk.Button(self.nav_frame, text="NEXT →",
                                   font=("Arial", 9, "bold"),
                                   bg=t["bg_mid"], fg=t["text_light"],
                                   relief="flat", padx=20, pady=10,
                                   cursor="hand2", bd=0,
                                   command=self._go_next)
        self.btn_next.pack(side="right", padx=20, pady=16)

        self.btn_create = tk.Button(self.nav_frame, text="✓  CREATE TAKEOFF",
                                     font=("Arial", 10, "bold"),
                                     bg=t["btn_create_bg"], fg=t["btn_create_fg"],
                                     relief="flat", padx=24, pady=10,
                                     cursor="hand2", bd=0,
                                     command=self._create)
        self.btn_create.pack(side="right", padx=8, pady=16)

    def apply_theme(self, key):
        # Single unified theme — key ignored, always uses T
        pass

    def _update_nav(self):
        total   = len(self.steps)
        is_last = self.step == total - 1
        t = self.current_theme

        self.btn_back.config(
            state="normal" if self.step > 0 else "disabled",
            bg=t["btn_back_bg"] if self.step > 0 else "#0D1B2A",
            fg=t["btn_back_fg"] if self.step > 0 else "#2A3D52"
        )
        if is_last:
            self.btn_next.pack_forget()
            self.btn_create.pack(side="right", padx=8, pady=16)
        else:
            self.btn_create.pack_forget()
            self.btn_next.pack(side="right", padx=20, pady=16)

        self.lbl_step.config(text=f"{self.step + 1} / {total}")

        # Step title label in the step bar
        step_class  = self.steps[self.step]
        step_title  = getattr(step_class, "TITLE", "")
        self.lbl_step_title.config(text=step_title)

        pct = (self.step + 1) / total
        self.prog_track.update_idletasks()
        self.prog_bar.place(width=int(self.prog_track.winfo_width() * pct))

    def _show_step(self, idx):
        for w in self.content.winfo_children():
            w.destroy()
        self.steps = self._compute_steps()
        self.step  = idx
        frame = self.steps[idx](self.content, self.config_data, self)
        frame.pack(fill="both", expand=True)
        self.current_frame = frame
        self._update_nav()

    def _compute_steps(self):
        steps = [StepType]
        t  = self.config_data.get("type", "")
        sc = self.config_data.get("scope", "")
        if t == "masonry":
            steps.append(StepScope)
            if sc in ("brickwork", "both"):  steps.append(StepBrickTypes)
            if sc in ("blockwork", "both"):  steps.append(StepBlockTypes)
        steps.append(StepJobName)
        return steps

    def _go_next(self):
        if not self.current_frame.validate(): return
        self.current_frame.collect(self.config_data)
        self._show_step(self.step + 1)

    def _go_back(self):
        if self.step > 0:
            self._show_step(self.step - 1)

    def _create(self):
        if not self.current_frame.validate(): return
        self.current_frame.collect(self.config_data)
        try:
            path = build_excel(self.config_data)
            # Remember this save folder for next time
            save_config({"last_save_folder": self.config_data.get("save_folder", "")})
            messagebox.showinfo("Takeoff Created ✓",
                                f"Saved to:\n\n{path}\n\nOpening now…")
            try: os.startfile(path)
            except Exception: pass
            self.destroy()
        except Exception as e:
            messagebox.showerror("Error", f"Something went wrong:\n\n{e}")

    def _on_close(self):
        if messagebox.askokcancel("Quit", "Cancel and exit the wizard?"):
            self.destroy()


# ══════════════════════════════════════════════════════════════════════════════
#  BASE STEP
# ══════════════════════════════════════════════════════════════════════════════
class BaseStep(tk.Frame):
    TITLE    = ""
    SUBTITLE = ""
    N_TYPES  = 20   # number of type entry slots

    def __init__(self, parent, config, app):
        t = app.current_theme
        super().__init__(parent, bg=t["content_bg"])
        self.config_data = config
        self.app         = app
        self.theme       = t
        self._build_header()
        self._build_body()

    def _build_header(self):
        t = self.theme
        tk.Label(self, text=self.TITLE, font=("Arial", 14, "bold"),
                 fg=t["step_title_fg"], bg=t["content_bg"]
                 ).pack(pady=(18, 3), padx=30, anchor="w")
        if self.SUBTITLE:
            tk.Label(self, text=self.SUBTITLE, font=("Arial", 9),
                     fg=t["subtitle_fg"], bg=t["content_bg"],
                     wraplength=550, justify="left"
                     ).pack(padx=30, anchor="w", pady=(0, 8))
        tk.Frame(self, bg=t["sep"], height=2).pack(fill="x", padx=30, pady=(0, 12))

    def _build_body(self): pass
    def validate(self):    return True
    def collect(self, config): pass

    def _radio_group(self, parent, var, options):
        t = self.theme
        for text, value in options:
            tk.Radiobutton(parent, text=text, variable=var, value=value,
                           font=("Arial", 11), bg=t["radio_bg"], fg=t["text_dark"],
                           activebackground=t["radio_bg"],
                           selectcolor=t["radio_select"],
                           pady=7, padx=6, anchor="w"
                           ).pack(fill="x", padx=8, pady=2)

    def _make_entry(self, parent, font_size=11):
        t = self.theme
        return tk.Entry(parent, font=("Arial", font_size), relief="solid",
                        highlightthickness=1, highlightcolor=t["entry_hl"],
                        highlightbackground=t["entry_hl_bg"],
                        bg=t["content_bg"], fg=t["text_dark"],
                        insertbackground=t["text_dark"], bd=0)

    def _add_placeholder(self, entry, placeholder):
        t = self.theme
        entry.insert(0, placeholder)
        entry.config(fg=t["text_muted"])
        def _fi(ev, en=entry, ph=placeholder):
            if en.get() == ph: en.delete(0, tk.END); en.config(fg=t["text_dark"])
        def _fo(ev, en=entry, ph=placeholder):
            if en.get() == "": en.insert(0, ph); en.config(fg=t["text_muted"])
        entry.bind("<FocusIn>",  _fi)
        entry.bind("<FocusOut>", _fo)

    def _type_entry_grid(self, parent, existing, placeholders):
        """Scrollable two-column grid of N_TYPES entries."""
        t = self.theme

        # ── scrollable canvas container ───────────────────────────────────
        CANVAS_H = 380   # visible height in pixels
        canvas = tk.Canvas(parent, bg=t["content_bg"],
                           highlightthickness=0, height=CANVAS_H)
        scrollbar = tk.Scrollbar(parent, orient="vertical",
                                 command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        # inner frame inside the canvas
        inner = tk.Frame(canvas, bg=t["content_bg"])
        window_id = canvas.create_window((0, 0), window=inner, anchor="nw")

        # make inner frame fill canvas width
        def _on_canvas_resize(event):
            canvas.itemconfig(window_id, width=event.width)
        canvas.bind("<Configure>", _on_canvas_resize)

        # update scroll region when inner frame changes size
        def _on_frame_resize(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        inner.bind("<Configure>", _on_frame_resize)

        # mousewheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # ── two-column grid ───────────────────────────────────────────────
        inner.columnconfigure(0, weight=1)
        inner.columnconfigure(1, weight=1)

        entries = []
        for i in range(self.N_TYPES):
            col_frame = tk.Frame(inner, bg=t["content_bg"])
            col_frame.grid(row=i % 10, column=i // 10, sticky="ew",
                           padx=(4 if i // 10 == 1 else 0, 4), pady=3)
            col_frame.columnconfigure(1, weight=1)

            tk.Label(col_frame, text=f"{i+1}.", font=("Arial", 10),
                     fg=t["subtitle_fg"], bg=t["content_bg"],
                     width=3, anchor="e").pack(side="left")

            e = self._make_entry(col_frame)
            e.pack(side="left", fill="x", expand=True, ipady=5, padx=(4, 0))

            ph = placeholders[i] if i < len(placeholders) else f"Type {i+1}"
            val = existing[i] if i < len(existing) else ""
            if val:
                e.insert(0, val); e.config(fg=t["text_dark"])
            else:
                self._add_placeholder(e, ph)
            entries.append((e, ph))

        return entries


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 1 — TYPE
# ══════════════════════════════════════════════════════════════════════════════
class StepType(BaseStep):
    TITLE    = "What type of takeoff is this?"
    SUBTITLE = "Your selection themes the wizard and determines which sheets are created."

    def _build_body(self):
        t = self.theme
        self.var = tk.StringVar(value=self.config_data.get("type", "masonry"))
        box = tk.Frame(self, bg=t["radio_bg"])
        box.pack(padx=30, fill="x", pady=4)
        self._radio_group(box, self.var, [
            ("Masonry Takeoff  —  brickwork, blockwork, or both", "masonry"),
            ("Demolition Takeoff",                                 "demo"),
        ])

    def collect(self, config):
        config["type"] = self.var.get()
        if config["type"] == "demo":
            config.pop("scope", None)
            config.pop("brick_types", None)
            config.pop("block_types", None)


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 2 — SCOPE
# ══════════════════════════════════════════════════════════════════════════════
class StepScope(BaseStep):
    TITLE    = "What is the masonry scope?"
    SUBTITLE = "Determines which worksheets are created in your takeoff."

    def _build_body(self):
        t = self.theme
        self.var = tk.StringVar(value=self.config_data.get("scope", "brickwork"))
        box = tk.Frame(self, bg=t["radio_bg"])
        box.pack(padx=30, fill="x", pady=4)
        self._radio_group(box, self.var, [
            ("Brickwork only",             "brickwork"),
            ("Blockwork only",             "blockwork"),
            ("Both brickwork & blockwork", "both"),
        ])

    def collect(self, config):
        config["scope"] = self.var.get()


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 3a — BRICK TYPES  (20 slots)
# ══════════════════════════════════════════════════════════════════════════════
class StepBrickTypes(BaseStep):
    TITLE    = "What brick types are on this job?"
    SUBTITLE = ("Up to 20 types — include size and finish in one field "
                "(e.g. '230mm Standard Brick - Painted'). "
                "These become your dropdowns and flow into the George Summary. "
                "Dimensions can be refined on the Rates sheet after creation.")

    PLACEHOLDERS = [
        "e.g. 230mm Standard Brick",
        "e.g. 230mm Standard Brick - Painted",
        "e.g. 230mm Recycled Brick",
        "e.g. Brick Veneer",
        "e.g. Double Skin Brick",
        "e.g. 50mm Brick",
        "e.g. 50mm Brick - Painted",
        "e.g. Face Brick",
        "e.g. Engineer Brick",
        "e.g. Clinker Brick",
        "e.g. 230mm Standard Brick - Bagged",
        "e.g. 230mm Standard Brick - Rendered",
        "e.g. 110mm Brick",
        "e.g. Glazed Brick",
        "e.g. Perforated Brick",
        "e.g. Split Face Brick",
        "e.g. Pressed Brick",
        "e.g. Handmade Brick",
        "e.g. Limestone Block",
        "e.g. Custom Brick Type",
    ]

    def _build_body(self):
        existing = self.config_data.get("brick_types", [])
        # hint above the scrollable grid
        tk.Label(self, text="Leave blank to skip — only filled entries appear in dropdowns.",
                 font=("Arial", 9), fg=self.theme["text_muted"],
                 bg=self.theme["content_bg"], wraplength=580, justify="left"
                 ).pack(padx=30, pady=(0, 6), anchor="w")
        container = tk.Frame(self, bg=self.theme["content_bg"])
        container.pack(padx=30, fill="both", expand=True)
        self.entries = self._type_entry_grid(container, existing, self.PLACEHOLDERS)

    def validate(self):
        vals = [e.get().strip() for e, ph in self.entries
                if e.get().strip() not in ("", ph)]
        if not vals:
            messagebox.showwarning("No brick types entered",
                                   "Please enter at least one brick type.")
            return False
        return True

    def collect(self, config):
        config["brick_types"] = [
            e.get().strip() for e, ph in self.entries
            if e.get().strip() not in ("", ph)
        ]


# ══════════════════════════════════════════════════════════════════════════════
#  STEP 3b — BLOCK TYPES  (20 slots)
# ══════════════════════════════════════════════════════════════════════════════
class StepBlockTypes(BaseStep):
    TITLE    = "What block types are on this job?"
    SUBTITLE = ("Up to 20 types. Enter the name, select block size and whether it's corefilled. "
                "Size and Corefilled? will auto-populate in the spreadsheet — "
                "you can still override them per row if needed.")

    N_TYPES = 20
    SIZES   = ["", "90mm", "140mm", "190mm", "290mm"]
    PNAMES  = [
        "e.g. 190mm Besser Block", "e.g. 290mm Besser Block",
        "e.g. 140mm Besser Block", "e.g. 90mm Besser Block",
        "e.g. 190mm Besser Block - Painted", "e.g. 190mm Besser Block - Bagged",
        "e.g. 190mm Besser Block - Rendered", "e.g. 290mm Besser Block - Painted",
        "e.g. 290mm Besser Block - Bagged", "e.g. 140mm Besser Block - Painted",
        "e.g. 90mm Besser Block - Painted", "e.g. 190mm Feature Block",
        "e.g. 190mm H Block", "e.g. 290mm H Block",
        "e.g. 190mm Lintel Block", "e.g. 290mm Lintel Block",
        "e.g. 140mm Feature Block", "e.g. 190mm Besser Block - Rendered",
        "e.g. 290mm Besser Block - Rendered", "e.g. Custom Block Type",
    ]

    def _guess_size(self, name):
        """Auto-detect size from type name for convenience."""
        for sz in ["290", "190", "140", "90"]:
            if sz in name:
                return sz + "mm"
        return ""

    def _build_body(self):
        t = self.theme
        existing_data = self.config_data.get("block_data", [])

        tk.Label(self,
                 text="Leave blank to skip. Size and Corefilled? auto-populate in the spreadsheet — override per row if needed.",
                 font=("Arial", 9), fg=t["text_muted"], bg=t["content_bg"],
                 wraplength=600, justify="left"
                 ).pack(padx=24, pady=(0, 4), anchor="w")

        # Column headers
        hdr_row = tk.Frame(self, bg=t["bg_mid"])
        hdr_row.pack(padx=24, fill="x", pady=(0, 2))
        tk.Label(hdr_row, text="#",      width=3,  font=("Arial", 8, "bold"),
                 fg=t["text_light"], bg=t["bg_mid"], anchor="e").pack(side="left", padx=(4,4))
        tk.Label(hdr_row, text="Block Type Name", font=("Arial", 8, "bold"),
                 fg=t["text_light"], bg=t["bg_mid"], anchor="w").pack(side="left", fill="x", expand=True)
        tk.Label(hdr_row, text="Size",   width=9,  font=("Arial", 8, "bold"),
                 fg=t["text_light"], bg=t["bg_mid"], anchor="c").pack(side="left", padx=(4,0))
        tk.Label(hdr_row, text="Corefill?", width=9, font=("Arial", 8, "bold"),
                 fg=t["text_light"], bg=t["bg_mid"], anchor="c").pack(side="left", padx=(4,4))

        # Scrollable canvas
        canvas = tk.Canvas(self, bg=t["content_bg"], highlightthickness=0, height=400)
        sb = tk.Scrollbar(self, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y", padx=(0,4))
        canvas.pack(padx=(24,0), fill="both", expand=True)

        inner = tk.Frame(canvas, bg=t["content_bg"])
        wid = canvas.create_window((0,0), window=inner, anchor="nw")
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(wid, width=e.width))
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        self.block_rows = []  # list of (name_entry, size_var, corefill_var)

        for i in range(self.N_TYPES):
            ex_name    = existing_data[i][0] if i < len(existing_data) else ""
            ex_size    = existing_data[i][1] if i < len(existing_data) else ""
            ex_corefill= existing_data[i][2] if i < len(existing_data) else ""
            ph = self.PNAMES[i] if i < len(self.PNAMES) else f"Block type {i+1}"

            row = tk.Frame(inner, bg=t["content_bg"])
            row.pack(fill="x", pady=2, padx=2)

            # Row number
            tk.Label(row, text=f"{i+1}.", width=3, font=("Arial", 9),
                     fg=t["subtitle_fg"], bg=t["content_bg"], anchor="e").pack(side="left", padx=(0,4))

            # Name entry
            e = self._make_entry(row, font_size=10)
            e.pack(side="left", fill="x", expand=True, ipady=4)
            if ex_name:
                e.insert(0, ex_name); e.config(fg=t["text_dark"])
            else:
                self._add_placeholder(e, ph)

            # Size dropdown
            size_var = tk.StringVar(value=ex_size)
            size_om  = tk.OptionMenu(row, size_var, *self.SIZES)
            size_om.config(width=6, font=("Arial", 9),
                           bg=t["bg_mid"], fg=t["text_light"],
                           activebackground=t["sep"], relief="flat",
                           highlightthickness=0)
            size_om["menu"].config(bg=t["bg_mid"], fg=t["text_light"],
                                   font=("Arial", 9))
            size_om.pack(side="left", padx=(6,0))

            # Auto-detect size when name is typed (if size not already set)
            def _auto_size(event, entry=e, svar=size_var, ph_=ph):
                name = entry.get().strip()
                if name and name != ph_ and not svar.get():
                    svar.set(self._guess_size(name))
            e.bind("<FocusOut>", _auto_size)

            # Corefill dropdown
            cf_var = tk.StringVar(value=ex_corefill)
            cf_om  = tk.OptionMenu(row, cf_var, "", "Yes", "No")
            cf_om.config(width=5, font=("Arial", 9),
                         bg=t["bg_mid"], fg=t["text_light"],
                         activebackground=t["sep"], relief="flat",
                         highlightthickness=0)
            cf_om["menu"].config(bg=t["bg_mid"], fg=t["text_light"],
                                 font=("Arial", 9))
            cf_om.pack(side="left", padx=(4,0))

            self.block_rows.append((e, size_var, cf_var, ph))

    def validate(self):
        vals = [e.get().strip() for e, sv, cv, ph in self.block_rows
                if e.get().strip() not in ("", ph)]
        if not vals:
            messagebox.showwarning("No block types entered",
                                   "Please enter at least one block type.")
            return False
        return True

    def collect(self, config):
        block_data  = []
        block_types = []
        for e, size_var, cf_var, ph in self.block_rows:
            name = e.get().strip()
            if name and name not in ("", ph):
                block_data.append((name, size_var.get(), cf_var.get()))
                block_types.append(name)
        config["block_data"]  = block_data
        config["block_types"] = block_types


# ══════════════════════════════════════════════════════════════════════════════
#  STEP FINAL — JOB NAME
# ══════════════════════════════════════════════════════════════════════════════
class StepJobName(BaseStep):
    TITLE    = "What is the job name?"
    SUBTITLE = "Used as the sheet title and file name. Saves to your chosen folder."

    def _build_body(self):
        t     = self.theme
        inner = tk.Frame(self, bg=t["content_bg"])
        inner.pack(padx=30, fill="x")

        tk.Label(inner, text="Job name", font=("Arial", 10, "bold"),
                 fg=t["step_title_fg"], bg=t["content_bg"]
                 ).pack(anchor="w", pady=(0, 5))

        self.job_entry = self._make_entry(inner, font_size=13)
        self.job_entry.config(highlightthickness=2)
        self.job_entry.pack(fill="x", ipady=9)
        existing = self.config_data.get("job_name", "")
        if existing:
            self.job_entry.insert(0, existing)
        self.job_entry.focus_set()

        self.preview_lbl = tk.Label(inner, text="", font=("Arial", 9, "italic"),
                                     fg=t["text_muted"], bg=t["content_bg"], anchor="w")
        self.preview_lbl.pack(anchor="w", pady=(5, 0))

        # save folder
        tk.Label(inner, text="\nSave location", font=("Arial", 10, "bold"),
                 fg=t["step_title_fg"], bg=t["content_bg"]
                 ).pack(anchor="w", pady=(10, 4))

        folder_row = tk.Frame(inner, bg=t["content_bg"])
        folder_row.pack(fill="x")
        cfg_saved = load_config().get("last_save_folder", "")
        default = self.config_data.get("save_folder",
                   cfg_saved if cfg_saved and os.path.isdir(cfg_saved)
                   else os.path.join(os.path.expanduser("~"), "Downloads"))
        self.folder_var = tk.StringVar(value=default)
        fe = self._make_entry(folder_row, font_size=10)
        fe.config(textvariable=self.folder_var)
        fe.pack(side="left", fill="x", expand=True, ipady=5)

        def browse():
            from tkinter import filedialog
            d = filedialog.askdirectory(initialdir=self.folder_var.get())
            if d:
                self.folder_var.set(d)
                self._refresh()

        tk.Button(folder_row, text="Browse…", font=("Arial", 10),
                  bg=t["btn_back_bg"], fg=t["btn_back_fg"],
                  relief="flat", padx=10, cursor="hand2", command=browse
                  ).pack(side="left", padx=(8, 0), ipady=4)

        # summary
        tk.Label(inner, text="\nJob summary", font=("Arial", 10, "bold"),
                 fg=t["step_title_fg"], bg=t["content_bg"]
                 ).pack(anchor="w")

        self.summary_lbl = tk.Label(inner, text="", font=("Arial", 9),
                                     fg=t["text_dark"], bg=t["summary_bg"],
                                     wraplength=540, justify="left",
                                     anchor="nw", padx=12, pady=10)
        self.summary_lbl.pack(fill="x")

        self.job_entry.bind("<KeyRelease>", lambda e: self._refresh())
        self._refresh()

    def _refresh(self):
        job = self.job_entry.get().strip()
        typ = self.config_data.get("type", "masonry").capitalize()
        self.preview_lbl.config(
            text=f"→  Will save as:  {job} - {typ} Takeoff.xlsx" if job
            else "→  Enter a job name above"
        )
        cfg  = self.config_data
        sc   = cfg.get("scope", "—").capitalize()
        bt   = ", ".join(cfg.get("brick_types", [])) or "—"
        blkt = ", ".join(cfg.get("block_types", [])) or "—"
        lines = [f"Type:       {typ}"]
        if cfg.get("type") == "masonry":
            lines.append(f"Scope:      {sc}")
            if cfg.get("scope") in ("brickwork", "both"): lines.append(f"Bricks:     {bt}")
            if cfg.get("scope") in ("blockwork", "both"): lines.append(f"Blocks:     {blkt}")
        lines.append(f"Job name:   {job or '—'}")
        self.summary_lbl.config(text="\n".join(lines))

    def validate(self):
        if not self.job_entry.get().strip():
            messagebox.showwarning("Job name required",
                                   "Please enter a job name before creating the takeoff.")
            return False
        return True

    def collect(self, config):
        config["job_name"]    = self.job_entry.get().strip()
        config["save_folder"] = self.folder_var.get()


# ══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    check_for_updates()   # Silent version check — no-op if offline or up to date
    app = WizardApp()
    app.mainloop()
